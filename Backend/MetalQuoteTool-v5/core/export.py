import datetime
from tkinter import messagebox


# ═══════════════════════════════════════════════════════════════
#  §7  EXPORT (PDF + Excel)
# ═══════════════════════════════════════════════════════════════

def export_pdf(q, fp):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors as rc
        from reportlab.platypus import SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
    except ImportError:
        with open(fp.replace(".pdf",".txt"),"w") as f:
            f.write(f"MEPL QUOTATION — {q.name}\n")
            for l in q.lines: f.write(f"{l.desc}: {l.amt}\n")
            f.write(f"TOTAL: {q.total}\n")
        return fp.replace(".pdf",".txt")

    doc=SimpleDocTemplate(fp,pagesize=A4); st=getSampleStyleSheet(); el=[]
    el.append(Paragraph(f"<b>MEPL — {q.name}</b>",st["Title"]))
    el.append(Paragraph("<i>MEPL Sheet Metal Quote Tool</i>",st["Normal"]))
    el.append(Paragraph(f"Date: {datetime.datetime.now().strftime('%d-%b-%Y')}",st["Normal"]))
    el.append(Spacer(1,3*mm))
    el.append(Paragraph(
        f"Material: {q.mat} | {q.t}mm | ₹{q.rate_kg}/kg<br/>"
        f"Part: {q.pl}×{q.pw}mm | Wt: {q.weight}kg | Qty: {q.qty}",st["Normal"]))
    if q.flat_info: el.append(Paragraph(f"<i>{q.flat_info}</i>",st["Normal"]))
    el.append(Spacer(1,4*mm))
    d=[["Description","Qty","Unit","Rate (₹)","Amount (₹)"]]
    for l in q.lines: d.append([l.desc,f"{l.qty}",l.unit,f"{l.rate:,.2f}",f"{l.amt:,.2f}"])
    for lb,v in [("Subtotal",q.sub),("Overhead",q.overhead),("Profit",q.profit),
                 ("RATE/PC",q.per_pc),(f"TOTAL ({q.qty} pcs)",q.total)]:
        d.append(["","","",lb,f"{v:,.2f}"])
    tbl=Table(d,colWidths=[170,45,40,80,80])
    tbl.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),rc.HexColor('#143b62')),('TEXTCOLOR',(0,0),(-1,0),rc.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),
        ('GRID',(0,0),(-1,-1),0.5,rc.grey),('ALIGN',(1,0),(-1,-1),'RIGHT'),
        ('BACKGROUND',(0,-2),(-1,-1),rc.HexColor('#f0fdf4')),
        ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'),
    ]))
    el.append(tbl)
    if q.n:
        el.append(Spacer(1,5*mm))
        el.append(Paragraph(f"<b>NESTING:</b> {q.n.best} pcs/sheet ({q.n.orient}) | "
            f"Util: {q.n.util}% | {q.n.sheets} sheets",st["Normal"]))
    doc.build(el); return fp

def export_xlsx(q, fp):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
    except ImportError: messagebox.showinfo("Missing","pip install openpyxl"); return
    wb=Workbook(); ws=wb.active; ws.title="Quote"
    hf=Font(bold=True,color="FFFFFF",size=11)
    hfill=PatternFill(start_color="143B62",end_color="143B62",fill_type="solid")
    bd=Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
    ws.merge_cells('A1:E1')
    ws['A1']=f"MEPL — {q.name}"; ws['A1'].font=Font(bold=True,size=14)
    ws['A2']=f"{q.mat} | {q.t}mm | ₹{q.rate_kg}/kg | {q.pl}×{q.pw}mm | Qty:{q.qty}"
    r=4
    for c,h in enumerate(["Description","Qty","Unit","Rate","Amount"],1):
        cl=ws.cell(r,c,h); cl.font=hf; cl.fill=hfill; cl.border=bd
    for l in q.lines:
        r+=1
        for c,v in enumerate([l.desc,l.qty,l.unit,l.rate,l.amt],1): ws.cell(r,c,v).border=bd
    r+=1
    for lb,v in [("Subtotal",q.sub),("Overhead",q.overhead),("Profit",q.profit),
                 ("RATE/PC",q.per_pc),(f"TOTAL ({q.qty})",q.total)]:
        ws.cell(r,4,lb).font=Font(bold=True); ws.cell(r,5,v).font=Font(bold=True); r+=1
    for col in ws.columns:
        mx=max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width=min(mx+4,40)
    wb.save(fp)
