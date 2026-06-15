"""FAB sheet file grouper.

Takes a list of .dxf + .pdf files, parses each for metadata, and groups them
into the hierarchical structure of a Kumar Enterprises standard FAB sheet:

    MASTER (e.g., JVD714E)            — header row, no dims
      └── VARIANT (JVD714E-1)          — CRCA, 2mm, 116×590
      └── VARIANT (JVD714E-2)          — CRCA, 2mm, 106×630
    FLAT_PART (JV7428E)                — no variants, has dims inline

Grouping rule: strip trailing "-<digits>" from filename stem. All files that
share the same stripped root become variants of one master.
"""
from __future__ import annotations
import os
import re
from dataclasses import dataclass, field


_VARIANT_SUFFIX = re.compile(r"[_-](\d{1,3})$")
_SHEET_SUFFIX = re.compile(r"_SHT\d+$", re.IGNORECASE)
_REV_SUFFIX = re.compile(r"_[A-Z]{1,3}$")  # trailing _AA, _AB, _AC, _AH…
# Trailing date stamps like " 20-Aug-2018", "_27-NOV-2023", " 9-FEB-2026".
# Stripped BEFORE the sheet/rev suffix passes.
_DATE_SUFFIX = re.compile(
    r"[_\s]+\d{1,2}[-\s](?:JAN|FEB|MAR|APR|APRIL|MAY|JUN|JUNE|JUL|JULY|"
    r"AUG|SEP|SEPT|OCT|NOV|DEC|"
    r"Jan|Feb|Mar|Apr|April|May|Jun|June|Jul|July|Aug|Sep|Sept|Oct|Nov|Dec)"
    r"[-\s]\d{4}$",
    re.IGNORECASE)


def _stem(path: str) -> str:
    """Filename without extension, preserving dashes and case."""
    return os.path.splitext(os.path.basename(path))[0]


def _clean_stem(stem: str) -> str:
    """Strip trailing date stamp from a filename stem for display purposes.

    `JVD1005E 14-Aug-2020` → `JVD1005E`. Preserves case and internal
    separators (unlike _match_key which normalizes everything). This is the
    name the operator should see in the DRG.NO. column when no better
    value was extracted from the title block.
    """
    return _DATE_SUFFIX.sub("", stem).strip(" _-")


# Trailing revision / sheet suffixes. Ordered — sheet markers stripped first,
# then the various rev formats. Digit-only revs require 2-3 digits to avoid
# eating single-digit components of the drawing number itself (e.g., the
# '_3' in '896_0066_3').
_SUFFIX_PATTERNS = [
    re.compile(r"_SHT\d*[A-Z]*$", re.IGNORECASE),    # _SHT, _SHT1, _SHT1AC
    re.compile(r"_[A-Z]{1,3}\d{1,2}$", re.IGNORECASE),   # _BL1, _AC2
    re.compile(r"_\d{1,2}[A-Z]{1,3}$", re.IGNORECASE),   # _1AC (reversed)
    re.compile(r"_\d{2,3}$"),                         # _00, _01, _001
    re.compile(r"_[A-Z]{1,3}$", re.IGNORECASE),      # _AA, _AH, _ABC
]


def _match_key(stem: str) -> str:
    """Normalize a filename stem so DXF + PDF + Excel variants of the same
    drawing collapse to the same key, even when they use different
    separators or carry sheet-number / revision / date suffixes.

    Preserves single-digit trailing components (like the `_3` in
    `896_0066_3`) because those are part of the drawing number, not a rev.
    """
    # 1. Strip trailing date stamp.
    s = _DATE_SUFFIX.sub("", stem)
    # 2. Excise SHT markers anywhere in the string — they're sheet
    # numbers (e.g. '_SHT1', '0SHT_1AC'), never part of the drawing
    # number.
    s = re.sub(r"[_\-.]?SHT\d*[A-Z]*", "", s, flags=re.IGNORECASE)
    # 3. Peel revision / suffix tokens from the end until no more match.
    for _ in range(4):
        before = s
        for rx in _SUFFIX_PATTERNS:
            s = rx.sub("", s)
        if s == before:
            break
    # 4. Normalize remaining separators and case.
    s = re.sub(r"[-.\s]+", "_", s).upper()
    # 5. Collapse repeated underscores and strip trailing ones left
    # behind by the SHT excision.
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _split_master_variant(stem: str) -> tuple[str, str | None]:
    """Return (master_name, variant_suffix) — e.g. 'JVD714E_1' → ('JVD714E', '_1').
    Accepts both '-N' and '_N' separators (after key normalization).
    For a stem without a variant suffix, returns (stem, None)."""
    m = _VARIANT_SUFFIX.search(stem)
    if m:
        return stem[:m.start()], m.group(0)
    return stem, None


@dataclass
class FabVariant:
    name: str                              # e.g., "JVD714E-1" or "JV7428E"
    dxf_path: str | None = None
    pdf_path: str | None = None
    description: str = ""
    rev: str = ""
    material: str | None = None
    thickness: float | None = None
    length: float | None = None
    width: float | None = None
    qty: int = 1
    process: str = ""                      # BOLTED PRT, WELDING PROCESS, …
    sheet_name: str = "1220 × 2440  (4'×8')"  # key into STANDARD_SHEETS
    # Extra DXF-derived features kept so we can decide the process.
    n_bends: int = 0
    n_holes: int = 0
    weld_len_mm: float = 0
    # Process hints extracted by each file type's parser (not the final
    # answer — `process` above is the one shown in the table).
    process_hint_pdf: str = ""
    process_hint_dxf: str = ""
    missing: list[str] = field(default_factory=list)
    # Bounding box of this variant's drawing region inside the DXF
    # (x0, y0, x1, y1) in model-space mm. None for single-drawing files —
    # the viewer will then show the whole DXF. Used by the "View" column
    # to crop multi-drawing sheets to just this drawing.
    cluster_bbox: tuple | None = None

    @property
    def has_dxf(self) -> bool: return self.dxf_path is not None
    @property
    def has_pdf(self) -> bool: return self.pdf_path is not None

    @property
    def weight_kg(self) -> float | None:
        """Total lot weight in kg — matches the Ace-JV Cost Sheet formula:
            L(mm) × W(mm) × T(mm) × 7.86 (g/cm³) / 1_000_000 × QTY
        Returns None if any of L/W/T is missing."""
        if self.length is None or self.width is None or self.thickness is None:
            return None
        return round(self.length * self.width * self.thickness * 7.86
                     / 1_000_000 * self.qty, 3)

    def _geometric_capacity(self, sheet_l: float, sheet_w: float,
                             kerf: float = 2.0) -> int:
        """Max copies of this part that fit on one sheet (geometry only,
        no qty awareness). Considers both orientations; `kerf` is the cut
        allowance added to each part dim."""
        if (self.length is None or self.width is None
                or self.length <= 0 or self.width <= 0):
            return 0
        pl, pw = self.length + kerf, self.width + kerf
        a = int(sheet_l // pl) * int(sheet_w // pw)       # 0° orientation
        b = int(sheet_l // pw) * int(sheet_w // pl)       # 90° orientation
        return max(a, b, 0)

    def parts_per_sheet(self, sheet_l: float, sheet_w: float,
                        kerf: float = 2.0) -> int:
        """Parts actually going on a sheet for this order.

        Reactive to both L/W (via geometric capacity) AND QTY:
          • If QTY ≤ capacity, Parts/Sheet = QTY (your order uses just
            part of one sheet — e.g., need 10 pieces, 264 fit → 10 on one
            sheet).
          • If QTY > capacity, Parts/Sheet = capacity (sheet fills up).
            The overflow drives No. Sheets up.

        This way, changing QTY always visibly changes Parts/Sheet for
        small orders, and No. Sheets changes once QTY crosses the
        geometric capacity threshold.
        """
        cap = self._geometric_capacity(sheet_l, sheet_w, kerf)
        if cap <= 0 or self.qty <= 0:
            return 0
        return min(self.qty, cap)

    def sheets_needed(self, sheet_l: float, sheet_w: float,
                      kerf: float = 2.0) -> int:
        """Ceil(QTY / geometric capacity). Returns 0 if part doesn't fit."""
        cap = self._geometric_capacity(sheet_l, sheet_w, kerf)
        if cap <= 0:
            return 0
        import math
        return math.ceil(self.qty / cap)


@dataclass
class FabPart:
    """A master part with zero or more variants.

    When `variants` is empty, the master itself is the sole physical part
    (a 'flat' part in the sheet). When non-empty, the master is a header
    and `variants` are the actual fabricated pieces.
    """
    name: str                              # master name, e.g., "JVD714E"
    description: str = ""
    rev: str = ""
    variants: list[FabVariant] = field(default_factory=list)

    @property
    def is_assembly(self) -> bool: return len(self.variants) > 1 or \
        (len(self.variants) == 1 and self.variants[0].name != self.name)


def _dwg_to_dxf(dwg_path: str) -> str | None:
    """Convert a DWG file to DXF on disk and return the DXF path.

    Tries converters in this order:
      1. `dwg2dxf` from LibreDWG (preferred — free, open-source, scriptable)
      2. `ODAFileConverter` from Open Design Alliance (if user installed it)

    Returns the DXF path on success, or None if no converter is available
    or the conversion fails. The DXF is written alongside the DWG with the
    extension swapped, so repeated conversions are cheap (skip if target
    already exists and is newer than the source).
    """
    import os, shutil, subprocess
    dxf_path = os.path.splitext(dwg_path)[0] + ".dxf"
    # Already converted? Reuse.
    if (os.path.exists(dxf_path)
            and os.path.getmtime(dxf_path) >= os.path.getmtime(dwg_path)):
        return dxf_path

    # Try LibreDWG's dwg2dxf first.
    dwg2dxf = shutil.which("dwg2dxf")
    if dwg2dxf:
        try:
            subprocess.run(
                [dwg2dxf, "-o", dxf_path, dwg_path],
                check=True, capture_output=True, timeout=60)
            if os.path.exists(dxf_path):
                return dxf_path
        except Exception:
            pass

    # Fallback: ODA File Converter (GUI install, hard to automate cleanly).
    # The binary lives inside the .app bundle.
    oda = "/Applications/ODAFileConverter.app/Contents/MacOS/ODAFileConverter"
    if os.path.exists(oda):
        try:
            in_dir = os.path.dirname(dwg_path)
            out_dir = in_dir
            # ODA args: input_dir output_dir out_ver out_fmt recurse audit [filter]
            subprocess.run(
                [oda, in_dir, out_dir, "ACAD2018", "DXF", "0", "1",
                 os.path.basename(dwg_path)],
                check=False, capture_output=True, timeout=120)
            if os.path.exists(dxf_path):
                return dxf_path
        except Exception:
            pass

    return None


# Indian fabrication drawings use two common title-block shorthand formats:
#   A) '<T> THK. X <W> X <L> [Lg.]'    — explicit THK keyword
#         "2 THK. X 84 X 700 Lg. CRCA SHEET"
#         "3THKx70x148Lg HRS SHEET"
#         "3.15THK.x148x630 Lg. HRS. SHEET"
#   B) '<T> X <W> X <L> Lg [material]' — no THK, just Lg + material keyword
#         "2.5x50x109 Lg HRS SHEET"
#         "3X60X167 Lg HRS SHEET."
# Pattern B is more permissive so the MATERIAL-keyword OR 'Lg.' context is
# required to avoid false positives (e.g. a dimension string "500x300x2"
# for a box volume is NOT a thickness×width×length spec).
_TITLE_SPEC_THK_RE = re.compile(
    r"([0-9]+(?:\.[0-9]+)?)\s*(?:MM)?\s*THK\.?\s*X?\s*"
    r"([0-9]+(?:\.[0-9]+)?)\s*[Xx×]\s*"
    r"([0-9]+(?:\.[0-9]+)?)\s*(?:L[Gg]?\.?)?",
    re.IGNORECASE)
# Without THK. Anchored on the 'Lg' suffix which only appears on Indian
# fab-drawing material specs.
_TITLE_SPEC_NOTHK_RE = re.compile(
    r"\b([0-9]+(?:\.[0-9]+)?)\s*[Xx×]\s*"
    r"([0-9]+(?:\.[0-9]+)?)\s*[Xx×]\s*"
    r"([0-9]+(?:\.[0-9]+)?)\s*L[Gg]\.?",
    re.IGNORECASE)


def _parse_title_block_spec(text: str) -> dict:
    """Extract {thickness, width, length, material} from free-text like
    'MATERIAL :- 2 THK. X 84 X 700 Lg. CRCA SHEET'. Silent on no match."""
    out = {}
    if not text:
        return out

    # Try the stricter THK-anchored pattern first.
    m = _TITLE_SPEC_THK_RE.search(text) or _TITLE_SPEC_NOTHK_RE.search(text)
    if m:
        try:
            t = float(m.group(1)); w = float(m.group(2)); l = float(m.group(3))
            # Sanity: part dims 1–5000mm, thickness 0.3–12mm.
            if 0.3 <= t <= 12 and 1 <= w <= 5000 and 1 <= l <= 5000:
                out["thickness"] = t
                # Convention is usually "<W> X <L>" (width × length), so the
                # larger number is length, smaller is width.
                out["width"] = min(w, l)
                out["length"] = max(w, l)
        except ValueError:
            pass

    # Thickness-only fallback: title blocks on simple drawings often just
    # say "3 THK" or "3 MM THK" without any W × L dims. Use this only if
    # the main pattern above didn't already produce a thickness.
    if "thickness" not in out:
        m = re.search(
            r"\b([0-9]+(?:\.[0-9]+)?)\s*(?:MM)?\s*THK\b",
            text, re.IGNORECASE)
        if m:
            try:
                t = float(m.group(1))
                if 0.3 <= t <= 12:
                    out["thickness"] = t
            except ValueError:
                pass

    # "<T> MM Lg CRCA SHEET" / "<T>mm Lg HRS SHEET" — common on drawings
    # that omit the THK keyword entirely, anchoring on the material word
    # that follows "Lg".
    if "thickness" not in out:
        m = re.search(
            r"\b([0-9]+(?:\.[0-9]+)?)\s*MM\s+L[Gg]\.?\s+"
            r"(?:CRCA|HRS?|HRC|HR|MS|GI|C\.?R\.?C\.?A\.?|H\.?R\.?S?\.?|M\.?S\.?)",
            text, re.IGNORECASE)
        if m:
            try:
                t = float(m.group(1))
                if 0.3 <= t <= 12:
                    out["thickness"] = t
            except ValueError:
                pass
    # Material keyword anywhere in the text (reuse pdf_reader logic).
    try:
        from core.pdf_reader import _detect_material
        mat = _detect_material(text)
        if mat:
            out["material"] = mat
    except Exception:
        pass
    return out


# A "drawing anchor" is an MTEXT containing BOTH:
#   (a) a T×W×L spec matching _TITLE_SPEC_THK_RE or _TITLE_SPEC_NOTHK_RE, AND
#   (b) a material keyword (MATL, MATERIAL, CRCA, HRS, MS, HR SHEET, etc.)
# This keeps generic feature-dimension callouts (e.g., "2x10x15" on a small
# notch) from being mistaken for a title block.
_MATERIAL_ANCHOR_RE = re.compile(
    r"\bMAT[LE](?:RIAL)?\b"
    r"|\b(?:CRCA|HRS?|HRC|HR|MS|GI)\s*(?:SHEET)?\b"
    r"|\bM\.?S\.?\b"
    r"|\bC\.?R\.?C\.?A\.?\b"
    r"|\bH\.?R\.?S?\.?\b",
    re.IGNORECASE)


def _is_drawing_anchor(text: str) -> bool:
    if not (_TITLE_SPEC_THK_RE.search(text) or _TITLE_SPEC_NOTHK_RE.search(text)):
        return False
    return bool(_MATERIAL_ANCHOR_RE.search(text))


import math as _math


def _line_crosses_bbox(line, bbox, pad: float = 0.0) -> bool:
    """Does this line segment intersect or sit inside the (padded) bbox?"""
    (sx, sy), (ex, ey), _ll = line
    x0, y0, x1, y1 = bbox
    x0 -= pad; y0 -= pad; x1 += pad; y1 += pad
    # Either endpoint inside the bbox?
    if x0 <= sx <= x1 and y0 <= sy <= y1: return True
    if x0 <= ex <= x1 and y0 <= ey <= y1: return True
    # Segment straddles the bbox — parametric test on each of the 4 edges.
    dx = ex - sx; dy = ey - sy
    if abs(dx) < 1e-9 and abs(dy) < 1e-9:
        return False
    # Test intersection with each edge by clipping parameter t ∈ [0,1].
    tmin, tmax = 0.0, 1.0
    for p, q in ((-dx, sx - x0), (dx, x1 - sx), (-dy, sy - y0), (dy, y1 - sy)):
        if abs(p) < 1e-9:
            if q < 0: return False
            continue
        t = q / p
        if p < 0:
            if t > tmax: return False
            if t > tmin: tmin = t
        else:
            if t < tmin: return False
            if t < tmax: tmax = t
    return tmin <= tmax


def _is_cluster_cancelled(cluster_bbox, all_lines: list) -> bool:
    """Return True if the cluster's region is crossed out by an 'X' —
    TWO long lines with opposite diagonal slopes both passing through the
    cluster bbox.

    Operators mark superseded drawings with a big X covering the entire
    drawing frame. The X lines may start outside the cluster's MTEXT bbox
    (because the drawing frame is usually bigger than the title-block
    text region), so we check for line-bbox intersection + diagonal angle
    rather than "endpoints near corners".
    """
    if cluster_bbox is None:
        return False
    x0, y0, x1, y1 = cluster_bbox
    w = max(x1 - x0, 1.0); h = max(y1 - y0, 1.0)
    diag = (w * w + h * h) ** 0.5
    # A credible X-mark diagonal is at least 60% of the cluster diagonal AND
    # has an angle clearly different from horizontal/vertical. Pad bbox
    # slightly so lines whose endpoints land outside the MTEXT cluster
    # still register when they clip the region.
    pad = max(w, h) * 0.25

    has_pos_slope = False
    has_neg_slope = False

    for line in all_lines:
        (sx, sy), (ex, ey), ll = line
        if ll < diag * 0.6:
            continue
        if not _line_crosses_bbox(line, cluster_bbox, pad=pad):
            continue
        angle = _math.degrees(_math.atan2(ey - sy, ex - sx))
        if angle > 90: angle -= 180
        if angle < -90: angle += 180
        if abs(angle) < 15 or abs(angle) > 75:
            continue
        if angle > 0:
            has_pos_slope = True
        else:
            has_neg_slope = True
    # Require BOTH diagonal slopes for cancel — single-slash detection
    # caused too many false positives on drawings with isometric-view
    # construction lines. The dedup safety net catches cases where one
    # diagonal is short enough to miss.
    return has_pos_slope and has_neg_slope


def _cluster_drawings(mtext_chunks: list) -> list[dict]:
    """Cluster positioned MTEXT/TEXT chunks into per-drawing regions.

    Strategy: every drawing's title block carries a material spec like
    "MATL: 3THKx25x142 Lg M.S." — such chunks are DRAWING ANCHORS. For each
    anchor, collect all nearby chunks (Voronoi cell). Number of distinct
    anchors ≈ number of drawings on the sheet.

    Returns a single cluster for zero/one-anchor cases (single-drawing).
    """
    if len(mtext_chunks) < 3:
        return [{"bbox": None, "chunks": list(mtext_chunks)}]

    # Find anchors that pass BOTH criteria (THK spec + material keyword).
    anchors = [(x, y, t) for x, y, t in mtext_chunks if _is_drawing_anchor(t)]

    # Deduplicate anchors too close together (within 150 mm) — duplicated
    # material callouts / mirrored views on the same drawing sheet.
    deduped = []
    for ax, ay, at in anchors:
        skip = False
        for bx, by, _ in deduped:
            if (ax - bx) ** 2 + (ay - by) ** 2 < 150 * 150:
                skip = True
                break
        if not skip:
            deduped.append((ax, ay, at))
    anchors = deduped

    if len(anchors) < 2:
        return [{"bbox": None, "chunks": list(mtext_chunks)}]

    # Assign every chunk to its nearest anchor (Euclidean).
    clusters = [[] for _ in anchors]
    for x, y, t in mtext_chunks:
        best = 0; best_d = 10 ** 18
        for i, (ax, ay, _) in enumerate(anchors):
            d = (x - ax) ** 2 + (y - ay) ** 2
            if d < best_d:
                best_d = d; best = i
        clusters[best].append((x, y, t))

    out = []
    for cl in clusters:
        if not cl:
            continue
        cxs = [c[0] for c in cl]; cys = [c[1] for c in cl]
        out.append({
            "bbox": (min(cxs), min(cys), max(cxs), max(cys)),
            "chunks": cl,
        })
    out.sort(key=lambda r: (r["bbox"] or (0, 0, 0, 0))[0])
    return out


def _nearest_value(chunks: list, label_xy: tuple, max_dist: float = 300.0) -> str:
    """Return the text of the chunk whose insertion point is closest to
    `label_xy`, within `max_dist`. Used to pair a title-block label (e.g.
    "DRG. NO.") with the value MTEXT that sits beside it."""
    lx, ly = label_xy
    best = None; best_d = 10**9
    for x, y, t in chunks:
        dx = x - lx; dy = y - ly
        d = (dx * dx + dy * dy) ** 0.5
        if d < best_d and d <= max_dist and 0 < d:
            best_d = d; best = t
    return (best or "").strip()


# Labels searched for per cluster.
_DRGNO_LABEL_RE = re.compile(r"\bDR?G\.?\s*NO\.?\s*:?\s*$|\bDRAWING\s*NO\.?\s*:?\s*$",
                              re.IGNORECASE)
_REV_LABEL_RE = re.compile(r"\bREV\.?\s*N?O?\.?\s*:?\s*$|\bREVISION\s*:?\s*$",
                            re.IGNORECASE)
_TITLE_LABEL_RE = re.compile(
    r"\bTITLE\s*:?\s*$|\bDRG\.?\s*TITLE\s*:?\s*$|\bDESCRIPTION\s*:?\s*$",
    re.IGNORECASE)


_DRG_JUNK_RE = re.compile(
    r"^(NTS|XXX|TITLE|MATERIAL|DRG|REV|SCALE|SHEET|DATE|SIGN|NAME|"
    r"ITEM|CHECKED|BY|DESCRIPTION|ASSLY|ASSEMBLY|SPEC|NOTES?)$",
    re.IGNORECASE)


def _looks_like_drg_no(val: str) -> bool:
    """A valid drawing number has letters + digits, 3-30 chars, contains a
    digit, and isn't a generic label word."""
    val = (val or "").strip()
    if not (3 <= len(val) <= 30):
        return False
    if _DRG_JUNK_RE.match(val.strip(" .,:-")):
        return False
    if not re.search(r"\d", val):
        return False        # real drawing numbers always have a digit
    if not re.search(r"[A-Za-z]", val):
        return False        # and at least one letter
    # Allowable characters: letters, digits, - _ / .
    if not re.match(r"^[A-Za-z0-9][A-Za-z0-9\-_/.\s]{1,28}[A-Za-z0-9]$", val):
        return False
    return True


def _extract_drg_no(chunks: list) -> str:
    """Pull the drawing number value from the cluster's MTEXT, preferring
    values that actually look like real drawing numbers."""
    candidates: list[str] = []

    # Path A — inline "DRG. NO.: <value>" within one MTEXT
    for _x, _y, t in chunks:
        for m in re.finditer(
                r"\bDR?G\.?\s*NO\.?\s*:?\s*([A-Z0-9][A-Z0-9\-_/.\s]{2,30})\b",
                t, re.IGNORECASE):
            candidates.append(m.group(1).strip(" .,:-"))

    # Path B — label chunk adjacent to a value chunk
    for x, y, t in chunks:
        if _DRGNO_LABEL_RE.search(t.strip()):
            val = _nearest_value(chunks, (x, y)).strip(" .,:-")
            if val:
                candidates.append(val)

    for c in candidates:
        if _looks_like_drg_no(c):
            return c
    return ""


_REV_VALUE_RE_STRICT = re.compile(r"^(\d{1,3}|[A-Z]\d{0,2}|[A-Z]{2})$",
                                   re.IGNORECASE)
_REV_VALUE_JUNK_SET = {
    "NO", "REV", "DRG", "DATE", "SIGN", "NAME",
    "NTS", "ASSY", "CHD", "DRN", "APPD",
    "ZN", "ZINC", "GI", "HDG", "PT", "NIL",
    "NA", "TBD", "TBC", "MM", "KG", "CM",
    "SHT", "QTY", "MTL", "BY", "CHK",
    "ISED", "USED", "OTHER", "AS", "PER",
}


def _is_valid_rev(val: str) -> bool:
    if not val: return False
    if not _REV_VALUE_RE_STRICT.match(val): return False
    if val.upper() in _REV_VALUE_JUNK_SET: return False
    return True


def _extract_rev(chunks: list) -> str:
    """Pull the revision letters/digits from the cluster.

    Three paths:
      A) Inline "REV: X" in same chunk (most reliable)
      B) "REV." label chunk + spatially-adjacent value chunk (older
         drawings where the rev value is a separate MTEXT next to the
         label, e.g. JV7428E with "REV." at (1488,172) and "0" at
         (1497,171))
    """
    # A: inline. Only accept VALID rev formats — rejects 'ISED' captured
    # from 'REVISED' / 'AUTHORISED' etc. which the lax regex used to grab.
    for _x, _y, t in chunks:
        # Require the captured token to be at a word-boundary AFTER 'REV.'
        # — that way 'REVISED' (no dot/space after REV) is treated as one
        # word and the rev capture won't fire.
        m = re.search(
            r"\bREV(?:\.?\s+|:\s*|\.\s*N?O?\.?\s*:?\s*)([A-Z0-9]{1,4})\b",
            t, re.IGNORECASE)
        if m and _is_valid_rev(m.group(1).strip()):
            return m.group(1).strip()

    # B: spatial proximity — rev label chunk + nearest rev-shaped value
    rev_label_at = []
    for x, y, t in chunks:
        clean = t.strip(" .,:-|")
        if re.match(r"^REV(\.|ISION|\.NO\.?|\.NO)?\s*$", clean, re.IGNORECASE):
            rev_label_at.append((x, y))
    for lx, ly in rev_label_at:
        best = None; best_d = 35.0
        for x, y, t in chunks:
            if (x, y) == (lx, ly): continue
            v = t.strip(" .,:-|")
            if not _is_valid_rev(v):
                continue
            d = ((x - lx) ** 2 + (y - ly) ** 2) ** 0.5
            if d < best_d:
                best_d = d; best = v
        if best:
            return best

    return ""


_DESC_JUNK_RE = re.compile(
    r"^(QTY|ITEM|PART\s*NUMBER|DRG\.?\s*NO\.?|SCALE|DATE|"
    r"REV|SIGN|NAME|CHECKED|BY|NOTES?|SPEC|MATL|MATERIAL|"
    r"DESCRIPTION|TITLE|SHEET|FORMAT|W/O\s*NO|"
    r"ALL\s+DIMENSIONS?\s+IN\s+MM|DO\s+NOT\s+SCALE|DRAWING\s+NOT\s+TO\s+SCALE|"
    r"THIRD\s+ANGLE\s+PROJECTION|FIRST\s+ANGLE\s+PROJECTION|"
    r"ASSLY\.?\s*DRG\.?\s*NO\.?|A\s*S\s*PER|UNLESS\s+OTHERWISE|"
    # Boilerplate phrases that leaked into heading-style detection.
    r"BEND(ING)?\s+LINE|BEND(ING)?\s+DOWN|BEND(ING)?\s+UP|"
    r"USED\s+ON|APPR?O?VED|APPROVED\s+BY|DRAWN(\s+BY)?|"
    r"CROMPTON\s+GREAVES|SIEMENS|CGL|COMPANY|LTD|LIMITED|"
    r"PHOSPHATED(\s+AND\s+PAINTED)?|POWDER\s+COAT(ED|ING)?|ZINC\s+PLAT|"
    r"FINISH|SURFACE\s+FINISH|NATURAL|"
    r"REFERENCE|PROJECTION|CONFIDENTIAL|"
    r"HINGE\s+PIN|PIN)\.?$",
    re.IGNORECASE)


# Substring patterns — reject the description if it contains any of these
# anywhere, not only when it matches in full. Catches cases like
# "Crompton Greaves Ltd" where the junk word is followed by more text.
_DESC_JUNK_CONTAINS_RE = re.compile(
    r"\b("
    r"CROMPTON\s+GREAVES|GE\s+T\s*&\s*D|SIEMENS|ABB|CGL|"
    r"IF\s+IN\s+DOUBT|DO\s+NOT\s+SCALE|DRAWING\s+NOT\s+TO\s+SCALE|"
    r"PHOSPHATED|POWDER\s+COAT|ZINC\s+PLAT|HOT[\s\-]DIP|GALVANI[SZ]ED?|"
    r"SHADE\s+AS\s+PER|SHADED?\s+AS\s+PER|AS\s+PER\s+MI|"
    r"THIRD\s+ANGLE\s+PROJ|FIRST\s+ANGLE\s+PROJ|"
    r"ALL\s+DIMENSIONS?|UNLESS\s+OTHERWISE|CONFIDENTIAL|"
    r"BENDING?\s+(LINE|UP|DOWN)|ASSLY\.?\s*DRG|"
    r"COMPANY|COMPANIES|LIMITED|\sLTD\.?\b|"
    r"WEIGHT|TOLERANCE|NTS|TRUE\s+SCALE|"
    # Recently-leaked boilerplate ——————
    r"ALL\s+INFORMATION|INFORMATION\s+CONTAINED|WITHOUT\s+PRIOR|PRIOR\s+CONSENT|"
    r"CODE\s+N[OR]|CODE\s+NUMBER|"
    r"(?:FRONT|TOP|SIDE|REAR|BACK|BOTTOM|LEFT|RIGHT|DEVELOPMENT|EXPLODED)"
    r"\s+(?:VIEW|CIEW|PLAN)|VIEW\s+[A-Z]\-[A-Z]|"
    r"SECTION\s+[A-Z]|SECTIONAL\s+VIEW|DETAIL\s+[A-Z]|"
    r"SWITCHGEAR|DIVISION|DEPARTMENT|"
    r"REMARKS?|REFERENCE|REFERENCES|"
    r"GENERAL\s+TOLERANC|TOLERANCES?\s+SHALL|"
    r"DRAWN|APPROVED|AUTHORI[SZ]ED|RELEASED?|"
    r"OTHERWISE\s+SPECIFIED|NX\s+CALIX|NX\s+CAD|AUTOCAD|SOLIDWORKS|CREO|"
    # Assembly / fabrication notes that aren't part descriptions.
    r"NO\s+WELD(ING)?|WELD(\s+HERE)?|NO\s+HOLE|NO\s+DRILL|"
    r"CHAMFER|CHAM\s+\d|FILLET|DEBURR|REMOVE|ADD(ED)?|"
    r"BEFORE\s+BENDING|AFTER\s+BENDING|AFTER\s+WELD|"
    r"HAS\s+BEEN|CHANGED?\s+TO|WAS\s+\d|MM\s+BEFORE"
    r")\b",
    re.IGNORECASE)


# Single-word mechanical part names that are valid descriptions on their own.
_VALID_SINGLE_WORDS = {
    "BRACKET", "COVER", "PLATE", "PANEL", "ANGLE", "SHEET", "CHANNEL",
    "BEAM", "TUBE", "BAR", "ROD", "CLIP", "FRAME", "MOUNT", "SUPPORT",
    "HOUSING", "ENCLOSURE", "DOOR", "LID", "FLANGE", "GUSSET", "RIB",
    "WASHER", "SPACER", "SHIM", "HINGE", "LATCCH", "LATCH", "LOCK",
    "HOOK", "LUG", "STUD", "PIN", "SHAFT", "RING", "CAP", "BASE",
    "BODY", "CASE", "CAP", "CASING", "GRILLE", "LOUVER", "LOUVRE",
    "STIFFENER", "REINFORCEMENT", "SEAL", "GASKET", "BUSBAR", "BUS-BAR",
    "SPLITTER", "DIVIDER", "PARTITION", "SHUTTER",
}

# Single words that should NOT be accepted as descriptions.
_INVALID_SINGLE_WORDS_RE = re.compile(
    r"^(BENDING|BEND|FRONT|SIDE|TOP|BACK|REAR|BOTTOM|DETAIL|SECTION|"
    r"SECTIONAL|INDEX|REFERENCE|NOTES?|GENERAL|TITLE|DESCRIPTION|"
    r"ITEM|QTY|CHECKED|APPROVED|DRAWN|REV|SIGN|DATE|NAME|SCALE|"
    r"SHEET|FORMAT|WEIGHT|TOLERANCE|MATL|MATERIAL|SPEC|FINISH)\.?$",
    re.IGNORECASE)


def _looks_like_description(val: str, drg_no: str = "") -> bool:
    """A valid part description is 3-60 chars, has letters, isn't generic
    header text / boilerplate, isn't the drawing number, and either is
    multi-word OR is a known mechanical part name."""
    val = (val or "").strip(" .,:-|")
    if not (3 <= len(val) <= 60): return False
    if _DESC_JUNK_RE.match(val): return False
    if _DESC_JUNK_CONTAINS_RE.search(val): return False
    letters = sum(1 for c in val if c.isalpha())
    if letters < 4: return False
    if drg_no:
        dn = re.sub(r"[\s\-_./]+", "", drg_no).upper()
        vn = re.sub(r"[\s\-_./]+", "", val).upper()
        if dn == vn: return False
        if vn.startswith(dn) and len(vn) - len(dn) < 3: return False
    # Single-word case: accept only known mechanical part names.
    if " " not in val:
        if _INVALID_SINGLE_WORDS_RE.match(val): return False
        token = re.sub(r"[^A-Za-z]", "", val).upper()
        if token in _VALID_SINGLE_WORDS:
            return True
        # Mixed-case single words like "Bracket" or "HingePlate" are also
        # reasonable (CamelCase).
        if re.search(r"[A-Z][a-z]+", val):
            return True
        return False
    return True


def _extract_description(chunks: list, drg_no: str = "") -> str:
    """Pull a short part description from the cluster. The DRG.NO. for this
    cluster is passed in so we can reject values that simply echo it.

    Three paths, in order:
      A) Inline "DESCRIPTION: <value>" in a single chunk — most reliable
      B) Label chunk → spatially-nearest value chunk within 250 mm
      C) Heading heuristic — first multi-word all-letter phrase near the
         DRG.NO. label (catches title blocks like "ANGLE FOR BUSBAR"
         that have NO explicit DESCRIPTION label).
    """
    try:
        from core.pdf_reader import _strip_acad_formatting
    except Exception:
        _strip_acad_formatting = lambda s: s

    candidates: list[str] = []
    cut_re = re.compile(
        r"\b(?:MATL\s*:|MATERIAL\s*:|SPEC\s*:|THK\b|MM\s+THK\b|"
        r"SHEET\s*:|FINISH\s*:|SCALE\s*:|REV\s*:|ITEM\b)",
        re.IGNORECASE)

    # A) Inline DESCRIPTION/TITLE
    for _x, _y, t in chunks:
        clean = _strip_acad_formatting(t)
        for m in re.finditer(
                r"\b(?:DRG\.?\s*)?(?:TITLE|DESCRIPTION)\s*[:\-]\s*([^\n|]{3,80})",
                clean, re.IGNORECASE):
            raw = m.group(1)
            cut = cut_re.search(raw)
            if cut:
                raw = raw[:cut.start()]
            candidates.append(raw)

    # B) Label chunk → nearest value
    for x, y, t in chunks:
        clean = _strip_acad_formatting(t).strip(" .,:-|")
        if _TITLE_LABEL_RE.search(clean):
            val = _nearest_value(chunks, (x, y), max_dist=250.0)
            if val:
                candidates.append(val)

    # C) Heading heuristic — heading-style phrases near DRG.NO. label.
    # Useful for Indian drawings whose title block uses a standalone part
    # name (e.g. "ANGLE FOR BUSBAR") with no explicit label.
    drgno_pos = None
    for x, y, t in chunks:
        if _DRGNO_LABEL_RE.search(t.strip()):
            drgno_pos = (x, y)
            break
    for x, y, t in chunks:
        clean = _strip_acad_formatting(t).strip(" .,:-|")
        # Must be multi-word, all letters + spaces (no digits / special chars).
        if not re.match(r"^[A-Za-z][A-Za-z\s\-'/]{4,59}$", clean):
            continue
        if " " not in clean:
            continue
        # If we know where DRG.NO. sits, prefer phrases close to it.
        if drgno_pos is not None:
            dx = x - drgno_pos[0]; dy = y - drgno_pos[1]
            if (dx * dx + dy * dy) > 800 * 800:   # > 800 mm away — skip
                continue
        candidates.append(clean)

    for c in candidates:
        clean = c.strip(" .,:-|")
        if _looks_like_description(clean, drg_no):
            return clean
    return ""


def _extract_sheet_size(chunks: list) -> str:
    """Paper size like A3, A4, A2. Usually shown as 'A3', 'B/A3', 'FORMAT A3'."""
    for _x, _y, t in chunks:
        m = re.search(r"\b(?:FORMAT\s*:?\s*)?([AB]/?A[0-4])\b", t, re.IGNORECASE)
        if m:
            val = m.group(1).upper().replace("/", "")
            return val if val else ""
    return ""


def _filter_dims_in_bbox(dim_positioned: list, bbox) -> list:
    """Return DIMENSION measurements whose insertion point falls inside bbox."""
    if bbox is None:
        return [v for _x, _y, v in dim_positioned]
    x0, y0, x1, y1 = bbox
    # Pad the cluster bbox since drawings extend beyond their title-block text.
    pad = 500.0
    return [v for x, y, v in dim_positioned
            if (x0 - pad) <= x <= (x1 + pad) and (y0 - pad) <= y <= (y1 + pad)]


# Normalize an AutoCAD ATTRIB tag → canonical field. Different templates
# use different tag names; collapse the common variants.
_ATTR_TAG_CANONICAL = {
    # Description / part name
    "DESCRIPTION": "description",
    "PART_NAME": "description",
    "PARTNAME": "description",
    "PART_TITLE": "description",
    "TITLE": "description",
    "DRAWING_TITLE": "description",
    # Drawing number
    "DRAWING_NUMBER": "drg_no",
    "DRAWINGNUMBER": "drg_no",
    "DRG_NO": "drg_no",
    "DRG_NUMBER": "drg_no",
    "DWG_NO": "drg_no",
    "DWG_NUMBER": "drg_no",
    "DOC_NUMBER": "drg_no",
    "NUMBER": "drg_no",
    # Revision
    "REVISION": "rev",
    "REVISION_NUMBER": "rev",
    "REVISION_NO": "rev",
    "REV": "rev",
    "REV_NO": "rev",
    # Paper / sheet size
    "FORMAT": "paper_size",
    "PAPER_SIZE": "paper_size",
    "SHEET_SIZE": "paper_size",
    "SIZE": "paper_size",
    # Material (rarely in ATTRIBs but handle it)
    "MATERIAL": "material",
    "MATL": "material",
    "MTL": "material",
}


def _extract_attr_fields(block_attrs: dict) -> dict:
    """Translate raw ATTRIB tag/value dict into canonical field names."""
    out = {}
    for raw_tag, val in block_attrs.items():
        if not val: continue
        canon = _ATTR_TAG_CANONICAL.get(raw_tag)
        if canon and canon not in out:
            out[canon] = val.strip()
    return out


def _fill_variant_from_cluster(v: FabVariant, cluster: dict,
                                bbox_l: float, bbox_w: float,
                                dim_positioned: list,
                                layers: list,
                                attr_fields: dict = None) -> None:
    """Populate a single FabVariant from one drawing-region cluster.

    `attr_fields` is the canonical dict from title-block ATTRIB entities
    (e.g. {'description': 'GLAND PLATE', 'rev': '0'}) — these take priority
    over any text-layer extraction because they're authoritative fields.
    """
    attr_fields = attr_fields or {}
    chunks = cluster["chunks"]
    cluster_text = " | ".join(c[2] for c in chunks)

    # Title-block spec: T × W × L + material
    spec = _parse_title_block_spec(cluster_text)

    # Accept bbox fallback only when cluster bbox is part-sized.
    cbbox = cluster["bbox"]
    if cbbox is not None:
        cw = cbbox[2] - cbbox[0]
        ch = cbbox[3] - cbbox[1]
        cbbox_is_part_like = (0 < cw <= 3000 and 0 < ch <= 3000)
    else:
        cbbox_is_part_like = False

    # Stash the cluster bbox on the variant so the "View" column can
    # crop the DXF render to just this drawing's region.
    v.cluster_bbox = cbbox

    if spec.get("length") and spec.get("width"):
        v.length = spec["length"]
        v.width = spec["width"]
    elif 0 < bbox_l <= 3000 and 0 < bbox_w <= 3000:
        # 3m threshold covers real sheet-metal parts (max sheet is 1524×3048
        # and largest fab parts rarely exceed 3m in any dimension) while
        # still rejecting full A0 drawing-sheet bboxes (~4m × 2.5m).
        v.length = bbox_l
        v.width = bbox_w
    else:
        # DIMENSION-entity fallback, filtered to this cluster's region.
        dims = _filter_dims_in_bbox(dim_positioned, cbbox)
        plausible = sorted({x for x in dims if 10 <= x <= 3000}, reverse=True)
        if len(plausible) >= 2:
            v.length = plausible[0]
            v.width = plausible[1]
        elif plausible:
            v.length = plausible[0]

    if spec.get("thickness") and v.thickness is None:
        v.thickness = spec["thickness"]
    if spec.get("material") and not v.material:
        v.material = spec["material"]

    # Drawing-level fields from this cluster. ATTRIB-based values (from
    # AutoCAD title-block template) are authoritative; fall back to text
    # extraction only when an attribute isn't present.
    drg = attr_fields.get("drg_no", "") or _extract_drg_no(chunks)
    if drg and not getattr(v, "_extracted_drg_no", None):
        v._extracted_drg_no = drg
    rev = attr_fields.get("rev", "") or _extract_rev(chunks)
    if rev and not v.rev:
        v.rev = rev
    desc = attr_fields.get("description", "") or _extract_description(chunks, drg_no=drg)
    if desc and not v.description:
        v.description = desc
    # Material from ATTRIB is a nice-to-have; text extraction usually wins
    # because it's more specific (grade + spec standard).
    if not v.material and attr_fields.get("material"):
        v.material = attr_fields["material"]

    # Process hint from cluster text + global layers.
    haystack = " ".join(layers) + " " + cluster_text
    hit = _match_process_keyword(haystack)
    if hit and not v.process_hint_dxf:
        v.process_hint_dxf = hit

    # ── Auto-pick the smallest standard sheet that holds the part ──
    # If the variant's dims don't fit the default 1220×2440 sheet,
    # bump the sheet_name up to the smallest one that DOES fit (or to
    # the 1300×3000 "Oversize" stock as the operator's preferred
    # fallback for oversized work).
    try:
        from data.constants import pick_sheet_for_part, DEFAULT_SHEET
        if v.length and v.width:
            best = pick_sheet_for_part(v.length, v.width)
            # Only override when the auto-pick is non-default — i.e. the
            # part actually needed something larger than 1220×2440.
            if best and best != DEFAULT_SHEET:
                v.sheet_name = best
    except Exception:
        pass


def _parse_dxf(fp: str, v: FabVariant) -> list[FabVariant]:
    """Fill DXF-derived fields on `v` for the FIRST drawing detected in the
    file. Returns a list of ADDITIONAL FabVariants for any extra drawings
    found on the same sheet — empty list for single-drawing DXFs.
    Silent on missing library / parse errors.
    """
    try:
        from core.cad_reader import HAS_EZDXF, read_dxf
        if not HAS_EZDXF:
            return []
        d = read_dxf(fp)
        if not d:
            return []

        # Common fields used by every cluster.
        bbox_l = d.get("length") or 0
        bbox_w = d.get("width") or 0
        dim_positioned = d.get("dim_positioned") or []
        layers = d.get("layers") or []
        mtext_chunks = d.get("mtext_chunks") or []

        # AutoCAD title-block ATTRIB entities — translated to canonical
        # field names (description, drg_no, rev, paper_size, material).
        attr_fields = _extract_attr_fields(d.get("block_attrs") or {})

        # Geometry features (same for all clusters — they share the DXF).
        v.n_bends = int(d.get("n_bends") or 0)
        v.n_holes = int(d.get("holes") or 0)
        v.weld_len_mm = float(d.get("weld_length") or 0)

        clusters = _cluster_drawings(mtext_chunks)
        # Drop clusters crossed out with an X (cancelled / superseded
        # drawings — operators put a big diagonal X over them on the sheet).
        all_lines = d.get("all_lines") or []
        if all_lines:
            filtered = [c for c in clusters
                        if not _is_cluster_cancelled(c["bbox"], all_lines)]
            if not filtered and clusters:
                # Every cluster on this sheet was X'd out — the operator
                # cancelled the entire drawing. Skip it entirely (don't
                # fetch any data, don't emit a row).
                v._skip = True
                return []
            clusters = filtered

        # First cluster fills `v`. Subsequent clusters become extra variants.
        if not clusters:
            # No chunks — very degenerate DXF. Fall back to the classic flow:
            # joined annotations + bbox/DIMENSION.
            ann = d.get("annotations") or ""
            spec = _parse_title_block_spec(ann)
            if spec.get("length") and spec.get("width"):
                v.length = spec["length"]; v.width = spec["width"]
            elif 0 < bbox_l <= 2000 and 0 < bbox_w <= 2000:
                v.length = bbox_l; v.width = bbox_w
            if spec.get("thickness") and v.thickness is None:
                v.thickness = spec["thickness"]
            if spec.get("material") and not v.material:
                v.material = spec["material"]
            # Apply ATTRIB fields even when there were no MTEXT clusters.
            if attr_fields.get("drg_no") and not getattr(v, "_extracted_drg_no", None):
                v._extracted_drg_no = attr_fields["drg_no"]
            if attr_fields.get("rev") and not v.rev:
                v.rev = attr_fields["rev"]
            if attr_fields.get("description") and not v.description:
                v.description = attr_fields["description"]
            return []

        _fill_variant_from_cluster(v, clusters[0], bbox_l, bbox_w,
                                    dim_positioned, layers,
                                    attr_fields=attr_fields)

        extras: list[FabVariant] = []
        for cluster in clusters[1:]:
            extra = FabVariant(name=v.name)
            # Inherit the shared geometry features (per-cluster features
            # aren't tracked by cad_reader today — they apply to the whole DXF).
            extra.n_bends = v.n_bends
            extra.n_holes = v.n_holes
            extra.weld_len_mm = v.weld_len_mm
            extra.dxf_path = v.dxf_path
            # ATTRIB fields apply to the whole DWG (single title-block
            # template) — subsequent drawings keep the same rev/desc
            # unless the operator overrides manually.
            _fill_variant_from_cluster(extra, cluster, bbox_l, bbox_w,
                                        dim_positioned, layers,
                                        attr_fields=attr_fields)
            extras.append(extra)

        # Safety-net dedup: collapse variants that resolve to the same
        # (L, W, T, material) signature to one row. This handles the case
        # where cancelled copies produce identical metadata but the X-mark
        # detector didn't recognise them (e.g. cancel drawn with a single
        # slash, or dashed lines). Keeps one copy of each signature.
        def _sig(var):
            return (var.length, var.width, var.thickness, var.material or "")
        seen_signatures = {_sig(v)}
        kept = []
        for var in extras:
            s = _sig(var)
            if s in seen_signatures:
                continue
            seen_signatures.add(s)
            kept.append(var)
        return kept
    except Exception:
        return []


# Process label taxonomy — matches the Siemens STD WD xlsx exactly.
PROCESS_OPTIONS = [
    "BOLTED PRT",
    "WELDING PROCESS",
    "ASSEMBLY PRT",
    "BENDING PROCESS",
    "HEMIG TOOL ADD",
    "LOUVER SHEET",
    "LOUVER SHEET/NUT WELD",
    "LOUVERS",
    "LOUVERS/WELDING PROCESS",
    "WELDING PRT/MASKING",
    "PUNCHING",
    "CUTTING",
]


# Shared process keyword lookup. Used identically by:
#   • the description/filename heuristic (fallback)
#   • the PDF OCR text scanner  (core/pdf_reader.read_pdf)
#   • the DXF layer/annotation scanner (_parse_dxf below)
# Most-specific patterns first so "LOUVER SHEET/NUT WELD" wins over "LOUVER".
PROCESS_KEYWORD_RULES = [
    (r"\bLOUVER\s*SHEET\s*[/+ ]*NUT\s*WELD", "LOUVER SHEET/NUT WELD"),
    (r"\bNUT\s*WELD\b",                       "LOUVER SHEET/NUT WELD"),
    (r"\bLOUVER\b",                           "LOUVER SHEET"),
    (r"\bWELD\w*\s+.*MASK",                   "WELDING PRT/MASKING"),
    (r"\bHEMIG\b",                             "HEMIG TOOL ADD"),
    (r"\bBOLT(ED|ING)?\b",                    "BOLTED PRT"),
    (r"\bWELD(ING|ED)?\b",                    "WELDING PROCESS"),
    (r"\bBEND(ING)?\b",                       "BENDING PROCESS"),
    (r"\bPUNCH\w*\b",                          "PUNCHING"),
    (r"\bASSEMBL\w*\b",                        "ASSEMBLY PRT"),
    (r"\bCUT(TING)?\b",                       "CUTTING"),
]

# Pre-compile once.
_PROCESS_KEYWORD_RE = [(re.compile(pat, re.IGNORECASE), proc)
                       for pat, proc in PROCESS_KEYWORD_RULES]


def _match_process_keyword(text: str) -> str:
    """Return the first matching canonical process label, or '' if none."""
    if not text:
        return ""
    for rx, proc in _PROCESS_KEYWORD_RE:
        if rx.search(text):
            return proc
    return ""


def _detect_process(v: FabVariant) -> str:
    """Detect every applicable process for this part. Returns a ' + '
    separated list — e.g. 'PUNCHING + BENDING + WELDING'.

    Sources combined:
      • Workflow category (text-based) — from PDF OCR, DXF
        annotations, or filename/description keywords
        (BOLTED PRT, WELDING PROCESS, LOUVER SHEET, etc.)
      • Feature-based operations from DXF geometry:
        - PUNCHING — drawing has any CIRCLE entities (holes)
        - BENDING — drawing has bend lines / bend angles
        - WELDING — drawing has weld symbols / weld length
      • Description keyword scan as a final fallback

    Excel-provided PROCESS values are applied before this runs (in
    `group_files`), so they always come through unchanged.
    """
    parts: list[str] = []

    # 1. Workflow category from text (Excel/PDF/DXF/description heuristic).
    main = ""
    if v.process_hint_pdf:
        main = v.process_hint_pdf
    elif v.process_hint_dxf:
        main = v.process_hint_dxf
    else:
        main = _match_process_keyword(f"{v.description} {v.name}")
    if main:
        parts.append(main)

    main_upper = " ".join(parts).upper()

    # 2. Feature-based operations from the DXF geometry. Each operation is
    # only added if not already implied by the workflow category.
    if v.n_holes and v.n_holes > 0 and "PUNCH" not in main_upper:
        parts.append("PUNCHING")
    if v.n_bends and v.n_bends > 0 and "BEND" not in main_upper:
        parts.append("BENDING")
    if (v.weld_len_mm and v.weld_len_mm > 0) and "WELD" not in main_upper:
        parts.append("WELDING")

    return " + ".join(parts)


def _parse_pdf(fp: str, v: FabVariant) -> None:
    """Fill PDF-derived fields (material, thickness, process). Silent on
    failure."""
    try:
        from core.pdf_reader import HAS_PDF, read_pdf
        if not HAS_PDF:
            return
        d = read_pdf(fp)
        if d.get("material"):
            v.material = d["material"]
        if d.get("thickness") is not None:
            v.thickness = d["thickness"]
        proc = d.get("process")
        if proc:
            v.process_hint_pdf = proc
    except Exception:
        pass


def _parse_3d(fp: str, v: FabVariant) -> None:
    """Fill dims + thickness from a STEP/IGES 3D model. Silent on missing OCP."""
    try:
        from core.cad_reader import HAS_OCC, read_step_iges
        if not HAS_OCC:
            return
        d = read_step_iges(fp)
        if not d:
            return
        if d.get("length"): v.length = d["length"]
        if d.get("width"): v.width = d["width"]
        if d.get("thickness_3d", 0) > 0 and v.thickness is None:
            v.thickness = d["thickness_3d"]
    except Exception:
        pass


def _parse_excel_rows(fp: str) -> list["FabVariant"]:
    """Parse a FAB-style Excel sheet — emit one FabVariant per data row.

    Expects columns similar to Kumar Enterprises 'STD.FAB. DETAILS.':
        DRG.NO. | REV | DESCRIPTION | Mate | T | W | L | QTY
    Header row detected by locating the row with 'DRG.NO.' (or 'DRG NO')
    somewhere in its cells. Returns variants with material/thickness/dims
    already populated so they can be merged with any DXF/PDF of the same
    drawing number.
    """
    out: list[FabVariant] = []
    try:
        import openpyxl
    except Exception:
        return out
    try:
        wb = openpyxl.load_workbook(fp, data_only=True)
    except Exception:
        return out

    for ws in wb.worksheets:
        # Find a header row containing DRG.NO. (within the first 5 rows).
        col_map = {}  # key → column index (1-based)
        header_row = None
        for r in range(1, min(6, ws.max_row + 1)):
            row_vals = [str(ws.cell(r, c).value or "").strip().upper()
                        for c in range(1, ws.max_column + 1)]
            if any("DRG" in v for v in row_vals):
                for c, v in enumerate(row_vals, 1):
                    if "DRG" in v: col_map["drg"] = c
                    elif v == "REV": col_map["rev"] = c
                    elif v in ("DESCRIPTION", "DESC"): col_map["desc"] = c
                    elif v in ("MATE", "MATL", "MATERIAL"): col_map["mat"] = c
                    elif v == "T": col_map["t"] = c
                    elif v == "W": col_map["w"] = c
                    elif v == "L": col_map["l"] = c
                    elif v == "QTY": col_map.setdefault("qty", c)
                    elif v in ("PROCESS", "OPERATION"): col_map["process"] = c
                header_row = r
                break
        if header_row is None or "drg" not in col_map:
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            drg = ws.cell(r, col_map["drg"]).value
            if not drg:
                continue
            name = str(drg).strip()
            v = FabVariant(name=name)
            if "rev" in col_map:
                rv = ws.cell(r, col_map["rev"]).value
                v.rev = "" if rv is None else str(rv)
            if "desc" in col_map:
                v.description = str(ws.cell(r, col_map["desc"]).value or "").strip()
            if "mat" in col_map:
                m = ws.cell(r, col_map["mat"]).value
                if m: v.material = str(m).strip()
            for attr, key in (("thickness", "t"), ("width", "w"), ("length", "l")):
                if key in col_map:
                    val = ws.cell(r, col_map[key]).value
                    try:
                        if val is not None: setattr(v, attr, float(val))
                    except (TypeError, ValueError): pass
            if "qty" in col_map:
                try:
                    q = ws.cell(r, col_map["qty"]).value
                    if q: v.qty = int(float(q))
                except (TypeError, ValueError): pass
            if "process" in col_map:
                p = ws.cell(r, col_map["process"]).value
                if p: v.process = str(p).strip()
            out.append(v)
    return out


def _recompute_missing(v: FabVariant) -> None:
    v.missing = []
    if not v.material: v.missing.append("Material")
    if v.thickness is None: v.missing.append("Thickness")
    if v.length is None: v.missing.append("Length")
    if v.width is None: v.missing.append("Width")


def group_files(file_paths: list[str], progress=None,
                db=None) -> list[FabPart]:
    """Parse every file and group into FabPart rows for the preview table.

    Pairing rule: filename stem (without extension) is the grouping key.
    `JVD714E-1.dxf` + `JVD714E-1.pdf` pair into one variant named
    'JVD714E-1'. The master row is derived by stripping any trailing '-N'.

    `db` is an optional QuoteDB — when provided, the process column gets
    auto-filled from the cached `process_cache` table and any PROCESS values
    read from uploaded Excel files are written back to the cache for future
    uploads.

    If `progress` is callable it will be invoked as progress(i, n, fp) after
    each file — the UI uses this to keep a status counter fresh and to
    redraw the table incrementally.
    """
    # Speed ordering: instant parsers first (xlsx, dxf, 3D) so rows appear
    # immediately, PDFs (OCR) last so their minutes-long pass doesn't block
    # the visible feedback. DWG files go last since they're no-ops without
    # a converter.
    def _speed_key(p):
        e = os.path.splitext(p)[1].lower()
        if e in (".xlsx", ".xls"): return 0
        if e == ".dxf": return 1
        if e in (".step", ".stp", ".iges", ".igs"): return 2
        if e == ".pdf": return 3
        if e == ".dwg": return 5
        return 4
    ordered = sorted(file_paths, key=_speed_key)

    # Group variants by match key (normalized stem). Display name is the
    # SHORTEST stem seen for that key so the output table looks clean.
    by_key: dict[str, FabVariant] = {}
    n = len(ordered)

    def _pick_display_name(v: FabVariant, stem: str):
        # Always strip trailing dates from the display name so the
        # DRG.NO. column reads like a real drawing number, not a
        # filename-with-date.
        clean = _clean_stem(stem)
        if not v.name or len(clean) < len(v.name):
            v.name = clean

    for i, fp in enumerate(ordered, 1):
        stem = _stem(fp)
        key = _match_key(stem)
        ext = os.path.splitext(fp)[1].lower()

        if ext in (".xlsx", ".xls"):
            for ev in _parse_excel_rows(fp):
                ekey = _match_key(ev.name)
                existing = by_key.get(ekey)
                if existing is None:
                    by_key[ekey] = ev
                else:
                    _pick_display_name(existing, ev.name)
                    if not existing.material and ev.material:
                        existing.material = ev.material
                    if existing.thickness is None and ev.thickness is not None:
                        existing.thickness = ev.thickness
                    if existing.length is None and ev.length is not None:
                        existing.length = ev.length
                    if existing.width is None and ev.width is not None:
                        existing.width = ev.width
                    if not existing.description and ev.description:
                        existing.description = ev.description
                    if not existing.rev and ev.rev:
                        existing.rev = ev.rev
        else:
            v = by_key.setdefault(key, FabVariant(name=stem))
            _pick_display_name(v, stem)
            extras: list[FabVariant] = []
            if ext == ".dxf":
                v.dxf_path = fp
                extras = _parse_dxf(fp, v) or []
            elif ext == ".pdf":
                v.pdf_path = fp
                _parse_pdf(fp, v)
            elif ext in (".step", ".stp", ".iges", ".igs"):
                _parse_3d(fp, v)
            elif ext == ".dwg":
                # Try auto-conversion to DXF (LibreDWG or ODA). If it works,
                # read the resulting DXF so L/W/holes/bends get extracted.
                converted = _dwg_to_dxf(fp)
                if converted:
                    v.dxf_path = converted
                    extras = _parse_dxf(converted, v) or []
                else:
                    v.dxf_path = fp
                    v.description = (v.description
                                     or "DWG — install a converter to extract L/W")

            # Multi-drawing DXF/DWG: file contained more than one title block.
            # Naming convention (per user spec): first drawing uses its real
            # DRG.NO. if extractable, else the cleaned filename stem;
            # subsequent drawings get `-1`, `-2`, `-3` on that base.
            if extras:
                first_drg = getattr(v, "_extracted_drg_no", "") or ""
                if first_drg and _looks_like_drg_no(first_drg):
                    v.name = first_drg
                    base = first_drg
                else:
                    v.name = _clean_stem(stem)
                    base = v.name
                for i, ev in enumerate(extras, 1):
                    ev.name = f"{base}-{i}"
                    ev_key = _match_key(ev.name)
                    uniq = ev_key
                    n = 1
                    while uniq in by_key and by_key[uniq] is not ev:
                        uniq = f"{ev_key}_{n}"; n += 1
                    by_key[uniq] = ev
            else:
                # Single-drawing DXF/DWG — still prefer an extracted real
                # drawing number if available (cleaner than the filename).
                single_drg = getattr(v, "_extracted_drg_no", "") or ""
                if single_drg and _looks_like_drg_no(single_drg):
                    v.name = single_drg
                else:
                    # Keep whatever _pick_display_name already chose (which
                    # now strips the date suffix).
                    pass

        if progress:
            try: progress(i, n, fp)
            except Exception: pass

    # Drop variants flagged as cancelled (entire DWG was X'd out — see
    # `_parse_dxf`). The user explicitly asked for these to be omitted
    # rather than shown as placeholder rows.
    by_key = {k: v for k, v in by_key.items() if not getattr(v, "_skip", False)}

    # Rebuild a by_stem-style dict for the rest of the flow below — but keyed
    # by the (now stable) match key so master/variant grouping works.
    by_stem = by_key

    # Group variants under their master name.
    by_master: dict[str, FabPart] = {}
    for stem, variant in sorted(by_stem.items()):
        master_name, suffix = _split_master_variant(stem)
        part = by_master.setdefault(master_name, FabPart(name=master_name))
        # If the stem IS the master name and has no suffix siblings yet,
        # treat it as a flat part for now — we'll reconsider once all
        # variants are known (a second pass below).
        part.variants.append(variant)

    # Second pass: for each FabPart, decide whether the master is itself
    # a variant (flat) or a header. If there's exactly one variant AND its
    # name equals the master, it's a flat part — collapse it into a single
    # variant list of one. If multiple variants exist, the master becomes
    # a header; if the master itself was uploaded, leave it as the first
    # variant row.
    for part in by_master.values():
        for v in part.variants:
            # Auto-fill the process if still blank — the description
            # heuristic falls back to the variant's name internally, so we
            # don't need to stamp the drawing number into the description
            # column (which would pollute it with DRG.NO. echoes).
            if not v.process:
                v.process = _detect_process(v)
            _recompute_missing(v)

    return list(by_master.values())
