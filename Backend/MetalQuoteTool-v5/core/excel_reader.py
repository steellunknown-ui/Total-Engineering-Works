"""Excel BOM / SAP-export parser.

Reads an .xlsx (or .xls) BOM-style sheet and returns a {drawing_no → quantity}
map that other readers / the UI can use to auto-fill the Quantity field after a
matching DXF or PDF is imported.

It is intentionally tolerant:
  • The header row may not be row 1 — we scan the first ~20 rows.
  • Column names vary across customers and ERP exports — we match by keyword.
  • All sheets in the workbook are merged; the first sheet to yield a usable
    map wins, but rows from later sheets are appended if they expose the same
    column shape.

Public surface
--------------
HAS_OPENPYXL : bool
    True if openpyxl could be imported (xlsx support).

normalize_drg(value: str | int | float | None) -> str
    Canonical form used as the lookup key (upper-case, whitespace and most
    punctuation removed). Use this on both the BOM side AND the DXF/PDF side
    before comparing.

read_excel(fp: str) -> dict
    Parses the file and returns:
        {
            "drg_qty":     {normalized_drg_no: int_qty, ...},
            "rows":        [ {drg_no, qty, sheet, row_index, raw_drg, raw_qty},
                             ... ],          # original rows, useful for UI list
            "sheet_used":  "Sheet1" or "<merged>",
            "drg_col":     "DRG. NO.",       # human-readable header found
            "qty_col":     "QTY",
            "missing":     []                # warnings (e.g. "No qty column")
        }
    On total failure raises RuntimeError with an explanatory message.
"""
from __future__ import annotations

import os
import re
from typing import Any

try:
    import openpyxl
    HAS_OPENPYXL = True
except Exception:                                   # pragma: no cover
    HAS_OPENPYXL = False


# ── Header recognition ──────────────────────────────────────────────────────
# Order matters: more specific patterns first.
_DRG_HEADER_PATTERNS = [
    r"^\s*dr?\.?w?g\.?\s*[ _]?(no\.?|number|num|#)\s*$",
    r"^\s*drawing\s*(no\.?|number|num|#)\s*$",
    r"^\s*part\s*(no\.?|number|num|#)\s*$",
    r"^\s*item\s*(no\.?|code|number)\s*$",
    r"^\s*sap\s*code\s*$",
    r"^\s*material\s*code\s*$",
    r"^\s*drg\.?\s*$",          # bare "DRG."
    r"^\s*dwg\.?\s*$",
]
_QTY_HEADER_PATTERNS = [
    r"^\s*qty(\.|/pnl|\s*per\s*panel|\s*per\s*pcs?)?\s*$",
    r"^\s*quantity\s*$",
    r"^\s*total\s*qty\.?\s*$",
    r"^\s*order\s*qty\.?\s*$",
    r"^\s*nos\.?\s*$",
    r"^\s*pcs?\.?\s*$",
    r"^\s*req\.?\s*qty\.?\s*$",
]

_DRG_RES = [re.compile(p, re.IGNORECASE) for p in _DRG_HEADER_PATTERNS]
_QTY_RES = [re.compile(p, re.IGNORECASE) for p in _QTY_HEADER_PATTERNS]

# A drawing number must contain at least one digit AND one letter, length 3-30.
# This is the same rule fab_grouper uses for DXF title-block parsing — kept
# in sync so the same key matches.
_DRG_JUNK_RE = re.compile(
    r"^(NTS|XXX|TITLE|MATERIAL|DRG|REV|SCALE|SHEET|DATE|SIGN|NAME|"
    r"ITEM|CHECKED|BY|DESCRIPTION|ASSLY|ASSEMBLY|SPEC|NOTES?|TOTAL|"
    r"SUM|GRAND)$",
    re.IGNORECASE,
)


def _looks_like_drg_no(val: str) -> bool:
    val = (val or "").strip()
    if not (3 <= len(val) <= 30):
        return False
    if _DRG_JUNK_RE.match(val.strip(" .,:-")):
        return False
    if not re.search(r"\d", val):
        return False
    if not re.search(r"[A-Za-z]", val):
        return False
    if not re.match(r"^[A-Za-z0-9][A-Za-z0-9\-_/.\s]{1,28}[A-Za-z0-9]$", val):
        return False
    return True


def normalize_drg(value: Any) -> str:
    """Return the canonical lookup key for a drawing number. Strips all
    whitespace and runs of punctuation, upper-cases, so that 'jvd 1206 q/b',
    'JVD1206Q-B' and 'JVD1206Q/B' all collapse to 'JVD1206QB'.

    Returns '' for empty / non-string-like input.
    """
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    # Strip everything that is not alnum.
    s = re.sub(r"[^A-Za-z0-9]+", "", s).upper()
    return s


# ── Header detection ────────────────────────────────────────────────────────
def _is_drg_header(cell: Any) -> bool:
    s = str(cell or "").strip()
    return any(r.match(s) for r in _DRG_RES)


def _is_qty_header(cell: Any) -> bool:
    s = str(cell or "").strip()
    return any(r.match(s) for r in _QTY_RES)


def _find_header_row(rows: list[tuple]) -> tuple[int, int, int, str, str]:
    """Scan the first 25 rows of a sheet and try to locate a row that contains
    both a 'drawing number' column AND a 'quantity' column.

    Returns (header_row_index, drg_col_index, qty_col_index, drg_label, qty_label).
    All -1 / "" if no usable header is found.
    """
    limit = min(25, len(rows))
    for r_idx in range(limit):
        row = rows[r_idx]
        drg_col = qty_col = -1
        drg_label = qty_label = ""
        for c_idx, cell in enumerate(row):
            if drg_col < 0 and _is_drg_header(cell):
                drg_col = c_idx
                drg_label = str(cell).strip()
            elif qty_col < 0 and _is_qty_header(cell):
                qty_col = c_idx
                qty_label = str(cell).strip()
        if drg_col >= 0 and qty_col >= 0:
            return r_idx, drg_col, qty_col, drg_label, qty_label
    return -1, -1, -1, "", ""


def _to_int_qty(value: Any) -> int | None:
    """Convert a cell value to an integer qty. Accepts '12', '12.0', 12, 12.0.
    Returns None for blank / 0 / non-numeric so callers can skip those rows."""
    if value is None:
        return None
    if isinstance(value, bool):                     # 'True' is not a quantity
        return None
    if isinstance(value, (int, float)):
        try:
            n = int(round(float(value)))
        except (ValueError, OverflowError):
            return None
        return n if n > 0 else None
    s = str(value).strip()
    if not s:
        return None
    # Strip trailing 'Nos' / 'pcs' / commas.
    s = re.sub(r"[,\s]+", "", s)
    s = re.sub(r"(?i)(nos|pcs|pc|qty)\.?$", "", s)
    try:
        n = int(round(float(s)))
        return n if n > 0 else None
    except ValueError:
        return None


# ── Main entry point ────────────────────────────────────────────────────────
def read_excel(fp: str) -> dict:
    if not HAS_OPENPYXL:
        raise RuntimeError(
            "openpyxl not installed. Run:\n\n    pip install openpyxl\n")
    if not os.path.exists(fp):
        raise FileNotFoundError(fp)

    wb = openpyxl.load_workbook(fp, data_only=True, read_only=False)

    all_rows: list[dict] = []
    drg_qty: dict[str, int] = {}
    sheets_used: list[str] = []
    drg_label_used = qty_label_used = ""
    missing: list[str] = []

    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.sheet_state != "visible":
            # Still read hidden sheets — SAP DATA / CU51 type sheets in MEPL
            # MI files often live there. The user can always re-hide them.
            pass
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        hdr_idx, drg_col, qty_col, drg_label, qty_label = _find_header_row(rows)
        if hdr_idx < 0:
            continue

        if not drg_label_used:
            drg_label_used, qty_label_used = drg_label, qty_label
        sheets_used.append(sn)

        # Iterate data rows (after the header).
        for r_idx in range(hdr_idx + 1, len(rows)):
            row = rows[r_idx]
            if drg_col >= len(row) or qty_col >= len(row):
                continue
            raw_drg = row[drg_col]
            raw_qty = row[qty_col]

            drg_str = str(raw_drg).strip() if raw_drg is not None else ""
            if not drg_str or not _looks_like_drg_no(drg_str):
                continue

            qty = _to_int_qty(raw_qty)
            if qty is None:
                continue

            key = normalize_drg(drg_str)
            if not key:
                continue

            all_rows.append({
                "drg_no": drg_str,
                "qty": qty,
                "sheet": sn,
                "row_index": r_idx + 1,         # 1-based for the user
                "raw_drg": raw_drg,
                "raw_qty": raw_qty,
                "key": key,
            })

            # If the same drawing appears on multiple sheets/rows, sum them —
            # SAP exports sometimes split a single PO across line items.
            drg_qty[key] = drg_qty.get(key, 0) + qty

    if not all_rows:
        missing.append(
            "Could not find a 'Drawing No.' + 'Quantity' column pair in any "
            "sheet. Make sure the headers contain something like 'DRG. NO.' "
            "and 'QTY' / 'Quantity' / 'Order Qty'.")

    return {
        "drg_qty": drg_qty,
        "rows": all_rows,
        "sheet_used": ", ".join(sheets_used) if sheets_used else "",
        "drg_col": drg_label_used,
        "qty_col": qty_label_used,
        "missing": missing,
        "n_matches": len(drg_qty),
        "n_rows": len(all_rows),
    }


def lookup_qty(drg_qty_map: dict, drg_no: str) -> int | None:
    """Helper for the UI: given a raw drawing number from DXF/PDF, look it up
    in a map produced by read_excel().  Returns the int qty, or None if no
    match (the caller should leave the qty field untouched)."""
    if not drg_qty_map or not drg_no:
        return None
    key = normalize_drg(drg_no)
    return drg_qty_map.get(key)
