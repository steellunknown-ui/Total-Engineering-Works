import math
import os
from services.dxf_svg_service import dxf_to_svg
from typing import Optional, List

from fastapi import APIRouter, Depends, Query, HTTPException, status
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, or_

from core.database import get_db
from core.security import get_current_user
from models.user import User
from models.rfq import RFQ, RFQFile
from models.customer import Customer
from models.setting import Setting, SettingAudit
from models.material import Material, MaterialThickness, MaterialRateBand
from models.surface_finish import SurfaceFinish
from models.quote import Quote, QuoteItem, QuoteStatusHistory
from schemas.rfq_schema import (
    RFQListItem,
    RFQListResponse,
    RFQDetailResponse,
    RFQCustomerDetail,
    RFQFileDetail,
)
from schemas.material_schema import (
    MaterialItem, MaterialCreate, MaterialUpdate,
    MaterialThicknessItem, MaterialThicknessCreate, MaterialThicknessUpdate,
    MaterialRateBandItem, MaterialRateBandCreate, MaterialRateBandUpdate,
    SurfaceFinishItem, SurfaceFinishCreate, SurfaceFinishUpdate,
)
from schemas.settings_schema import SettingsResponse, SettingUpsertRequest, SettingAuditItem, SettingAuditResponse
from schemas.customer_schema import (
    CustomerListItem, CustomerListResponse,
    CustomerDetailResponse, CustomerRFQItem, CustomerQuoteItem,
    CustomerCreate, CustomerUpdate,
)

router = APIRouter(prefix="/api/admin", tags=["Admin"])


@router.get("/rfqs", response_model=RFQListResponse)
def list_rfqs(
    # Pagination
    page: int = Query(1, ge=1, description="Page number (1-indexed)"),
    page_size: int = Query(20, ge=1, le=100, description="Records per page (max 100)"),
    # Search
    search: Optional[str] = Query(None, description="Search by RFQ number, company name, or email"),
    # Filter
    status: Optional[str] = Query(None, description="Filter by status"),
    # Auth
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Protected admin endpoint: returns a paginated list of all RFQs.

    - Joins rfqs → customers for company/contact/email
    - Aggregates rfq_files for file_count
    - Supports search by rfq_number, company_name, email
    - Supports filter by status
    - Returns backend pagination metadata

    Requires: valid JWT (HttpOnly cookie or Authorization: Bearer header)
    """

    # ── Base query: join RFQ → Customer, subquery file counts ─────────────────
    # Subquery: count files per rfq
    file_count_sq = (
        db.query(
            RFQFile.rfq_id.label("rfq_id"),
            func.count(RFQFile.id).label("file_count"),
        )
        .group_by(RFQFile.rfq_id)
        .subquery()
    )

    # Main query
    query = (
        db.query(
            RFQ.id.label("rfq_id"),
            RFQ.rfq_number,
            Customer.company_name,
            Customer.contact_person,
            Customer.email,
            RFQ.status,
            RFQ.created_at,
            func.coalesce(file_count_sq.c.file_count, 0).label("file_count"),
            RFQ.customer_id,
        )
        .join(Customer, RFQ.customer_id == Customer.id)
        .outerjoin(file_count_sq, RFQ.id == file_count_sq.c.rfq_id)
    )

    # ── Search filter ─────────────────────────────────────────────────────────
    if search and search.strip():
        search_term = f"%{search.strip()}%"
        query = query.filter(
            or_(
                RFQ.rfq_number.ilike(search_term),
                Customer.company_name.ilike(search_term),
                Customer.email.ilike(search_term),
            )
        )

    # ── Status filter ─────────────────────────────────────────────────────────
    if status and status.strip() and status.lower() != "all":
        query = query.filter(RFQ.status == status.strip())

    # ── Total count (before pagination) ───────────────────────────────────────
    total = query.count()

    # ── Pagination ─────────────────────────────────────────────────────────────
    total_pages = math.ceil(total / page_size) if total > 0 else 1
    offset = (page - 1) * page_size

    rows = (
        query
        .order_by(RFQ.created_at.desc())
        .offset(offset)
        .limit(page_size)
        .all()
    )

    # ── Serialize rows ─────────────────────────────────────────────────────────
    rfq_items = [
        RFQListItem(
            rfq_id=row.rfq_id,
            rfq_number=row.rfq_number,
            company_name=row.company_name,
            contact_person=row.contact_person,
            email=row.email,
            status=row.status,
            created_at=row.created_at,
            file_count=row.file_count,
            customer_id=row.customer_id,
        )
        for row in rows
    ]

    return RFQListResponse(
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages,
        rfqs=rfq_items,
    )


# ─── Dashboard Stats ─────────────────────────────────────────────────────────

@router.get("/dashboard-stats")
def get_dashboard_stats(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Returns real statistics for the admin dashboard.
    """
    total = db.query(RFQ).count()
    pending = db.query(RFQ).filter(RFQ.status == 'Pending Review').count()
    quoted = db.query(RFQ).filter(RFQ.status == 'Quoted').count()
    
    # Simple conversion rate (quoted / total)
    conversion_rate = round((quoted / total * 100)) if total > 0 else 0

    return {
        "pending_rfqs": pending,
        "quoted_rfqs": quoted,
        "total_rfqs": total,
        "conversion_rate": conversion_rate
    }


# ─── Phase 4: GET /api/admin/rfqs/{rfq_id} ────────────────────────────────────

@router.get("/rfqs/{rfq_id}", response_model=RFQDetailResponse)
def get_rfq_detail(
    rfq_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Protected admin endpoint: returns the full detail of a single RFQ.

    - Eager-loads the Customer relationship (no N+1 query)
    - Eager-loads all RFQFile records
    - Returns 404 if the RFQ does not exist

    Requires: valid JWT (HttpOnly cookie or Authorization: Bearer header)
    """
    rfq = (
        db.query(RFQ)
        .options(
            joinedload(RFQ.customer),   # single JOIN to customers
            joinedload(RFQ.files),      # single JOIN to rfq_files
        )
        .filter(RFQ.id == rfq_id)
        .first()
    )

    if rfq is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"RFQ with id {rfq_id} not found.",
        )

    customer_detail = RFQCustomerDetail(
        id=rfq.customer.id,
        company_name=rfq.customer.company_name,
        contact_person=rfq.customer.contact_person,
        email=rfq.customer.email,
        phone=rfq.customer.phone,
    )

    file_details = [
        RFQFileDetail(
            id=f.id,
            file_name=f.file_name,
            file_type=f.file_type,
            storage_path=f.storage_path,
            uploaded_at=f.uploaded_at,
            storage_status=f.storage_status,
            archived_at=f.archived_at,
            archived_reason=f.archived_reason,
        )
        for f in rfq.files
    ]

    return RFQDetailResponse(
        rfq_id=rfq.id,
        rfq_number=rfq.rfq_number,
        status=rfq.status,
        project_description=rfq.project_description,
        created_at=rfq.created_at,
        customer=customer_detail,
        files=file_details,
        lead_source=rfq.lead_source or "Manual Upload",
        material=rfq.material,
        thickness=rfq.thickness,
        quantity=rfq.quantity,
        estimate_min=rfq.estimate_min,
        estimate_max=rfq.estimate_max,
    )


# ─── Phase 4: GET /api/admin/rfqs/{rfq_id}/files/{file_id}/download ───────────

@router.get("/rfqs/{rfq_id}/files/{file_id}/download")
def get_file_download_url(
    rfq_id: int,
    file_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Protected: generates a short-lived Supabase signed URL for a specific file.

    Strategy:
    - The 'rfq-files' bucket is PRIVATE — never publicly readable.
    - This endpoint verifies the requesting user is authenticated, then
      asks Supabase (via the service-role key on the server) for a
      60-second signed URL.
    - The signed URL is returned to the client; it expires automatically.
    - The Supabase service-role key never leaves the server.

    Returns 404 if the file does not belong to the specified RFQ.
    Returns 401 if the caller is unauthenticated.
    """
    # Verify the file exists and belongs to the given RFQ (prevents enumeration)
    rfq_file = (
        db.query(RFQFile)
        .filter(RFQFile.id == file_id, RFQFile.rfq_id == rfq_id)
        .first()
    )
    if rfq_file is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"File {file_id} not found for RFQ {rfq_id}.",
        )

    # Generate signed URL via Supabase service-role client
    supabase_url = os.environ.get("SUPABASE_URL")
    supabase_key = os.environ.get("SUPABASE_KEY")   # service-role key (never client-facing)

    if not supabase_url or not supabase_key:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Supabase credentials not configured.",
        )

    try:
        from supabase import create_client
        client = create_client(supabase_url, supabase_key)
        # 60-second expiry — short enough to be safe, long enough for a browser download
        signed = client.storage.from_("rfq-files").create_signed_url(
            path=rfq_file.storage_path,
            expires_in=60,
        )
        url = signed.get("signedURL") or signed.get("signed_url") or signed.get("data", {}).get("signedUrl")
        if not url:
            raise ValueError(f"Unexpected Supabase response: {signed}")
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Could not generate download URL: {str(exc)}",
        )

    return {
        "file_id": file_id,
        "file_name": rfq_file.file_name,
        "signed_url": url,
        "expires_in_seconds": 60,
    }


# ─── File Retention ──────────────────────────────────────────────────────────

from services.storage_service import archive_rfq_file, delete_rfq_file_permanently

@router.post("/rfq-files/{file_id}/archive")
def archive_file_endpoint(
    file_id: int, 
    current_user: User = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    file = db.query(RFQFile).filter(RFQFile.id == file_id).first()
    if not file:
        raise HTTPException(status_code=404, detail="File not found")
    if file.storage_status != "active":
        raise HTTPException(status_code=400, detail="File is not active")
    
    archive_rfq_file(file, db, reason="Manual Archive")
    return {"success": True, "message": "File archived successfully", "status": file.storage_status}

@router.post("/rfq-files/{file_id}/delete")
def delete_file_endpoint(
    file_id: int, 
    current_user: User = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    # Restrict to Admin
    if current_user.role not in ("Admin", "Super Admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Not authorized to permanently delete files")
        
    file = db.query(RFQFile).filter(RFQFile.id == file_id).first()
    if not file:
        raise HTTPException(status_code=404, detail="File not found")
        
    delete_rfq_file_permanently(file, db)
    return {"success": True, "message": "File permanently deleted", "status": file.storage_status}


# ─── Phase 5: GET /api/admin/rfqs/{rfq_id}/quote-data ────────────────────────

@router.get("/rfqs/{rfq_id}/quote-data")
def get_rfq_quote_data(
    rfq_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Phase 5: Fetches RFQ files from Supabase, parses them via existing CAD/PDF readers,
    and returns pre-shaped FabPartSpec data for the quote tool.
    Guarantees temp file cleanup.
    """
    import tempfile
    import urllib.request
    from pathlib import Path
    from core.cad_reader import read_cad
    from core.pdf_reader import read_pdf

    rfq = (
        db.query(RFQ)
        .options(
            joinedload(RFQ.customer),
            joinedload(RFQ.files),
        )
        .filter(RFQ.id == rfq_id)
        .first()
    )

    if not rfq:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"RFQ {rfq_id} not found.",
        )

    supabase_url = os.environ.get("SUPABASE_URL")
    supabase_key = os.environ.get("SUPABASE_KEY")

    if not supabase_url or not supabase_key:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Supabase credentials not configured.",
        )

    from supabase import create_client
    client = create_client(supabase_url, supabase_key)

    parsed_parts = []
    parse_errors = []

    for rfq_file in rfq.files:
        tmp_path = None
        try:
            # 1. Generate signed URL
            signed = client.storage.from_("rfq-files").create_signed_url(
                path=rfq_file.storage_path,
                expires_in=60,
            )
            url = signed.get("signedURL") or signed.get("signed_url") or signed.get("data", {}).get("signedUrl")
            if not url:
                raise ValueError("Could not get signed URL from Supabase")

            import urllib.request
            import ssl
            import tempfile
            from pathlib import Path
            from core.cad_reader import read_cad
            from core.pdf_reader import read_pdf
            
            ext = os.path.splitext(rfq_file.file_name)[1].lower()
            fd, tmp_path = tempfile.mkstemp(suffix=ext)
            
            ctx = ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
            
            with os.fdopen(fd, "wb") as f:
                # Add headers to avoid some basic blocks if any, though Supabase is fine
                req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=30, context=ctx) as response:
                    f.write(response.read())

            # 3. Parse and normalize
            file_stem = Path(rfq_file.file_name).stem

            def _resolve_material(raw: str, file_name: str) -> str:
                """3-tier material fallback: parsed → filename → customer history."""
                if raw:
                    return raw
                # Tier 2: filename pattern
                fu = file_name.upper()
                if "CRCA" in fu or "CRC" in fu:          return "CRCA"
                if "HR" in fu or "HOT" in fu or "HRS" in fu: return "HR Sheet"
                if "CR" in fu or "COLD" in fu:            return "CR Sheet"
                if "GI" in fu or "GALV" in fu:            return "GI Sheet"
                if "SS" in fu or "STAINLESS" in fu:       return "SS-304"
                if "AL" in fu or "ALUM" in fu:            return "Aluminium"
                if "MS" in fu or "MILD" in fu:            return "MS Sheet"
                # Tier 3: most-used material for this customer from quote history
                try:
                    from models.quote import Quote as _Q, QuoteItem as _QI
                    from sqlalchemy import func as _func
                    row = (
                        db.query(_QI.material, _func.count(_QI.material).label("cnt"))
                        .join(_Q, _QI.quote_id == _Q.id)
                        .filter(_Q.customer_id == rfq.customer_id)
                        .group_by(_QI.material)
                        .order_by(_func.count(_QI.material).desc())
                        .first()
                    )
                    if row and row.material:
                        return row.material
                except Exception:
                    pass
                return ""

            def _resolve_thickness(raw: float, annotations: str, block_attrs: dict,
                                   dim_measurements: list, file_name: str) -> float:
                """3-tier thickness fallback for DXF files."""
                import re as _re
                if raw and raw > 0:
                    return raw
                # Tier 1b: scan annotations/MTEXT/ATTRIB for thickness patterns
                text = ""
                if annotations:
                    text += " " + annotations
                if block_attrs:
                    for tag in ("THICKNESS", "THK", "THICK", "GAUGE", "MATERIAL_THICKNESS", "T"):
                        v = block_attrs.get(tag, "")
                        if v:
                            text += " " + str(v)
                if text.strip():
                    # Pattern: number followed by mm or T= or THK
                    for pat in [
                        r"\bT\s*[=:]\s*([0-9]+(?:\.[0-9]+)?)\s*(?:mm|MM)?",
                        r"([0-9]+(?:\.[0-9]+)?)\s*(?:mm|MM)?\s*(?:THK|THICK|THICKNESS)",
                        r"(?:THK|THICK|THICKNESS)\s*[=:]?\s*([0-9]+(?:\.[0-9]+)?)",
                    ]:
                        m = _re.search(pat, text, _re.IGNORECASE)
                        if m:
                            try:
                                v = float(m.group(1))
                                if 0.3 <= v <= 12.0:
                                    return v
                            except ValueError:
                                pass
                # Tier 1c: from dimension measurements — smallest plausible sheet gauge
                if dim_measurements:
                    for v in sorted(dim_measurements):
                        if 0.3 <= v <= 6.0:
                            return v
                from core.filename_parser import parse_thickness_from_filename
                
                # Tier 2: filename — e.g. "bracket_2mm.dxf", "1.5thk_panel.dxf", "1p5mm"
                thk = parse_thickness_from_filename(file_name)
                if thk > 0:
                    return thk
                # Tier 3: most-used thickness for this customer from quote history
                try:
                    from models.quote import Quote as _Q, QuoteItem as _QI
                    from sqlalchemy import func as _func
                    row = (
                        db.query(_QI.thickness, _func.count(_QI.thickness).label("cnt"))
                        .join(_Q, _QI.quote_id == _Q.id)
                        .filter(_Q.customer_id == rfq.customer_id)
                        .group_by(_QI.thickness)
                        .order_by(_func.count(_QI.thickness).desc())
                        .first()
                    )
                    if row and row.thickness and 0.3 <= float(row.thickness) <= 12.0:
                        return float(row.thickness)
                except Exception:
                    pass
                return 0.0

            if ext in (".dxf", ".step", ".stp", ".iges", ".igs"):
                cad_data = read_cad(tmp_path)
                if cad_data:
                    resolved_thickness = _resolve_thickness(
                        raw=float(cad_data.get("thickness_hint") or 0),
                        annotations=cad_data.get("annotations", ""),
                        block_attrs=cad_data.get("block_attrs", {}),
                        dim_measurements=cad_data.get("dim_measurements", []),
                        file_name=rfq_file.file_name,
                    )
                    parsed_parts.append({
                        "name": file_stem,
                        "material": _resolve_material(cad_data.get("material_hint") or "", rfq_file.file_name),
                        "thickness_mm": resolved_thickness,
                        "length_mm": float(cad_data.get("length") or cad_data.get("length_mm") or 0),
                        "width_mm": float(cad_data.get("width") or cad_data.get("width_mm") or 0),
                        "qty": 1,
                        "process": "",
                        "do_cut": True,
                        "do_bend": False,
                        "do_punch": False,
                        "do_weld": False,
                        "do_powder_dual": False,
                        "perim_mm": float(cad_data.get("perimeter_mm") or 0),
                        "int_cuts_mm": 0,
                        "bend_count": 0,
                        "bend_len_mm": 0,
                        "punch_count": int(cad_data.get("holes") or 0),
                        "punch_dia_mm": 0,
                        "weld_len_mm": 0,
                        "weld_spots": 0,
                        "drg_no": file_stem,
                        "geometry_svg": dxf_to_svg(tmp_path, target_width=380) if ext == ".dxf" else "",
                        "rfq_file_id": rfq_file.id,   # Phase 7A: authoritative DXF FK
                        "autoDetected": {
                            "material": bool(cad_data.get("material_hint")),
                            "thickness_mm": bool(cad_data.get("thickness_hint")),
                            "length_mm": bool(cad_data.get("length") or cad_data.get("length_mm")),
                            "width_mm": bool(cad_data.get("width") or cad_data.get("width_mm"))
                        }
                    })
                else:
                    parse_errors.append({"file_name": rfq_file.file_name, "reason": "Empty or invalid CAD file"})
            elif ext == ".pdf":
                pdf_data = read_pdf(tmp_path)
                if pdf_data:
                    parsed_parts.append({
                        "name": file_stem,
                        "material": _resolve_material(pdf_data.get("material") or "", rfq_file.file_name),
                        "thickness_mm": float(pdf_data.get("thickness_mm") or pdf_data.get("thickness") or 0),
                        "length_mm": float(pdf_data.get("length_mm") or pdf_data.get("length") or 0),
                        "width_mm": float(pdf_data.get("width_mm") or pdf_data.get("width") or 0),
                        "qty": 1,
                        "process": pdf_data.get("process") or "Parsed from PDF",
                        "do_cut": True,
                        "do_bend": "bend" in (pdf_data.get("process") or "").lower(),
                        "do_punch": False,
                        "do_weld": "weld" in (pdf_data.get("process") or "").lower(),
                        "do_powder_dual": "powder" in (pdf_data.get("process") or "").lower() or "pc" in (pdf_data.get("process") or "").lower(),
                        "perim_mm": 0,
                        "int_cuts_mm": 0,
                        "bend_count": 0,
                        "bend_len_mm": 0,
                        "punch_count": 0,
                        "punch_dia_mm": 0,
                        "weld_len_mm": 0,
                        "weld_spots": 0,
                        "drg_no": pdf_data.get("drg_no") or file_stem,
                        "rfq_file_id": rfq_file.id,   # Phase 7A: authoritative FK
                        "autoDetected": {
                            "material": bool(pdf_data.get("material")),
                            "thickness_mm": bool(pdf_data.get("thickness_mm") or pdf_data.get("thickness")),
                            "length_mm": bool(pdf_data.get("length_mm") or pdf_data.get("length")),
                            "width_mm": bool(pdf_data.get("width_mm") or pdf_data.get("width"))
                        }
                    })
                else:
                    parse_errors.append({"file_name": rfq_file.file_name, "reason": "Empty or invalid PDF file"})
            else:
                 parse_errors.append({"file_name": rfq_file.file_name, "reason": "Unsupported format for auto-parse"})
        except Exception as e:
            parse_errors.append({"file_name": rfq_file.file_name, "reason": str(e)})
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass

    return {
        "rfq_id": rfq.id,
        "rfq_number": rfq.rfq_number,
        "status": rfq.status,
        "customer": {
            "id": rfq.customer.id,
            "company_name": rfq.customer.company_name,
            "contact_person": rfq.customer.contact_person,
            "email": rfq.customer.email
        },
        "parsed_parts": parsed_parts,
        "parse_errors": parse_errors,
        "total_files": len(rfq.files),
        "parsed_count": len(parsed_parts)
    }

# ─── Phase 6: Quotes & Estimates ──────────────────────────────────────────────

from schemas.quote_schema import (
    EstimateRequest,
    EstimateResponse,
    CreateQuoteRequest,
    QuoteHistoryResponse,
    QuoteHistoryItem,
    QuoteStatusUpdateRequest,
    PdfGenerateResponse,
    PdfUrlResponse,
)
from services.estimate_service import generate_estimate
from services.quote_number_service import generate_quote_number
from services.quote_status_service import QuoteStatusMachine
from services.pdf_quote_service import generate_quote_pdf, get_quote_pdf_url
from models.quote import Quote, QuoteItem, QuoteStatusHistory

@router.post("/quotes/estimate", response_model=EstimateResponse)
def get_estimate(
    request: EstimateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Returns an unpersisted financial estimate based on current settings and provided part metrics.
    """
    return generate_estimate(db, request)

@router.post("/quotes")
def create_draft_quote(
    request: CreateQuoteRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Saves a draft quote to the database, capturing a snapshot of the current rates.
    Updates the RFQ status to 'In Review'.
    """
    rfq = db.query(RFQ).filter(RFQ.id == request.rfq_id).first()
    if not rfq:
        raise HTTPException(status_code=404, detail="RFQ not found")

    quote_number = generate_quote_number(db)
    est = request.estimate

    new_quote = Quote(
        quote_number=quote_number,
        rfq_id=request.rfq_id,
        customer_id=request.customer_id,
        subtotal=est.subtotal,
        margin_amount=est.margin_amount,
        gst_amount=est.gst_amount,
        grand_total=est.grand_total,
        status=QuoteStatusMachine.DRAFT,
        notes=request.notes,
        created_by=current_user.id
    )
    db.add(new_quote)
    db.flush() # get new_quote.id

    for item in est.items:
        db.add(QuoteItem(
            quote_id=new_quote.id,
            rfq_file_id=item.rfq_file_id,
            geometry_svg=item.geometry_svg,
            part_name=item.part_name,
            material=item.material,
            thickness=item.thickness,
            quantity=item.quantity,
            weight=item.weight,
            material_cost=item.material_cost,
            cutting_cost=item.cutting_cost,
            bending_cost=item.bending_cost,
            welding_cost=item.welding_cost,
            machining_cost=item.machining_cost,
            labour_cost=item.labour_cost,
            part_total=item.part_total,
            material_rate_snapshot=est.snapshots.get("material_rate_snapshot", 0.0),
            laser_rate_snapshot=est.snapshots.get("laser_rate_snapshot", 0.0),
            bending_rate_snapshot=est.snapshots.get("bending_rate_snapshot", 0.0),
            welding_rate_snapshot=est.snapshots.get("welding_rate_snapshot", 0.0),
            machining_rate_snapshot=est.snapshots.get("machining_rate_snapshot", 0.0),
            labour_rate_snapshot=est.snapshots.get("labour_rate_snapshot", 0.0),
            margin_snapshot=est.snapshots.get("margin_snapshot", 0.0),
            gst_snapshot=est.snapshots.get("gst_snapshot", 0.0),
        ))

    # Initial status history
    history = QuoteStatusHistory(
        quote_id=new_quote.id,
        old_status=None,
        new_status=QuoteStatusMachine.DRAFT,
        changed_by=current_user.id,
        notes="Initial draft creation"
    )
    db.add(history)

    # Update RFQ Status as requested
    rfq.status = "In Review"

    db.commit()
    return {"message": "Draft quote created successfully", "quote_number": quote_number, "quote_id": new_quote.id}

@router.get("/quote-history", response_model=QuoteHistoryResponse)
def list_quote_history(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    query = (
        db.query(Quote, RFQ, Customer, User)
        .join(RFQ, Quote.rfq_id == RFQ.id)
        .join(Customer, Quote.customer_id == Customer.id)
        .outerjoin(User, Quote.created_by == User.id)
    )

    if search and search.strip():
        search_term = f"%{search.strip()}%"
        query = query.filter(
            or_(
                Quote.quote_number.ilike(search_term),
                Customer.company_name.ilike(search_term),
                RFQ.rfq_number.ilike(search_term)
            )
        )
        
    if status and status.strip():
        query = query.filter(Quote.status == status.strip())
    
    total = query.count()
    total_pages = math.ceil(total / page_size) if total > 0 else 1
    offset = (page - 1) * page_size

    rows = query.order_by(Quote.created_at.desc()).offset(offset).limit(page_size).all()

    items = []
    for q, r, c, u in rows:
        items.append(QuoteHistoryItem(
            id=q.id,
            quote_number=q.quote_number,
            rfq_number=r.rfq_number,
            customer_name=c.company_name,
            status=q.status,
            grand_total=q.grand_total,
            created_by_name=u.email if u else None,
            created_at=q.created_at,
            # Phase 7A: PDF fields
            pdf_storage_path=q.pdf_storage_path,
            pdf_version=q.pdf_version,
            pdf_generated_at=q.pdf_generated_at,
        ))

    return QuoteHistoryResponse(
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages,
        quotes=items
    )

@router.patch("/quotes/{quote_id}/status")
def update_quote_status(
    quote_id: int,
    request: QuoteStatusUpdateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Updates the status of a quote, enforcing state machine rules and writing to audit history.
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")

    QuoteStatusMachine.transition_status(
        db=db,
        quote=quote,
        new_status=request.new_status,
        user=current_user,
        notes=request.notes
    )

    return {"message": f"Quote status updated to {request.new_status}"}


# ─── Phase 7A: PDF Generation Endpoints ───────────────────────────────────────

@router.post("/quotes/{quote_id}/generate-pdf", response_model=PdfGenerateResponse)
def generate_pdf(
    quote_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Generate a professional PDF for an Approved, Sent, or Accepted quote.

    - DXF geometry is fetched via rfq_file_id (no filename matching)
    - SVG geometry is cached in quote_item_svg_cache
    - Each generation creates a NEW versioned file (Q-2026-0001-v1.pdf, v2.pdf, ...)
    - Old versions are never overwritten
    - FAIL quality status → 422, PDF not stored
    - Returns signed URL (300s), version, quality_status, warnings
    """
    return generate_quote_pdf(quote_id=quote_id, db=db, current_user=current_user)


@router.get("/quotes/{quote_id}/pdf", response_model=PdfUrlResponse)
def get_pdf_url(
    quote_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Return a fresh 300-second signed URL for the latest generated PDF.
    Returns 404 if no PDF has been generated yet for this quote.
    """
    return get_quote_pdf_url(quote_id=quote_id, db=db)

# ─── Export Routes for Customer Approx ───────────────────────────────────────
import io
import datetime
from fastapi.responses import StreamingResponse
from jinja2 import Environment, FileSystemLoader, select_autoescape
from xhtml2pdf import pisa
from services.dxf_svg_service import render_part_preview_base64

jinja_env = Environment(
    loader=FileSystemLoader(os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates")),
    autoescape=select_autoescape(["html"]),
)
def datefmt(v):
    if isinstance(v, datetime.datetime) or isinstance(v, datetime.date):
        return v.strftime("%d %b %Y")
    return str(v) if v else "—"
jinja_env.filters["datefmt"] = datefmt

@router.post("/quotes/export-approx-pdf")
def export_approx_pdf(est: EstimateResponse):
    """
    Generates an approximate estimate PDF on the fly without saving to the database.
    Used for customer preview.
    """
    logo_path = os.path.join(os.path.dirname(__file__), "..", "templates", "logo.png")
    logo_base64 = ""
    if os.path.exists(logo_path):
        import base64
        with open(logo_path, "rb") as lf:
            logo_base64 = base64.b64encode(lf.read()).decode("utf-8")

    context = {
        "logo_base64": f"data:image/png;base64,{logo_base64}" if logo_base64 else "",
        "company": {
            "name": "Total Engineering Works",
            "address": "B-79 Ambad, MIDC, Nasik - 422010, Maharashtra, India",
            "phone": "+91 9545 450 786",
            "email": "rfq@totalengineeringworks.com",
            "website": "www.totalengineeringworks.com",
            "gst": "N/A",
            "logo_url": "",
        },
        "customer": {"company_name": "Customer", "contact_person": "Contact", "email": "customer@example.com", "phone": ""},
        "rfq": {"rfq_number": "Draft", "project_description": "Approximate Estimate"},
        "quote": {
            "quote_number": "Approximate",
            "subtotal": est.subtotal,
            "margin_amount": est.margin_amount,
            "gst_amount": est.gst_amount,
            "grand_total": est.grand_total,
            "notes": "This is an approximate cost estimate based on ±5% of the calculated total.",
            "created_at": datetime.datetime.now(),
            "valid_until": datetime.datetime.now() + datetime.timedelta(days=30),
        },
        "items": [
            {
                "part_name": item.part_name,
                "drg_no": item.part_name,
                "material": item.material,
                "thickness": item.thickness,
                "quantity": item.quantity,
                "weight": item.weight,
                "part_total": item.part_total,
                "line_total": item.line_total,
                "nesting_png": "",
                "part_preview_png": render_part_preview_base64(item.geometry_svg) if item.geometry_svg else "",
                "sheet_l": 2500, "sheet_w": 1250, "part_l": 0, "part_w": 0, "kerf": 0,
                "orientation": "0", "util": 0, "waste": 0, "pcs": 0, "used_area": 0, "waste_area": 0, "total_area": 0
            } for item in est.items
        ],
        "prepared_by": "System",
        "validity_days": 30,
        "terms": ["This is an approximate estimation and not a final binding quotation."],
        "margin_pct": 15.0,
        "gst_pct": 18.0,
        "total_weight": sum((item.weight * item.quantity) for item in est.items),
        "total_qty": sum(item.quantity for item in est.items),
        "total_parts": len(est.items),
        "estimate_date": datetime.datetime.now().strftime("%d %b %Y"),
        "total_cost": est.grand_total
    }
    
    html_str = jinja_env.get_template("estimate_pdf.html").render(**context)
    buf = io.BytesIO()
    pisa.CreatePDF(io.StringIO(html_str), dest=buf, encoding="utf-8")
    
    return StreamingResponse(
        io.BytesIO(buf.getvalue()),
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="Approximate_Estimate.pdf"'}
    )

@router.post("/quotes/export-excel")
def export_excel_bom(est: EstimateResponse):
    """
    Generates a professional Excel BOM from the estimate data without saving to the database.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
        
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="143B62", end_color="143B62", fill_type="solid")
    bd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    
    ws.merge_cells('A1:F1')
    ws['A1'] = "TATVA DYNAMICS PVT. LTD. - Bill of Materials"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    headers = ["Sr No", "Part Name", "Material", "Thickness (mm)", "Quantity", "Weight (kg)"]
    r = 3
    for c, h in enumerate(headers, 1):
        cl = ws.cell(r, c, h)
        cl.font = hf
        cl.fill = hfill
        cl.border = bd
        
    for i, item in enumerate(est.items, 1):
        r += 1
        row_data = [i, item.part_name, item.material, item.thickness, item.quantity, round(item.weight, 2)]
        for c, v in enumerate(row_data, 1):
            ws.cell(r, c, v).border = bd
            
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        mx = max(len(str(c.value or "")) for c in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(mx + 4, 40)
        
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="BOM_Export.xlsx"'}
    )


# ═══════════════════════════════════════════════════════════════════════════════
#  Phase 8.5 — Settings, Materials, Surface Finishes, Customers APIs
# ═══════════════════════════════════════════════════════════════════════════════

# ─── Settings ────────────────────────────────────────────────────────────────

@router.get("/settings")
def get_all_settings(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Returns all settings as a flat {key: value} dict."""
    rows = db.query(Setting).all()
    result = {}
    for row in rows:
        result[row.key] = row.str_value if row.str_value is not None else row.value
    return {"settings": result}


@router.put("/settings")
def upsert_settings(
    request: SettingUpsertRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Upserts settings and writes an audit row per changed key."""
    STRING_KEYS = {
        "company_name", "company_address", "company_phone", "company_email",
        "company_website", "company_gst", "company_logo_url", "terms_and_conditions",
    }
    updated = []
    for key, new_val in request.updates.items():
        row = db.query(Setting).filter(Setting.key == key).first()
        old_str = row.str_value if row and row.str_value is not None else (str(row.value) if row and row.value is not None else None)
        is_string = key in STRING_KEYS or isinstance(new_val, str)
        if row is None:
            row = Setting(key=key)
            db.add(row)
        if is_string:
            row.str_value = str(new_val)
            row.value = None
        else:
            row.value = float(new_val)
            row.str_value = None
        new_str = str(new_val)
        if old_str != new_str:
            db.add(SettingAudit(setting_key=key, old_value=old_str, new_value=new_str, changed_by=current_user.id))
        updated.append(key)
    db.commit()
    return {"updated": updated, "count": len(updated)}


@router.get("/settings/audit", response_model=SettingAuditResponse)
def get_settings_audit(
    page: int = Query(1, ge=1),
    page_size: int = Query(30, ge=1, le=100),
    key_filter: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    query = db.query(SettingAudit)
    if key_filter:
        query = query.filter(SettingAudit.setting_key.ilike(f"%{key_filter}%"))
    total = query.count()
    total_pages = math.ceil(total / page_size) if total > 0 else 1
    rows = query.order_by(SettingAudit.changed_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return SettingAuditResponse(
        total=total, page=page, page_size=page_size, total_pages=total_pages,
        items=[SettingAuditItem.model_validate(r) for r in rows]
    )


# ─── Materials ────────────────────────────────────────────────────────────────

@router.get("/materials", response_model=List[MaterialItem])
def list_materials(
    include_inactive: bool = Query(False),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(Material).options(joinedload(Material.thicknesses))
    if not include_inactive:
        q = q.filter(Material.active == True)
    return q.order_by(Material.name).all()


@router.post("/materials", response_model=MaterialItem)
def create_material(
    body: MaterialCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if db.query(Material).filter(Material.name == body.name).first():
        raise HTTPException(status_code=400, detail=f"Material '{body.name}' already exists.")
    mat = Material(name=body.name, density=body.density, active=body.active, created_by=current_user.id)
    db.add(mat)
    db.commit()
    db.refresh(mat)
    return mat


@router.put("/materials/{material_id}", response_model=MaterialItem)
def update_material(
    material_id: int,
    body: MaterialUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    mat = db.query(Material).filter(Material.id == material_id).first()
    if not mat:
        raise HTTPException(status_code=404, detail="Material not found.")
    if body.name is not None: mat.name = body.name
    if body.density is not None: mat.density = body.density
    if body.active is not None: mat.active = body.active
    mat.updated_by = current_user.id
    db.commit()
    db.refresh(mat)
    return mat


@router.delete("/materials/{material_id}")
def soft_delete_material(
    material_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    mat = db.query(Material).filter(Material.id == material_id).first()
    if not mat:
        raise HTTPException(status_code=404, detail="Material not found.")
    mat.active = False
    mat.updated_by = current_user.id
    db.commit()
    return {"message": f"Material '{mat.name}' deactivated."}


# ─── Material Thicknesses ─────────────────────────────────────────────────────

@router.post("/material-thicknesses", response_model=MaterialThicknessItem)
def add_thickness(
    body: MaterialThicknessCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not db.query(Material).filter(Material.id == body.material_id).first():
        raise HTTPException(status_code=404, detail="Material not found.")
    existing = db.query(MaterialThickness).filter(
        MaterialThickness.material_id == body.material_id,
        MaterialThickness.thickness_mm == body.thickness_mm
    ).first()
    if existing:
        existing.active = True
        db.commit()
        db.refresh(existing)
        return existing
    t = MaterialThickness(material_id=body.material_id, thickness_mm=body.thickness_mm, active=True)
    db.add(t)
    db.commit()
    db.refresh(t)
    return t


@router.put("/material-thicknesses/{thickness_id}", response_model=MaterialThicknessItem)
def update_thickness(
    thickness_id: int,
    body: MaterialThicknessUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    t = db.query(MaterialThickness).filter(MaterialThickness.id == thickness_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Thickness not found.")
    if body.thickness_mm is not None: t.thickness_mm = body.thickness_mm
    if body.active is not None: t.active = body.active
    db.commit()
    db.refresh(t)
    return t


@router.delete("/material-thicknesses/{thickness_id}")
def soft_delete_thickness(
    thickness_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    t = db.query(MaterialThickness).filter(MaterialThickness.id == thickness_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Thickness not found.")
    t.active = False
    db.commit()
    return {"message": f"Thickness {t.thickness_mm}mm deactivated."}


# ─── Material Rate Bands ──────────────────────────────────────────────────────

@router.get("/material-rate-bands", response_model=List[MaterialRateBandItem])
def list_rate_bands(
    material_name: Optional[str] = Query(None),
    include_inactive: bool = Query(False),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(MaterialRateBand)
    if material_name:
        q = q.filter(MaterialRateBand.material_name == material_name)
    if not include_inactive:
        q = q.filter(MaterialRateBand.active == True)
    return q.order_by(MaterialRateBand.material_name, MaterialRateBand.thickness_min).all()


@router.post("/material-rate-bands", response_model=MaterialRateBandItem)
def create_rate_band(
    body: MaterialRateBandCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    band = MaterialRateBand(
        material_name=body.material_name, thickness_min=body.thickness_min,
        thickness_max=body.thickness_max, rate_low=body.rate_low, rate_high=body.rate_high,
        active=True, created_by=current_user.id,
    )
    db.add(band)
    db.commit()
    db.refresh(band)
    return band


@router.put("/material-rate-bands/{band_id}", response_model=MaterialRateBandItem)
def update_rate_band(
    band_id: int,
    body: MaterialRateBandUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    band = db.query(MaterialRateBand).filter(MaterialRateBand.id == band_id).first()
    if not band:
        raise HTTPException(status_code=404, detail="Rate band not found.")
    for field, val in body.model_dump(exclude_none=True).items():
        setattr(band, field, val)
    band.updated_by = current_user.id
    db.commit()
    db.refresh(band)
    return band


@router.delete("/material-rate-bands/{band_id}")
def soft_delete_rate_band(
    band_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    band = db.query(MaterialRateBand).filter(MaterialRateBand.id == band_id).first()
    if not band:
        raise HTTPException(status_code=404, detail="Rate band not found.")
    band.active = False
    band.updated_by = current_user.id
    db.commit()
    return {"message": "Rate band deactivated."}


# ─── Surface Finishes ─────────────────────────────────────────────────────────

@router.get("/surface-finishes", response_model=List[SurfaceFinishItem])
def list_surface_finishes(
    include_inactive: bool = Query(False),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(SurfaceFinish)
    if not include_inactive:
        q = q.filter(SurfaceFinish.active == True)
    return q.order_by(SurfaceFinish.name).all()


@router.post("/surface-finishes", response_model=SurfaceFinishItem)
def create_surface_finish(
    body: SurfaceFinishCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if db.query(SurfaceFinish).filter(SurfaceFinish.name == body.name).first():
        raise HTTPException(status_code=400, detail=f"Surface finish '{body.name}' already exists.")
    sf = SurfaceFinish(name=body.name, rate=body.rate, unit=body.unit, active=body.active, created_by=current_user.id)
    db.add(sf)
    db.commit()
    db.refresh(sf)
    return sf


@router.put("/surface-finishes/{finish_id}", response_model=SurfaceFinishItem)
def update_surface_finish(
    finish_id: int,
    body: SurfaceFinishUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    sf = db.query(SurfaceFinish).filter(SurfaceFinish.id == finish_id).first()
    if not sf:
        raise HTTPException(status_code=404, detail="Surface finish not found.")
    for field, val in body.model_dump(exclude_none=True).items():
        setattr(sf, field, val)
    sf.updated_by = current_user.id
    db.commit()
    db.refresh(sf)
    return sf


@router.delete("/surface-finishes/{finish_id}")
def soft_delete_surface_finish(
    finish_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    sf = db.query(SurfaceFinish).filter(SurfaceFinish.id == finish_id).first()
    if not sf:
        raise HTTPException(status_code=404, detail="Surface finish not found.")
    sf.active = False
    sf.updated_by = current_user.id
    db.commit()
    return {"message": f"Surface finish '{sf.name}' deactivated."}


# ─── Customers CRM ────────────────────────────────────────────────────────────

def _build_customer_stats(customer_id: int, db: Session) -> dict:
    from models.quote import Quote as Q
    from models.rfq import RFQ as R
    rfq_count = db.query(R).filter(R.customer_id == customer_id).count()
    quote_rows = db.query(Q).filter(Q.customer_id == customer_id).all()
    quote_count = len(quote_rows)
    accepted_count = sum(1 for q in quote_rows if q.status == "Accepted")
    rejected_count = sum(1 for q in quote_rows if q.status == "Rejected")
    total_revenue = sum(q.grand_total for q in quote_rows if q.status == "Accepted")
    avg_quote_value = (sum(q.grand_total for q in quote_rows) / quote_count) if quote_count > 0 else 0.0
    conversion_rate = round((accepted_count / quote_count * 100), 1) if quote_count > 0 else 0.0
    last_rfq = db.query(func.max(R.created_at)).filter(R.customer_id == customer_id).scalar()
    last_quote = db.query(func.max(Q.created_at)).filter(Q.customer_id == customer_id).scalar()
    last_activity = max(filter(None, [last_rfq, last_quote]), default=None)
    return dict(rfq_count=rfq_count, quote_count=quote_count, accepted_count=accepted_count,
                rejected_count=rejected_count, total_revenue=round(total_revenue, 2),
                avg_quote_value=round(avg_quote_value, 2), conversion_rate=conversion_rate,
                last_activity=last_activity, last_rfq_date=last_rfq, last_quote_date=last_quote)


def _build_material_stats(customer_id: int, db: Session) -> dict:
    from models.quote import Quote as Q, QuoteItem as QI
    mat_row = (db.query(QI.material, func.count(QI.material).label("cnt"))
               .join(Q, QI.quote_id == Q.id).filter(Q.customer_id == customer_id)
               .group_by(QI.material).order_by(func.count(QI.material).desc()).first())
    thk_row = (db.query(QI.thickness, func.count(QI.thickness).label("cnt"))
               .join(Q, QI.quote_id == Q.id).filter(Q.customer_id == customer_id)
               .group_by(QI.thickness).order_by(func.count(QI.thickness).desc()).first())
    return dict(most_used_material=mat_row.material if mat_row else None,
                most_used_thickness=float(thk_row.thickness) if thk_row else None)


@router.get("/customers", response_model=CustomerListResponse)
def list_customers(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None),
    lead_source: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(Customer)
    if search and search.strip():
        term = f"%{search.strip()}%"
        q = q.filter(or_(Customer.company_name.ilike(term), Customer.email.ilike(term),
                         Customer.contact_person.ilike(term)))
    if lead_source:
        q = q.filter(Customer.lead_source == lead_source)
    total = q.count()
    total_pages = math.ceil(total / page_size) if total > 0 else 1
    customers = q.order_by(Customer.company_name).offset((page - 1) * page_size).limit(page_size).all()
    items = []
    for c in customers:
        stats = _build_customer_stats(c.id, db)
        items.append(CustomerListItem(
            id=c.id, company_name=c.company_name, contact_person=c.contact_person,
            email=c.email, phone=c.phone, website=c.website, lead_source=c.lead_source,
            created_at=c.created_at, updated_at=c.updated_at, **stats))
    return CustomerListResponse(total=total, page=page, page_size=page_size, total_pages=total_pages, customers=items)


@router.get("/customers/export")
def export_customers(
    export_all: bool = Query(False),
    search: Optional[str] = Query(None),
    lead_source: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export customers as Excel — filtered view or all."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    q = db.query(Customer)
    if not export_all:
        if search and search.strip():
            term = f"%{search.strip()}%"
            q = q.filter(or_(Customer.company_name.ilike(term), Customer.email.ilike(term),
                             Customer.contact_person.ilike(term)))
        if lead_source:
            q = q.filter(Customer.lead_source == lead_source)
    customers = q.order_by(Customer.company_name).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="143B62", end_color="143B62", fill_type="solid")
    bd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    headers = ["Company Name", "Contact Person", "Email", "Phone", "Website",
               "Lead Source", "RFQs", "Quotes", "Accepted", "Revenue (INR)",
               "Conversion %", "Last Activity", "Created At"]
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(1, c_idx, h)
        cell.font = hf
        cell.fill = hfill
        cell.border = bd
    for row_idx, c in enumerate(customers, 2):
        stats = _build_customer_stats(c.id, db)
        row_data = [c.company_name, c.contact_person, c.email, c.phone or "",
                    c.website or "", c.lead_source or "", stats["rfq_count"],
                    stats["quote_count"], stats["accepted_count"], stats["total_revenue"],
                    stats["conversion_rate"],
                    stats["last_activity"].strftime("%Y-%m-%d") if stats["last_activity"] else "",
                    c.created_at.strftime("%Y-%m-%d")]
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row_idx, c_idx, val).border = bd
    for col in ws.columns:
        mx = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 40)
    import io
    from fastapi.responses import StreamingResponse
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "All_Customers.xlsx" if export_all else "Customer_Export.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/customers/{customer_id}", response_model=CustomerDetailResponse)
def get_customer_detail(
    customer_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from models.rfq import RFQ as R
    from models.quote import Quote as Q
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Customer not found.")
    stats = _build_customer_stats(customer_id, db)
    mat_stats = _build_material_stats(customer_id, db)
    rfqs = db.query(R).filter(R.customer_id == customer_id).order_by(R.created_at.desc()).limit(50).all()
    quotes = db.query(Q).filter(Q.customer_id == customer_id).order_by(Q.created_at.desc()).limit(50).all()
    return CustomerDetailResponse(
        id=c.id, company_name=c.company_name, contact_person=c.contact_person,
        email=c.email, phone=c.phone, website=c.website, lead_source=c.lead_source,
        notes=c.notes, created_at=c.created_at, updated_at=c.updated_at,
        **stats, **mat_stats,
        rfqs=[CustomerRFQItem(id=r.id, rfq_number=r.rfq_number, status=r.status,
                              lead_source=r.lead_source, created_at=r.created_at) for r in rfqs],
        quotes=[CustomerQuoteItem(id=q.id, quote_number=q.quote_number, status=q.status,
                                  grand_total=q.grand_total, created_at=q.created_at) for q in quotes],
    )


@router.post("/customers", response_model=CustomerDetailResponse)
def create_customer(
    body: CustomerCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if db.query(Customer).filter(Customer.email == body.email).first():
        raise HTTPException(status_code=400, detail="A customer with this email already exists.")
    c = Customer(company_name=body.company_name, contact_person=body.contact_person,
                 email=body.email, phone=body.phone, website=body.website,
                 lead_source=body.lead_source or "Manual Entry", notes=body.notes)
    db.add(c)
    db.commit()
    db.refresh(c)
    return get_customer_detail(c.id, current_user, db)


@router.put("/customers/{customer_id}", response_model=CustomerDetailResponse)
def update_customer(
    customer_id: int,
    body: CustomerUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Customer not found.")
    for field, val in body.model_dump(exclude_none=True).items():
        setattr(c, field, val)
    db.commit()
    db.refresh(c)
    return get_customer_detail(c.id, current_user, db)
