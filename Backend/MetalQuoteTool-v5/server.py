#!/usr/bin/env python3
"""
Total Engineering Works Quote System — FastAPI HTTP Server
----------------------------------------
Wraps existing QuoteDB + constants logic and exposes a REST API
for the Next.js frontend running on port 3000.

Run:
    python server.py                  (default: port 8000)
    python server.py --port 9000      (custom port)

100% offline / LAN — no cloud, no internet required.
"""
from __future__ import annotations

import argparse
import sys
import threading
from contextlib import asynccontextmanager
from pathlib import Path
from typing import Optional

# ── Ensure project root is on sys.path (same trick as app.py) ──
ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware

from data.db import QuoteDB
from data import constants


# ═══════════════════════════════════════════════════════════════
#  App lifecycle — per-request DB connections (SQLite thread-safe)
# ═══════════════════════════════════════════════════════════════
#
#  SQLite connections cannot be shared across threads (FastAPI uses
#  a thread pool). Solution: store the DB path at startup and open
#  a fresh QuoteDB per request with check_same_thread=False patched
#  via a threadlocal. This keeps existing QuoteDB code untouched.

_db_path: Optional[Path] = None
_db_write_lock = threading.Lock()   # serialise writes across threads
_thread_local  = threading.local()  # per-thread connection cache


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Startup: resolve DB path + load constants. Shutdown: cleanup."""
    global _db_path
    # Open once just to resolve path + create schema
    _bootstrap = QuoteDB()
    constants.load_overrides(_bootstrap)
    _db_path = _bootstrap.db_path
    _bootstrap.close()
    print(f"[server] DB path -> {_db_path}", flush=True)
    print("[server] Total Engineering Works Quote API running -- http://localhost:8000", flush=True)
    
    import asyncio
    from services.cleanup_job import storage_cleanup_loop
    cleanup_task = asyncio.create_task(storage_cleanup_loop())
    
    yield
    
    cleanup_task.cancel()
    print("[server] Shutdown complete.", flush=True)


def get_db() -> QuoteDB:
    """
    Returns a per-thread QuoteDB instance.
    Opens a new connection if none exists for this thread, then caches it.
    Uses check_same_thread=False so FastAPI's thread pool can reuse it.
    """
    if _db_path is None:
        raise RuntimeError("Database not initialised — server not started yet.")
    db: Optional[QuoteDB] = getattr(_thread_local, 'db', None)
    if db is None:
        import sqlite3 as _sqlite3
        db = QuoteDB.__new__(QuoteDB)
        db.db_path = _db_path
        db.conn = _sqlite3.connect(
            str(_db_path),
            check_same_thread=False,
        )
        db.conn.row_factory = _sqlite3.Row
        db.conn.execute("PRAGMA foreign_keys = ON")
        db._init_schema()
        _thread_local.db = db
    return db


# ═══════════════════════════════════════════════════════════════
#  FastAPI app
# ═══════════════════════════════════════════════════════════════

app = FastAPI(
    title="Total Engineering Works Quote System API",
    description="REST API for the Total Engineering Works metal-sheet quoting tool.",
    version="1.0.0",
    lifespan=lifespan,
)

# CORS — allow Next.js dev server (port 3000) and production build
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://127.0.0.1:3000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Import and include routers
from routers.auth_router import router as auth_router
from routers.public_router import router as public_router
from routers.admin_router import router as admin_router
app.include_router(auth_router)
app.include_router(public_router)
app.include_router(admin_router)


# ═══════════════════════════════════════════════════════════════
#  Routes
# ═══════════════════════════════════════════════════════════════

@app.get("/api/health", tags=["System"])
def health_check():
    """
    Quick liveness probe.
    Returns {"status": "ok"} when the server is up.
    """
    return {"status": "ok"}


@app.get("/api/stats", tags=["System"])
def get_stats():
    """
    Aggregate statistics from the local SQLite DB.
    Returns total quotes, total value (INR), quotes this month,
    last backup timestamp, and DB file size in KB.
    """
    db = get_db()
    stats = db.stats()
    # Add DB file size
    try:
        size_kb = round(db.db_path.stat().st_size / 1024, 1)
    except Exception:
        size_kb = 0.0
    stats["db_size_kb"] = size_kb
    return stats


@app.get("/api/quotes", tags=["Quotes"])
def list_quotes(
    customer: Optional[str] = Query(None, description="Filter by exact customer name"),
    material: Optional[str] = Query(None, description="Filter by material (e.g. CRCA, HR Sheet)"),
    status: Optional[str]   = Query(None, description="Filter by status (draft, sent, accepted, rejected)"),
    date_from: Optional[str] = Query(None, description="Filter created_at >= date (YYYY-MM-DD)"),
    date_to:   Optional[str] = Query(None, description="Filter created_at <= date (YYYY-MM-DD)"),
    search: Optional[str]   = Query(None, description="Full-text search on part_name, customer, quote_no"),
    limit: int               = Query(100, ge=1, le=500, description="Max rows to return (1–500)"),
):
    """
    List quotes from the local DB with optional filters.
    All parameters are optional — omit them to get the latest 100 quotes.
    """
    db = get_db()
    try:
        quotes = db.list_quotes(
            customer=customer,
            material=material,
            status=status,
            date_from=date_from,
            date_to=date_to,
            search=search,
            limit=limit,
        )
        return {"count": len(quotes), "quotes": quotes}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


# ═══════════════════════════════════════════════════════════════
#  Phase 2 — Quote Generation Routes
# ═══════════════════════════════════════════════════════════════

import dataclasses
import datetime
import json
import sqlite3
from decimal import Decimal
from typing import Any

from fastapi import Body
from pydantic import BaseModel

from core.quote_engine import gen_quote, Quote, QLine
from core.nesting import Nest
from data import constants as _c


# ── Serialization helper ────────────────────────────────────────

def _convert(obj: Any) -> Any:
    """Recursively converts any value to a JSON-safe Python primitive."""
    if dataclasses.is_dataclass(obj) and not isinstance(obj, type):
        return {k: _convert(v) for k, v in dataclasses.asdict(obj).items()}
    if isinstance(obj, (datetime.datetime, datetime.date)):
        return obj.isoformat()
    if isinstance(obj, Decimal):
        return float(obj)
    if isinstance(obj, list):
        return [_convert(i) for i in obj]
    if isinstance(obj, dict):
        return {k: _convert(v) for k, v in obj.items()}
    return obj


def quote_to_dict(quote: Quote) -> dict:
    """
    Safely converts a Quote dataclass to a JSON-serializable dict.
    Handles all edge cases: nested dataclasses, datetime, Decimal.
    """
    return _convert(quote)


# ── Request model ───────────────────────────────────────────────

class QuoteRequest(BaseModel):
    # Part identity
    part_name: str
    customer: str = ""
    drg_no: str = ""
    # Geometry
    material: str
    thickness_mm: float
    length_mm: float
    width_mm: float
    box_height_mm: float = 0.0
    qty: int = 1
    # Sheet selection
    sheet_name: Optional[str] = None
    kerf_mm: float = 2.0
    # Rate
    rate_pct: float = 50.0
    manual_rate: float = 0.0
    # Cutting
    do_cut: bool = False
    cut_method: str = "laser"
    perim_mm: float = 0.0
    int_cuts_mm: float = 0.0
    # Bending
    do_bend: bool = False
    bend_count: int = 0
    bend_len_mm: float = 0.0
    # Punching
    do_punch: bool = False
    punch_count: int = 0
    punch_dia_mm: float = 0.0
    # Welding
    do_weld: bool = False
    weld_type: str = "mig"
    weld_len_mm: float = 0.0
    weld_spots: int = 0
    # Surface
    do_powder_dual: bool = False
    surface: str = "None"
    # Costing
    overhead_pct: float = 15.0
    profit_pct: float = 20.0
    # Extras
    hardware_rs: float = 0.0
    stretch_wrap_rs: float = 0.0
    packaging_rs: float = 0.0


# ── ROUTE 1: GET /api/constants ─────────────────────────────────

@app.get("/api/constants", tags=["Meta"])
def get_constants():
    """
    Returns all dropdown/reference data the frontend needs to
    populate forms: materials, thicknesses, surfaces, sheets, etc.
    All values are read live from constants.py — never hardcoded.
    """
    # Build standard_sheets as a list of {name, l, w} objects
    sheets = [
        {"name": name, "l": dims[0], "w": dims[1]}
        for name, dims in _c.STANDARD_SHEETS.items()
    ]
    return {
        "materials": _c.MATERIALS,
        "thicknesses": {mat: thk for mat, thk in _c.THICKNESSES.items()},
        "surfaces": _c.SURFACES,
        "standard_sheets": sheets,
        "cut_methods": ["laser", "plasma", "waterjet", "shearing"],
        "weld_types": ["mig", "tig", "arc", "spot"],
        "std_rates": _c.STD_RATES_PER_KG,
        "rate_bands": {
            mat: [{"t": t, "lo": lo, "hi": hi} for t, lo, hi in bands]
            for mat, bands in _c.RATE_BANDS.items()
        },
    }


# ── ROUTE 2: POST /api/quote/generate ──────────────────────────

@app.post("/api/quote/generate", tags=["Quotes"])
def generate_quote(req: QuoteRequest):
    """
    Core quote calculation endpoint. Maps request fields to gen_quote()
    and returns the full Quote object as JSON.
    """
    try:
        # cut_p = external perimeter; int_c = internal cut length
        cut_p = req.perim_mm if req.do_cut else 0.0
        int_c = req.int_cuts_mm if req.do_cut else 0.0

        quote = gen_quote(
            name=req.part_name,
            mat=req.material,
            t=req.thickness_mm,
            pl=req.length_mm,
            pw=req.width_mm,
            qty=req.qty,
            slider=req.rate_pct,
            cut_m=req.cut_method,
            cut_p=cut_p,
            int_c=int_c,
            n_bends=req.bend_count,
            b_len=req.bend_len_mm,
            n_holes=req.punch_count,
            h_dia=req.punch_dia_mm,
            w_type=req.weld_type,
            w_len=req.weld_len_mm,
            n_spots=req.weld_spots,
            surface=req.surface,
            oh_pct=req.overhead_pct,
            pr_pct=req.profit_pct,
            sheet_n=req.sheet_name,
            kerf=req.kerf_mm,
            box_h=req.box_height_mm,
            hardware=req.hardware_rs,
            stretch_wrap=req.stretch_wrap_rs,
            packaging=req.packaging_rs,
            apply_punch=req.do_punch,
            apply_bend=req.do_bend,
            apply_weld=req.do_weld,
            apply_pc_dual=req.do_powder_dual,
            manual_rate=req.manual_rate,
        )
    except Exception as exc:
        raise HTTPException(
            status_code=422,
            detail={"error": True, "message": str(exc), "route": "generate"},
        ) from exc

    try:
        result = quote_to_dict(quote)
        # Attach the customer + drg_no from request (not stored in Quote dataclass)
        result["customer"] = req.customer
        result["drg_no"] = req.drg_no
        return result
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={"error": True, "message": f"Serialization failed: {exc}", "route": "generate"},
        ) from exc


# ── ROUTE 3: POST /api/quote/save ──────────────────────────────

class SaveQuoteRequest(BaseModel):
    quote: dict                   # Full Quote JSON from /generate
    customer: str = ""
    cut_method: str = "laser"
    weld_type: str = "mig"
    box_height_mm: float = 0.0
    cad_file: str = ""
    status: str = "draft"


@app.post("/api/quote/save", tags=["Quotes"])
def save_quote(req: SaveQuoteRequest):
    """
    Reconstructs a Quote from its JSON dict and saves it to SQLite.
    Returns the assigned quote_no and database row id.
    """
    db = get_db()
    try:
        q_data = req.quote
        # Rebuild QLine objects
        lines = [
            QLine(
                desc=l["desc"], unit=l["unit"],
                qty=l["qty"], rate=l["rate"], amt=l["amt"]
            )
            for l in q_data.get("lines", [])
        ]
        # Rebuild optional Nest
        nest_obj = None
        if q_data.get("n"):
            nest_obj = Nest(**q_data["n"])

        quote = Quote(
            name=q_data.get("name", ""),
            mat=q_data.get("mat", ""),
            t=float(q_data.get("t", 0)),
            pl=float(q_data.get("pl", 0)),
            pw=float(q_data.get("pw", 0)),
            qty=int(q_data.get("qty", 1)),
            weight=float(q_data.get("weight", 0)),
            rate_kg=float(q_data.get("rate_kg", 0)),
            band_lo=float(q_data.get("band_lo", 0)),
            band_hi=float(q_data.get("band_hi", 0)),
            slider=float(q_data.get("slider", 50)),
            lines=lines,
            sub=float(q_data.get("sub", 0)),
            overhead=float(q_data.get("overhead", 0)),
            profit=float(q_data.get("profit", 0)),
            per_pc=float(q_data.get("per_pc", 0)),
            total=float(q_data.get("total", 0)),
            n=nest_obj,
            sheet_cost=float(q_data.get("sheet_cost", 0)),
            flat_info=q_data.get("flat_info", ""),
            overhead_pct=float(q_data.get("overhead_pct", 15)),
            profit_pct=float(q_data.get("profit_pct", 20)),
            surface=q_data.get("surface", "None"),
        )

        quote_id = db.save_quote(
            quote,
            customer=req.customer,
            cad_file=req.cad_file,
            cut_method=req.cut_method,
            weld_type=req.weld_type,
            box_height_mm=req.box_height_mm,
        )

        # Set status if not draft
        if req.status != "draft":
            db.update_status(quote_id, req.status)

        # Fetch back the assigned quote_no
        row = db.conn.execute(
            "SELECT quote_no FROM quotes WHERE id = ?", (quote_id,)
        ).fetchone()
        quote_no = row["quote_no"] if row else ""

        return {"success": True, "quote_id": quote_id, "quote_no": quote_no}

    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


# ── ROUTE 4: GET /api/quotes/{quote_id} ────────────────────────

@app.get("/api/quotes/{quote_id}", tags=["Quotes"])
def get_quote(quote_id: int):
    """
    Returns a single saved quote by its database ID.
    Includes the full quote_json parsed back to a dict.
    """
    db = get_db()
    row = db.conn.execute(
        "SELECT * FROM quotes WHERE id = ?", (quote_id,)
    ).fetchone()

    if not row:
        raise HTTPException(
            status_code=404,
            detail={"error": True, "message": f"Quote {quote_id} not found"},
        )

    result = dict(row)
    # Parse the stored quote_json back to a dict for easy frontend consumption
    if result.get("quote_json"):
        try:
            result["quote_data"] = json.loads(result["quote_json"])
        except (ValueError, TypeError):
            result["quote_data"] = None

    return result


# ── ROUTE 5: DELETE /api/quotes/{quote_id} ─────────────────────

@app.delete("/api/quotes/{quote_id}", tags=["Quotes"])
def delete_quote(quote_id: int):
    """
    Permanently deletes a quote by ID.
    Returns 404 if the quote does not exist.
    """
    db = get_db()
    row = db.conn.execute(
        "SELECT id FROM quotes WHERE id = ?", (quote_id,)
    ).fetchone()

    if not row:
        raise HTTPException(
            status_code=404,
            detail={"error": True, "message": f"Quote {quote_id} not found"},
        )

    db.delete_quote(quote_id)
    return {"success": True, "deleted_id": quote_id}


# ── ROUTE 6: PATCH /api/quotes/{quote_id}/status ──────────────────

class StatusUpdate(BaseModel):
    status: str  # draft | sent | accepted | rejected


@app.patch("/api/quotes/{quote_id}/status", tags=["Quotes"])
def update_quote_status(quote_id: int, body: StatusUpdate):
    """
    Update the status of a saved quote.
    Allowed values: draft, sent, accepted, rejected.
    """
    VALID = {"draft", "sent", "accepted", "rejected"}
    if body.status not in VALID:
        raise HTTPException(
            status_code=422,
            detail={"error": True, "message": f"Invalid status '{body.status}'. Must be one of: {sorted(VALID)}"},
        )
    db = get_db()
    row = db.conn.execute(
        "SELECT id FROM quotes WHERE id = ?", (quote_id,)
    ).fetchone()
    if not row:
        raise HTTPException(
            status_code=404,
            detail={"error": True, "message": f"Quote {quote_id} not found"},
        )
    db.update_status(quote_id, body.status)
    return {"success": True, "quote_id": quote_id, "status": body.status}


# ═══════════════════════════════════════════════════════════════
#  Phase 5 — FAB Sheet Routes
# ═══════════════════════════════════════════════════════════════

import shutil
import tempfile
from fastapi import File, UploadFile, Form
from core.cad_reader import read_cad
from core.pdf_reader import read_pdf
from core.excel_reader import read_excel, HAS_OPENPYXL
from services.dxf_svg_service import dxf_to_svg

@app.post("/api/fab/upload-file", tags=["FAB Sheet"])
async def fab_upload_file(file: UploadFile = File(...)):
    """
    Accepts a DXF, PDF, or Excel file and parses its metadata.
    Returns a standardized FabUploadResult matching the frontend.
    """
    # Create temp file
    ext = Path(file.filename).suffix.lower()
    fd, path = tempfile.mkstemp(suffix=ext)
    try:
        with open(path, "wb") as f:
            shutil.copyfileobj(file.file, f)
            
        result = {"detected": False, "reason": "Unsupported format"}
        
        # Route to appropriate reader
        if ext in (".dxf", ".step", ".stp", ".iges", ".igs"):
            try:
                import ezdxf
                cad_data = read_cad(path)
                if cad_data:
                    from core.pdf_reader import _detect_material as _dm, _detect_thickness as _dt
                    from core.filename_parser import parse_material_from_filename, parse_thickness_from_filename
                    import re as _re

                    _ann = cad_data.get("annotations", "")
                    _block_attrs = cad_data.get("block_attrs", {})
                    _dim_meas = cad_data.get("dim_measurements", [])

                    # --- MATERIAL resolution (3 tiers) ---
                    # Tier 1a: DXF text annotations scan
                    _raw_mat = _dm(_ann) or ""
                    # Tier 1b: block ATTRIB tags
                    if not _raw_mat and _block_attrs:
                        for _tag in ("MATERIAL", "MAT", "MATL", "MATERIAL_TYPE", "MAT_TYPE"):
                            _v = _block_attrs.get(_tag, "")
                            if _v:
                                _raw_mat = _v
                                break
                    # Tier 2: filename
                    detected_material = _raw_mat or parse_material_from_filename(file.filename) or ""

                    # --- THICKNESS resolution (3 tiers) ---
                    _raw_thk = 0.0
                    # Tier 1a: annotations
                    _dt_val = _dt(_ann)
                    if _dt_val and _dt_val > 0:
                        _raw_thk = float(_dt_val)
                    # Tier 1b: scan annotations + block_attrs for explicit T= / THK patterns
                    if not _raw_thk:
                        _text = _ann
                        if _block_attrs:
                            for _tag in ("THICKNESS", "THK", "THICK", "GAUGE", "MATERIAL_THICKNESS", "T"):
                                _v = _block_attrs.get(_tag, "")
                                if _v:
                                    _text += " " + str(_v)
                        for _pat in [
                            r"\bT\s*[=:]\s*([0-9]+(?:\.[0-9]+)?)\s*(?:mm|MM)?",
                            r"([0-9]+(?:\.[0-9]+)?)\s*(?:mm|MM)?\s*(?:THK|THICK|THICKNESS)",
                            r"(?:THK|THICK|THICKNESS)\s*[=:]?\s*([0-9]+(?:\.[0-9]+)?)",
                        ]:
                            _m = _re.search(_pat, _text, _re.IGNORECASE)
                            if _m:
                                try:
                                    _v = float(_m.group(1))
                                    if 0.3 <= _v <= 25.0:
                                        _raw_thk = _v
                                        break
                                except ValueError:
                                    pass
                    # Tier 1c: dimension measurements (smallest plausible gauge)
                    if not _raw_thk and _dim_meas:
                        for _v in sorted(_dim_meas):
                            if 0.3 <= _v <= 25.0:
                                _raw_thk = _v
                                break
                    # Tier 2: filename
                    detected_thickness = _raw_thk if _raw_thk > 0 else parse_thickness_from_filename(file.filename)

                    parts = [{
                        "temp_id": file.filename,
                        "drg_no": Path(file.filename).stem,
                        "description": "",
                        "material": detected_material,
                        "thickness_mm": detected_thickness,
                        "length_mm": float(cad_data.get("length") or cad_data.get("length_mm") or 0),
                        "width_mm": float(cad_data.get("width") or cad_data.get("width_mm") or 0),
                        "qty": 1,
                        "process": "",
                        "holes": int(cad_data.get("holes") or 0),
                        "perimeter_mm": float(cad_data.get("outer_perimeter") or cad_data.get("perimeter_mm") or 0),
                        "geometry_svg": dxf_to_svg(path, target_width=380) if ext == ".dxf" else "",
                        "confidence": "high",
                        "missing_fields": [
                            *([] if detected_material else ["Material"]),
                            *([] if detected_thickness else ["Thickness"]),
                        ]
                    }]
                    result = {
                        "file_name": file.filename,
                        "detected": True,
                        "type": "cad",
                        "parts": parts
                    }
                else:
                    result = {"file_name": file.filename, "detected": False, "reason": "Empty or invalid CAD file"}
            except ImportError:
                result = {"file_name": file.filename, "detected": False, "reason": "ezdxf not installed"}
            except Exception as e:
                result = {"file_name": file.filename, "detected": False, "reason": f"CAD parse error: {str(e)}"}
                
        elif ext == ".pdf":
            try:
                pdf_data = read_pdf(path)
                from core.filename_parser import parse_material_from_filename, parse_thickness_from_filename
                
                detected_material = pdf_data.get("material") or parse_material_from_filename(file.filename) or ""
                
                pdf_thk = float(pdf_data.get("thickness_mm") or pdf_data.get("thickness") or 0)
                detected_thickness = pdf_thk if pdf_thk > 0 else parse_thickness_from_filename(file.filename)
                
                parts = [{
                    "temp_id": file.filename,
                    "drg_no": pdf_data.get("drg_no") or Path(file.filename).stem,
                    "description": "",
                    "material": detected_material,
                    "thickness_mm": detected_thickness,
                    "length_mm": float(pdf_data.get("length_mm") or pdf_data.get("length") or 0),
                    "width_mm": float(pdf_data.get("width_mm") or pdf_data.get("width") or 0),
                    "qty": 1,
                    "process": pdf_data.get("process") or "Parsed from PDF",
                    "holes": 0,
                    "perimeter_mm": 0,
                    "confidence": "medium",
                    "missing_fields": pdf_data.get("missing_fields") or pdf_data.get("missing") or []
                }]
                result = {
                    "file_name": file.filename,
                    "detected": True,
                    "type": "pdf",
                    "parts": parts
                }
            except Exception as e:
                result = {"file_name": file.filename, "detected": False, "reason": f"PDF parse error: {str(e)}"}
                
        elif ext in (".xlsx", ".xls"):
            if not HAS_OPENPYXL:
                result = {"file_name": file.filename, "detected": False, "reason": "openpyxl not installed"}
            else:
                try:
                    excel_data = read_excel(path)
                    parts = []
                    for row in excel_data:
                        parts.append({
                            "temp_id": row.get("name") or "Part",
                            "drg_no": row.get("drg_no") or "",
                            "description": row.get("name") or "",
                            "material": row.get("material") or "",
                            "thickness_mm": float(row.get("t") or 0),
                            "length_mm": float(row.get("pl") or 0),
                            "width_mm": float(row.get("pw") or 0),
                            "qty": int(row.get("qty") or 1),
                            "process": row.get("process") or "",
                            "holes": 0,
                            "perimeter_mm": 0,
                            "confidence": "high",
                            "missing_fields": []
                        })
                    result = {
                        "file_name": file.filename,
                        "detected": True,
                        "type": "excel",
                        "parts": parts
                    }
                except Exception as e:
                    result = {"file_name": file.filename, "detected": False, "reason": f"Excel parse error: {str(e)}"}

        return result
    finally:
        import os
        os.close(fd)
        try:
            os.remove(path)
        except OSError:
            pass

class FabPartSpec(BaseModel):
    name: str
    material: str
    thickness_mm: float
    length_mm: float
    width_mm: float
    qty: int = 1
    # Process toggles
    do_cut: bool = True
    do_bend: bool = False
    do_punch: bool = False
    do_weld: bool = False
    do_powder_dual: bool = False
    # Extra CAD stats
    perim_mm: float = 0.0
    int_cuts_mm: float = 0.0
    bend_count: int = 0
    bend_len_mm: float = 0.0
    punch_count: int = 0
    punch_dia_mm: float = 0.0
    weld_len_mm: float = 0.0
    weld_spots: int = 0
    # Process
    process: str = ""
    drg_no: str = ""
    geometry_svg: str = ""

class GenerateBulkRequest(BaseModel):
    parts: list[FabPartSpec]
    customer: str = ""
    # Global settings for this batch
    overhead_pct: float = 15.0
    profit_pct: float = 20.0
    cut_method: str = "laser"
    kerf_mm: float = 2.0
    rate_pct: float = 50.0

@app.post("/api/fab/generate-bulk", tags=["FAB Sheet"])
def fab_generate_bulk(req: GenerateBulkRequest):
    """
    Evaluates a list of part specifications and generates Quotes.
    """
    results = []
    total_amount = 0.0
    total_weight = 0.0
    
    for p in req.parts:
        try:
            quote = gen_quote(
                name=p.name,
                mat=p.material,
                t=p.thickness_mm,
                pl=p.length_mm,
                pw=p.width_mm,
                qty=p.qty,
                slider=req.rate_pct,
                cut_m=req.cut_method,
                cut_p=p.perim_mm if p.do_cut else 0.0,
                int_c=p.int_cuts_mm if p.do_cut else 0.0,
                n_bends=p.bend_count,
                b_len=p.bend_len_mm,
                n_holes=p.punch_count,
                h_dia=p.punch_dia_mm,
                w_type="mig",
                w_len=p.weld_len_mm,
                n_spots=p.weld_spots,
                surface="Powder Coating" if p.do_powder_dual else "None",
                oh_pct=req.overhead_pct,
                pr_pct=req.profit_pct,
                kerf=req.kerf_mm,
                apply_punch=p.do_punch,
                apply_bend=p.do_bend,
                apply_weld=p.do_weld,
                apply_pc_dual=p.do_powder_dual
            )
            q_dict = quote_to_dict(quote)
            q_dict["drg_no"] = p.drg_no
            q_dict["geometry_svg"] = p.geometry_svg
            results.append({"success": True, "quote": q_dict})
            total_amount += quote.total
            total_weight += quote.weight * p.qty
        except Exception as e:
            results.append({"success": False, "error": str(e), "part": p.name})

    return {
        "items": results,
        "summary": {
            "total_amount": total_amount,
            "total_weight": total_weight,
            "item_count": len(req.parts)
        }
    }

class SaveBulkRequest(BaseModel):
    parts: list[FabPartSpec]
    customer: str = ""
    note: str = ""
    # Global settings for this batch
    overhead_pct: float = 15.0
    profit_pct: float = 20.0
    cut_method: str = "laser"
    kerf_mm: float = 2.0
    rate_pct: float = 50.0

@app.post("/api/fab/save-bulk", tags=["FAB Sheet"])
def fab_save_bulk(req: SaveBulkRequest):
    """
    Generates and saves a list of Quotes as a single batch.
    """
    db = get_db()
    quotes = []
    for p in req.parts:
        try:
            quote = gen_quote(
                name=p.name,
                mat=p.material,
                t=p.thickness_mm,
                pl=p.length_mm,
                pw=p.width_mm,
                qty=p.qty,
                slider=req.rate_pct,
                cut_m=req.cut_method,
                cut_p=p.perim_mm if p.do_cut else 0.0,
                int_c=p.int_cuts_mm if p.do_cut else 0.0,
                n_bends=p.bend_count,
                b_len=p.bend_len_mm,
                n_holes=p.punch_count,
                h_dia=p.punch_dia_mm,
                w_type="mig",
                w_len=p.weld_len_mm,
                n_spots=p.weld_spots,
                surface="Powder Coating" if p.do_powder_dual else "None",
                oh_pct=req.overhead_pct,
                pr_pct=req.profit_pct,
                kerf=req.kerf_mm,
                apply_punch=p.do_punch,
                apply_bend=p.do_bend,
                apply_weld=p.do_weld,
                apply_pc_dual=p.do_powder_dual
            )
            quotes.append(quote)
        except Exception as e:
            raise HTTPException(status_code=422, detail=f"Failed to generate quote for {p.name}: {e}")

    try:
        if hasattr(db, 'save_batch'):
            batch_id = db.save_batch(quotes, customer=req.customer, note=req.note)
            return {"success": True, "batch_id": batch_id, "items_saved": len(quotes)}
        else:
            # Fallback per instructions
            quote_ids = []
            for q in quotes:
                qid = db.save_quote(q, customer=req.customer)
                quote_ids.append(qid)
            return {"success": True, "batch_id": None, "items_saved": len(quote_ids), "note": "Saved individually (no save_batch method)"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error: {e}")

from fastapi.responses import Response

class EstimateDownloadRequest(GenerateBulkRequest):
    pass

@app.post("/api/fab/estimate/pdf", tags=["FAB Sheet"])
def fab_estimate_pdf(req: EstimateDownloadRequest):
    import io
    import jinja2
    from xhtml2pdf import pisa
    from datetime import datetime

    items_ctx = []
    total_cost = 0.0
    total_weight = 0.0

    for p in req.parts:
        try:
            quote = gen_quote(
                name=p.name,
                mat=p.material,
                t=p.thickness_mm,
                pl=p.length_mm,
                pw=p.width_mm,
                qty=p.qty,
                slider=req.rate_pct,
                cut_m=req.cut_method,
                cut_p=p.perim_mm if p.do_cut else 0.0,
                int_c=p.int_cuts_mm if p.do_cut else 0.0,
                n_bends=p.bend_count,
                b_len=p.bend_len_mm,
                n_holes=p.punch_count,
                h_dia=p.punch_dia_mm,
                w_type="mig",
                w_len=p.weld_len_mm,
                n_spots=p.weld_spots,
                surface="Powder Coating" if p.do_powder_dual else "None",
                oh_pct=req.overhead_pct,
                pr_pct=req.profit_pct,
                kerf=req.kerf_mm,
                apply_punch=p.do_punch,
                apply_bend=p.do_bend,
                apply_weld=p.do_weld,
                apply_pc_dual=p.do_powder_dual
            )
            total_cost += quote.total
            w = quote.weight * p.qty
            total_weight += w
            items_ctx.append({
                "part_name": p.name,
                "material": p.material,
                "thickness": p.thickness_mm,
                "quantity": p.qty,
                "weight": w,
                "line_total": quote.total
            })
        except:
            pass

    import os
    env = jinja2.Environment(
        loader=jinja2.FileSystemLoader(os.path.join(os.path.dirname(__file__), "templates")),
        autoescape=jinja2.select_autoescape(["html"])
    )
    def datefmt(v):
        if isinstance(v, datetime): return v.strftime("%d %b %Y")
        return str(v)
    env.filters["datefmt"] = datefmt

    template = env.get_template("estimate_pdf.html")
    html_str = template.render(
        company={"name": "Tatva Dynamics", "phone": "+91 98765 43210", "email": "info@tatvadynamics.in"},
        customer={"company_name": req.customer or "Guest", "contact_person": "N/A"},
        items=items_ctx,
        total_parts=sum(p.qty for p in req.parts),
        total_weight=total_weight,
        total_cost=total_cost,
        estimate_date=datetime.now().strftime("%d %b %Y")
    )

    buf = io.BytesIO()
    pisa.CreatePDF(io.StringIO(html_str), dest=buf, encoding="utf-8")

    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=Estimate.pdf"}
    )

@app.post("/api/fab/estimate/excel", tags=["FAB Sheet"])
def fab_estimate_excel(req: EstimateDownloadRequest):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from datetime import datetime

    items_ctx = []
    total_cost = 0.0
    total_weight = 0.0

    for p in req.parts:
        try:
            quote = gen_quote(
                name=p.name,
                mat=p.material,
                t=p.thickness_mm,
                pl=p.length_mm,
                pw=p.width_mm,
                qty=p.qty,
                slider=req.rate_pct,
                cut_m=req.cut_method,
                cut_p=p.perim_mm if p.do_cut else 0.0,
                int_c=p.int_cuts_mm if p.do_cut else 0.0,
                n_bends=p.bend_count,
                b_len=p.bend_len_mm,
                n_holes=p.punch_count,
                h_dia=p.punch_dia_mm,
                w_type="mig",
                w_len=p.weld_len_mm,
                n_spots=p.weld_spots,
                surface="Powder Coating" if p.do_powder_dual else "None",
                oh_pct=req.overhead_pct,
                pr_pct=req.profit_pct,
                kerf=req.kerf_mm,
                apply_punch=p.do_punch,
                apply_bend=p.do_bend,
                apply_weld=p.do_weld,
                apply_pc_dual=p.do_powder_dual
            )
            total_cost += quote.total
            w = quote.weight * p.qty
            total_weight += w
            
            util = 0.0
            if quote.n and quote.n.util:
                util = quote.n.util
                
            items_ctx.append({
                "name": p.name,
                "mat": p.material,
                "thk": p.thickness_mm,
                "qty": p.qty,
                "wt": w,
                "sheet": quote.n.sheet_n if quote.n else "N/A",
                "util": util,
                "cost": quote.total
            })
        except:
            pass

    wb = Workbook()
    
    # Summary Sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
    
    ws_sum.append(["APPROXIMATE COST ESTIMATE (NOT A FINAL QUOTATION)"])
    ws_sum["A1"].font = Font(bold=True, size=14, color="B91C1C")
    ws_sum.append([])
    
    data = [
        ("Customer Name", req.customer or "Guest"),
        ("Estimate Date", datetime.now().strftime("%d %b %Y")),
        ("Total Parts", sum(p.qty for p in req.parts)),
        ("Total Weight (KG)", round(total_weight, 2)),
        ("Estimated Cost Range", f"INR {round(total_cost * 0.95):,} - {round(total_cost * 1.05):,}"),
        ("Estimated Lead Time", "To be confirmed on final quote")
    ]
    for row in data:
        ws_sum.append([row[0], row[1]])
        ws_sum.cell(row=ws_sum.max_row, column=1).font = Font(bold=True)
    
    for col in ws_sum.columns:
        ws_sum.column_dimensions[col[0].column_letter].width = 25
        
    # Parts List Sheet
    ws_parts = wb.create_sheet(title="Parts List")
    headers = ["Part Name", "Material", "Thickness (MM)", "Quantity", "Weight (KG)", "Sheet Size", "Utilization %", "Estimated Cost", "Estimated Lead Time"]
    ws_parts.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws_parts.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        ws_parts.column_dimensions[cell.column_letter].width = 18

    ws_parts.column_dimensions['A'].width = 30
    ws_parts.column_dimensions['B'].width = 20

    for it in items_ctx:
        ws_parts.append([
            it["name"],
            it["mat"],
            it["thk"],
            it["qty"],
            round(it["wt"], 2),
            it["sheet"],
            f"{round(it['util'], 1)}%" if it['util'] else "N/A",
            round(it["cost"], 2),
            "TBC"
        ])

    buf = io.BytesIO()
    wb.save(buf)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Estimate.xlsx"}
    )


# ═══════════════════════════════════════════════════════════════
#  Nesting Diagram Helper (matplotlib → PNG bytes)
# ═══════════════════════════════════════════════════════════════

def _nesting_diagram_bytes(n: dict, drg_no: str = "", part_l: float = 0,
                           part_w: float = 0, part_svg_str: str = "") -> bytes:
    """
    Renders a 2D proportional nesting layout using matplotlib.
    Returns PNG bytes.
    """
    import matplotlib
    matplotlib.use("Agg")          # headless — no GUI needed
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    import io as _io

    sl = float(n.get("sl", 0))
    sw = float(n.get("sw", 0))
    pl = part_l or float(n.get("pl", 0))
    pw = part_w or float(n.get("pw", 0))
    kerf = float(n.get("kerf", 2))
    if not part_svg_str or sl <= 0 or sw <= 0 or pl <= 0 or pw <= 0:
        return b""
    try:
        import base64
        from services.dxf_svg_service import generate_nesting_diagram_base64
        b64 = generate_nesting_diagram_base64(
            sheet_l=max(sl, sw),
            sheet_w=min(sl, sw),
            part_l=pl,
            part_w=pw,
            qty=int(n.get("qty", 1) or 1),
            part_svg_str=part_svg_str,
            kerf=kerf,
        )
        return base64.b64decode(b64) if b64 else b""
    except Exception:
        pass

    # Matplotlib fallback (runs only when dxf_svg_service unavailable)
    best = int(n.get("best", 1))
    util = float(n.get("util", 0))
    waste = float(n.get("waste", 0))
    sheets = int(n.get("sheets", 1))
    orient = n.get("orient", "")
    sheet_name = n.get("name", "")

    # For tall standard sheets (e.g. 1250x2500) we orient it proportionally
    # We will draw sl on the X axis and sw on the Y axis.
    # To match standard display, typically long side is Y.
    # But sl is usually 1250, sw is 2500 for a 4x8 sheet. So X=1250, Y=2500.
    fig_w, fig_h = 2.4, 3.6
    if sl > sw:
        fig_w, fig_h = 3.6, 2.4

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    fig.patch.set_facecolor("#ffffff")
    ax.set_facecolor("#ffffff")

    # Sheet background - hatched red for waste
    sheet_rect = mpatches.Rectangle((0, 0), sl, sw, facecolor="white", edgecolor="#cc0000", linewidth=1.5, hatch="xxx")
    ax.add_patch(sheet_rect)
    ax.text(sl/2, sw/2, "WASTE", ha="center", va="center", color="#cc0000", alpha=0.4, fontsize=11, fontweight="bold", rotation=45)

    # Calculate part placements based on `orient` logic from nesting.py
    el, ew = pl + kerf, pw + kerf
    parts = [] # list of (x, y, w, h)
    
    def add_grid(startX, startY, w, h, stepX, stepY, cols, rows):
        for c in range(cols):
            for r in range(rows):
                parts.append((startX + c * stepX, startY + r * stepY, w, h))

    if orient == "Normal":
        cn, rn = int(sl // el), int(sw // ew)
        add_grid(0, 0, pl, pw, el, ew, cn, rn)
    elif orient == "Rotated 90°":
        cr, rr = int(sl // ew), int(sw // el)
        add_grid(0, 0, pw, pl, ew, el, cr, rr)
    elif orient == "Mixed":
        cn, rn = int(sl // el), int(sw // ew)
        mx1 = cn * rn
        ll = sl - cn * el
        extra1_c, extra1_r = 0, 0
        if ll >= ew and sw >= el:
            extra1_c = int(ll // ew)
            extra1_r = int(sw // el)
            mx1 += extra1_c * extra1_r
            
        cr, rr = int(sl // ew), int(sw // el)
        mx2 = cr * rr
        ll2 = sl - cr * ew
        extra2_c, extra2_r = 0, 0
        if ll2 >= el and sw >= ew:
            extra2_c = int(ll2 // el)
            extra2_r = int(sw // ew)
            mx2 += extra2_c * extra2_r
            
        if mx1 >= mx2:
            add_grid(0, 0, pl, pw, el, ew, cn, rn)
            if extra1_c > 0 and extra1_r > 0:
                add_grid(cn * el, 0, pw, pl, ew, el, extra1_c, extra1_r)
        else:
            add_grid(0, 0, pw, pl, ew, el, cr, rr)
            if extra2_c > 0 and extra2_r > 0:
                add_grid(cr * ew, 0, pl, pw, el, ew, extra2_c, extra2_r)

    # Draw parts (orange)
    for (x, y, w, h) in parts:
        p_rect = mpatches.Rectangle((x, y), w, h, facecolor="#E87722", edgecolor="white", linewidth=0.6)
        ax.add_patch(p_rect)

    ax.set_xlim(0, sl)
    ax.set_ylim(0, sw)
    ax.set_aspect('equal', adjustable='box') # Keep physical proportions

    # Styling axes
    ax.spines['top'].set_linewidth(1.5)
    ax.spines['right'].set_linewidth(1.5)
    ax.spines['bottom'].set_linewidth(1.5)
    ax.spines['left'].set_linewidth(1.5)
    ax.spines['top'].set_color('#cc0000')
    ax.spines['right'].set_color('#cc0000')
    ax.spines['bottom'].set_color('#cc0000')
    ax.spines['left'].set_color('#cc0000')
    
    ax.set_xticks([0, sl])
    ax.set_yticks([0, sw])
    ax.tick_params(axis='both', colors='#555555', labelsize=6, length=3)
    
    # Titles
    title = f"{best} pcs/sheet · Util {util:.1f}% · Waste {waste:.1f}%"
    ax.set_title(title, fontsize=7.5, fontweight="bold", pad=6, color="#1a1a2e")
    
    sub = f"{drg_no or sheet_name}\n{pl:.0f}×{pw:.0f} mm on {sl:.0f}×{sw:.0f} mm"
    ax.set_xlabel(sub, fontsize=6, color="#555555", labelpad=4)

    plt.tight_layout(pad=0.2)
    buf = _io.BytesIO()
    plt.savefig(buf, format="png", dpi=180, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════
#  Phase 5b — FAB Sheet Bulk PDF Export
# ═══════════════════════════════════════════════════════════════

class FabExportPdfRequest(BaseModel):
    items: list[dict]           # items[] from generate-bulk response
    summary: dict               # summary{} from generate-bulk response
    customer: str = ""
    parts: list[dict] = []      # original FabPartSpec list (for drg_no)

@app.post("/api/fab/export-pdf", tags=["FAB Sheet"])
def fab_export_pdf(req: FabExportPdfRequest):
    """
    Accepts the full generate-bulk response and renders a
    multi-page grouped PDF.
      Page 1 : Summary table (all materials grouped)
      Page 2+: Per-material detail pages (Landscape A4)
    """
    import io as _io
    import datetime as _dt2
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors as rc
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable, PageBreak, Image as RLImage,
        KeepTogether,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT

    # ── Colour palette ─────────────────────────────────────────
    DARK_BG    = rc.HexColor("#1a1a2e")
    ACCENT     = rc.HexColor("#E87722")
    GREY_ROW   = rc.HexColor("#f5f5f5")
    GREEN_BG   = rc.HexColor("#edf7ed")
    ORANGE_BG  = rc.HexColor("#fef0e0")
    BORDER     = rc.HexColor("#cccccc")
    MAT_HDR_BG = rc.HexColor("#2a2a4a")
    SUB_TEXT   = rc.HexColor("#555555")

    date_fmt = _dt2.date.today().strftime("%d %b %Y")
    batch_no = f"FAB-{_dt2.datetime.now().strftime('%Y%m%d-%H%M')}"

    buf = _io.BytesIO()

    # ── Page 1 uses Portrait A4; detail pages use Landscape A4 ─
    # We build Page 1 content first (portrait), then append a
    # PageBreak and switch to landscape via a custom on-page method.
    # Easiest approach: use Landscape A4 throughout (wider = better).
    PAGE_SIZE = landscape(A4)          # 297 × 210 mm  (w × h)
    doc = SimpleDocTemplate(
        buf, pagesize=PAGE_SIZE,
        leftMargin=12*mm, rightMargin=12*mm,
        topMargin=12*mm, bottomMargin=12*mm,
    )
    W = doc.width   # usable width ≈ 273 mm

    # ── Paragraph styles ───────────────────────────────────────
    st = getSampleStyleSheet()
    base = st["Normal"]

    def ps(name, **kw):
        return ParagraphStyle(name, parent=base, **kw)

    bold10   = ps("b10",  fontName="Helvetica-Bold",  fontSize=10)
    bold9    = ps("b9",   fontName="Helvetica-Bold",  fontSize=9)
    bold8    = ps("b8",   fontName="Helvetica-Bold",  fontSize=8)
    normal8  = ps("n8",   fontSize=8)
    small7   = ps("sm7",  fontSize=7,  textColor=SUB_TEXT)
    cell8    = ps("c8",   fontSize=8,  leading=10)
    cell8r   = ps("c8r",  fontSize=8,  leading=10, alignment=TA_RIGHT)
    cell8c   = ps("c8c",  fontSize=8,  leading=10, alignment=TA_CENTER)
    hdr_wh   = ps("hwh",  fontName="Helvetica-Bold", fontSize=12, textColor=rc.white)
    hdr_whr  = ps("hwr",  fontName="Helvetica-Bold", fontSize=10, textColor=rc.white, alignment=TA_RIGHT)
    hdr_sub  = ps("hsb",  fontSize=8,  textColor=rc.HexColor("#cccccc"))
    hdr_subr = ps("hsbr", fontSize=8,  textColor=rc.HexColor("#cccccc"), alignment=TA_RIGHT)
    mat_lbl  = ps("matlb",fontName="Helvetica-Bold", fontSize=10, textColor=rc.white)
    sub_row  = ps("subr", fontName="Helvetica-Bold", fontSize=8, textColor=ACCENT)
    sub_rowr = ps("subrr",fontName="Helvetica-Bold", fontSize=8, textColor=ACCENT, alignment=TA_RIGHT)

    # ── DRG lookup ─────────────────────────────────────────────
    drg_lookup: dict[str, str] = {}
    geometry_lookup: dict[str, str] = {}
    for p in req.parts:
        drg_lookup[p.get("name", "")] = p.get("drg_no", "")
        geometry_lookup[p.get("name", "")] = p.get("geometry_svg", "")

    # ── Successful quotes only ─────────────────────────────────
    quotes = [it["quote"] for it in req.items if it.get("success") and it.get("quote")]

    # ── Group by material ──────────────────────────────────────
    from collections import defaultdict
    groups: dict[str, list[dict]] = defaultdict(list)
    for q in quotes:
        groups[q["mat"]].append(q)

    el = []   # ReportLab story

    # ══════════════════════════════════════════════════════════
    #  HEADER BANNER
    # ══════════════════════════════════════════════════════════
    hdr_tbl = Table([[
        [Paragraph("TATVA DYNAMICS PVT. LTD.", hdr_wh),
         Paragraph("Sheet Metal Quotation — Bulk FAB Sheet", hdr_sub)],
        [Paragraph(batch_no, hdr_whr),
         Paragraph(date_fmt, hdr_subr)],
    ]], colWidths=[W * 0.6, W * 0.4])
    hdr_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), DARK_BG),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ("LEFTPADDING",   (0, 0), (0, -1),  12),
        ("RIGHTPADDING",  (-1, 0), (-1, -1), 12),
    ]))
    el.append(hdr_tbl)
    el.append(Spacer(1, 4 * mm))

    # ── Customer / batch info bar ──────────────────────────────
    customer_val = req.customer or "—"
    info_tbl = Table([[
        Paragraph("Customer", bold8), Paragraph(customer_val, cell8),
        Paragraph("Batch", bold8),    Paragraph(batch_no, cell8),
        Paragraph("Parts", bold8),    Paragraph(str(len(quotes)), cell8c),
        Paragraph("Date", bold8),     Paragraph(date_fmt, cell8),
    ]], colWidths=[18*mm, 50*mm, 14*mm, 44*mm, 13*mm, 16*mm, 14*mm, 30*mm])
    info_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (0, -1),  GREY_ROW),
        ("BACKGROUND",    (2, 0), (2, -1),  GREY_ROW),
        ("BACKGROUND",    (4, 0), (4, -1),  GREY_ROW),
        ("BACKGROUND",    (6, 0), (6, -1),  GREY_ROW),
        ("TEXTCOLOR",     (0, 0), (0, -1),  SUB_TEXT),
        ("TEXTCOLOR",     (2, 0), (2, -1),  SUB_TEXT),
        ("TEXTCOLOR",     (4, 0), (4, -1),  SUB_TEXT),
        ("TEXTCOLOR",     (6, 0), (6, -1),  SUB_TEXT),
        ("GRID",          (0, 0), (-1, -1), 0.3, BORDER),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    el.append(info_tbl)
    el.append(Spacer(1, 6 * mm))

    # ══════════════════════════════════════════════════════════
    #  PAGE 1 — Material-wise Summary Table
    # ══════════════════════════════════════════════════════════
    el.append(Paragraph("<b>Material-wise Summary</b>", bold9))
    el.append(Spacer(1, 2 * mm))

    # Column widths for summary (total = W)
    # #(10) | Material(40) | Thickness(22) | Parts(18) | Total Qty(22) | Total Wt(30) | Amount(35)
    sum_cw = [10*mm, 50*mm, 22*mm, 18*mm, 22*mm, 30*mm, 38*mm]

    def _p(txt, style=cell8):
        return Paragraph(str(txt), style)

    sum_hdr_row = [
        _p("#",           cell8c),
        _p("Material",    bold8),
        _p("Thickness",   bold8),
        _p("Parts",       cell8c),
        _p("Total Qty",   cell8c),
        _p("Total Wt (kg)", cell8r),
        _p("Amount (Rs.)",  cell8r),
    ]
    sum_data = [sum_hdr_row]
    grand_total = 0.0
    grand_wt    = 0.0
    row_num = 1

    for mat, qs in groups.items():
        by_t: dict[float, list[dict]] = defaultdict(list)
        for q in qs:
            by_t[float(q.get("t", 0))].append(q)
        for t_val, tqs in sorted(by_t.items()):
            part_count = len(tqs)
            total_qty  = sum(int(q.get("qty", 1))                              for q in tqs)
            total_wt   = sum(float(q.get("weight", 0)) * int(q.get("qty", 1)) for q in tqs)
            total_amt  = sum(float(q.get("total", 0))                          for q in tqs)
            grand_total += total_amt
            grand_wt    += total_wt
            sum_data.append([
                _p(str(row_num), cell8c),
                _p(mat,          cell8),
                _p(f"{t_val:g} mm", cell8c),
                _p(str(part_count), cell8c),
                _p(str(total_qty),  cell8c),
                _p(f"{total_wt:,.2f}", cell8r),
                _p(f"{total_amt:,.2f}", cell8r),
            ])
            row_num += 1

    # Grand total footer row
    sum_data.append([
        _p("GRAND TOTAL", ps("gt", fontName="Helvetica-Bold", fontSize=8, alignment=TA_RIGHT)),
        _p(""), _p(""), _p(""), _p(""),
        _p(f"{grand_wt:,.2f}",  ps("gtr", fontName="Helvetica-Bold", fontSize=8, alignment=TA_RIGHT, textColor=ACCENT)),
        _p(f"{grand_total:,.2f}", ps("gta", fontName="Helvetica-Bold", fontSize=9, alignment=TA_RIGHT, textColor=ACCENT)),
    ])

    sum_tbl = Table(sum_data, colWidths=sum_cw)
    ts_sum = [
        ("BACKGROUND",    (0, 0),  (-1, 0),  DARK_BG),
        ("TEXTCOLOR",     (0, 0),  (-1, 0),  rc.white),
        ("GRID",          (0, 0),  (-1, -1), 0.4, BORDER),
        ("VALIGN",        (0, 0),  (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0),  (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0),  (-1, -1), 4),
        ("LEFTPADDING",   (0, 0),  (-1, -1), 5),
        ("BACKGROUND",    (0, -1), (-1, -1), GREEN_BG),
        ("SPAN",          (0, -1), (4, -1)),
    ]
    for i in range(1, len(sum_data) - 1):
        if i % 2 == 0:
            ts_sum.append(("BACKGROUND", (0, i), (-1, i), GREY_ROW))
    sum_tbl.setStyle(TableStyle(ts_sum))
    el.append(sum_tbl)
    el.append(Spacer(1, 4 * mm))

    # ── Grand total highlight box ──────────────────────────────
    el.append(HRFlowable(width="100%", thickness=0.5, color=BORDER))
    el.append(Spacer(1, 3 * mm))
    tot_tbl = Table([[
        Paragraph(f"Total Parts: {len(quotes)}  |  Total Weight: {grand_wt:,.2f} kg", normal8),
        Paragraph(f"GRAND TOTAL:   Rs. {grand_total:,.2f}",
                  ps("gttxt", fontName="Helvetica-Bold", fontSize=12, textColor=ACCENT, alignment=TA_RIGHT)),
    ]], colWidths=[W * 0.45, W * 0.55])
    tot_tbl.setStyle(TableStyle([
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    el.append(tot_tbl)

    # ══════════════════════════════════════════════════════════
    #  PAGES 2+ — One page per material group (detail)
    # ══════════════════════════════════════════════════════════
    # Column layout for detail table (landscape A4, W ≈ 273 mm):
    #   DRG No | Part Name | T | L×W | Qty | Wt/pc | Tot Wt | Per Pc | Total
    #     28   |    60     | 12|  28 |  12 |  20   |  20    |  25    |  28    = 233 mm (pad rest)
    # We spread the remaining ~40 mm across Part Name and numeric columns.
    det_cw = [28*mm, 62*mm, 12*mm, 28*mm, 12*mm, 20*mm, 20*mm, 26*mm, 28*mm]
    # Verify total ≤ W  (≈ 236 mm — fits with margins)

    for mat, qs in groups.items():
        el.append(PageBreak())

        # Material banner
        mat_total = sum(float(q.get("total", 0)) for q in qs)
        mat_banner = Table([[
            Paragraph(f"Material: {mat}", mat_lbl),
            Paragraph(f"{len(qs)} parts  ·  Rs. {mat_total:,.2f}", hdr_subr),
        ]], colWidths=[W * 0.55, W * 0.45])
        mat_banner.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), MAT_HDR_BG),
            ("TOPPADDING",    (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING",   (0, 0), (0, -1),  12),
            ("RIGHTPADDING",  (-1, 0), (-1, -1), 12),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        el.append(mat_banner)
        el.append(Spacer(1, 3 * mm))

        # ── Detail table ──────────────────────────────────────
        # Header row uses bold Paragraphs — they wrap correctly
        det_hdr_row = [
            Paragraph("DRG No",      bold8),
            Paragraph("Part Name",   bold8),
            Paragraph("T\n(mm)",     ps("thdr", fontName="Helvetica-Bold", fontSize=8, alignment=TA_CENTER)),
            Paragraph("L × W\n(mm)", ps("lwhdr",fontName="Helvetica-Bold", fontSize=8, alignment=TA_CENTER)),
            Paragraph("Qty",         ps("qhdr", fontName="Helvetica-Bold", fontSize=8, alignment=TA_CENTER)),
            Paragraph("Wt/pc\n(kg)", ps("wthdr",fontName="Helvetica-Bold", fontSize=8, alignment=TA_RIGHT)),
            Paragraph("Tot Wt\n(kg)",ps("twthdr",fontName="Helvetica-Bold",fontSize=8, alignment=TA_RIGHT)),
            Paragraph("Per Pc\n(Rs.)",ps("ppchdr",fontName="Helvetica-Bold",fontSize=8,alignment=TA_RIGHT)),
            Paragraph("Total\n(Rs.)",ps("tothdr",fontName="Helvetica-Bold",fontSize=8, alignment=TA_RIGHT)),
        ]
        det_data = [det_hdr_row]

        t_subtotals: dict[float, dict] = {}
        for q in qs:
            t_v     = float(q.get("t", 0))
            name    = q.get("name", "")
            drg     = drg_lookup.get(name, q.get("drg_no", "")) or "—"
            l_v     = float(q.get("pl", 0))
            w_v     = float(q.get("pw", 0))
            qty_v   = int(q.get("qty", 1))
            wt_pc   = float(q.get("weight", 0))
            per_pc  = float(q.get("per_pc", 0))
            total_v = float(q.get("total", 0))
            wt_tot  = wt_pc * qty_v

            det_data.append([
                Paragraph(drg,                       cell8),
                Paragraph(name,                      cell8),
                Paragraph(f"{t_v:g}",                cell8c),
                Paragraph(f"{l_v:.0f}×{w_v:.0f}",   cell8c),
                Paragraph(str(qty_v),                cell8c),
                Paragraph(f"{wt_pc:.3f}",            cell8r),
                Paragraph(f"{wt_tot:.2f}",           cell8r),
                Paragraph(f"{per_pc:,.2f}",          cell8r),
                Paragraph(f"{total_v:,.2f}",         cell8r),
            ])

            if t_v not in t_subtotals:
                t_subtotals[t_v] = {"wt": 0.0, "amt": 0.0, "qty": 0}
            t_subtotals[t_v]["wt"]  += wt_tot
            t_subtotals[t_v]["amt"] += total_v
            t_subtotals[t_v]["qty"] += qty_v

        n_data_rows = len(qs)  # number of data rows (excluding header)

        # Subtotal rows per thickness
        for t_v, sub in sorted(t_subtotals.items()):
            label = f"Sub-total  {mat}  {t_v:g} mm"
            det_data.append([
                Paragraph("", cell8),
                Paragraph(label, sub_row),
                Paragraph("", cell8),
                Paragraph("", cell8),
                Paragraph(str(sub["qty"]), ps("sqty", fontName="Helvetica-Bold", fontSize=8, alignment=TA_CENTER, textColor=ACCENT)),
                Paragraph("", cell8),
                Paragraph(f"{sub['wt']:.2f}", sub_rowr),
                Paragraph("", cell8),
                Paragraph(f"{sub['amt']:,.2f}", sub_rowr),
            ])

        det_tbl = Table(det_data, colWidths=det_cw)
        n_det = len(det_data)

        ts_det = [
            # Header
            ("BACKGROUND",    (0, 0),  (-1, 0),  DARK_BG),
            ("TEXTCOLOR",     (0, 0),  (-1, 0),  rc.white),
            # Grid
            ("GRID",          (0, 0),  (-1, -1), 0.35, BORDER),
            # Alignment
            ("VALIGN",        (0, 0),  (-1, -1), "TOP"),
            # Padding — generous so content doesn't feel cramped
            ("TOPPADDING",    (0, 0),  (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0),  (-1, -1), 4),
            ("LEFTPADDING",   (0, 0),  (-1, -1), 4),
            ("RIGHTPADDING",  (0, 0),  (-1, -1), 4),
        ]
        # Alternating background for data rows
        for i in range(1, n_data_rows + 1):
            if i % 2 == 0:
                ts_det.append(("BACKGROUND", (0, i), (-1, i), GREY_ROW))
        # Subtotal rows
        for i in range(n_data_rows + 1, n_det):
            ts_det.append(("BACKGROUND", (0, i), (-1, i), ORANGE_BG))

        det_tbl.setStyle(TableStyle(ts_det))
        el.append(det_tbl)
        el.append(Spacer(1, 5 * mm))

        # ── Nesting diagrams ──────────────────────────────────
        any_nest = any(q.get("n") for q in qs)
        if any_nest:
            el.append(Paragraph("<b>Nesting Diagrams</b>", bold9))
            el.append(Spacer(1, 2 * mm))
            for q in qs:
                n_data = q.get("n")
                if not n_data:
                    continue
                name  = q.get("name", "")
                drg   = drg_lookup.get(name, q.get("drg_no", ""))
                part_svg = q.get("geometry_svg", "") or geometry_lookup.get(name, "")
                l_v   = float(q.get("pl", 0))
                w_v   = float(q.get("pw", 0))
                if not part_svg:
                    continue
                try:
                    png_bytes = _nesting_diagram_bytes(
                        n_data, drg_no=drg, part_l=l_v, part_w=w_v, part_svg_str=part_svg
                    )
                    if not png_bytes:
                        continue
                    img_buf = _io.BytesIO(png_bytes)
                    img = RLImage(img_buf, width=W * 0.68, height=W * 0.68 * 0.32)
                    from services.dxf_svg_service import calculate_nesting_metrics
                    sheet_l = max(float(n_data.get("sl", 0)), float(n_data.get("sw", 0)))
                    sheet_w = min(float(n_data.get("sl", 0)), float(n_data.get("sw", 0)))
                    m = calculate_nesting_metrics(sheet_l, sheet_w, l_v, w_v, l_v * w_v, float(n_data.get("kerf", 2)))
                    meta_tbl = Table([[
                        Paragraph(f"<b>Part / Drawing</b>: {drg or name}", cell8),
                        Paragraph(f"<b>Material</b>: {q.get('mat', '')}", cell8),
                        Paragraph(f"<b>Thickness</b>: {float(q.get('t', 0)):g} mm", cell8),
                        Paragraph(f"<b>Sheet Size</b>: {sheet_l:.0f} x {sheet_w:.0f} mm", cell8),
                    ]], colWidths=[W * 0.32, W * 0.22, W * 0.18, W * 0.28])
                    meta_tbl.setStyle(TableStyle([
                        ("LINEBELOW", (0, 0), (-1, -1), 0.5, rc.black),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]))
                    summary_tbl = Table([
                        [Paragraph("<b>NESTING SUMMARY</b>", ps("ns_hdr", fontName="Helvetica-Bold", fontSize=7, textColor=rc.white, alignment=TA_CENTER))],
                        [Paragraph(f"Parts per Sheet<br/><b>{m['pcs']} pcs</b>", cell8)],
                        [Paragraph(f"Utilization<br/><font color='green'><b>{m['util']:.1f}%</b></font>", cell8)],
                        [Paragraph(f"Waste<br/><font color='red'><b>{m['waste']:.1f}%</b></font>", cell8)],
                        [Paragraph(f"Nested Area<br/><b>{m['nested_area']:,.0f} mm2</b>", cell8)],
                        [Paragraph(f"Waste Area<br/><b>{m['waste_area']:,.0f} mm2</b>", cell8)],
                        [Paragraph(f"Total Area<br/><b>{m['total_area']:,.0f} mm2</b>", cell8)],
                    ], colWidths=[W * 0.24])
                    summary_tbl.setStyle(TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), DARK_BG),
                        ("GRID", (0, 0), (-1, -1), 0.35, BORDER),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]))
                    nest_tbl = Table([[img, summary_tbl]], colWidths=[W * 0.72, W * 0.24])
                    nest_tbl.setStyle(TableStyle([
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ]))
                    el.append(KeepTogether([
                        Paragraph("NESTING LAYOUT", bold9),
                        Spacer(1, 1 * mm),
                        meta_tbl,
                        Spacer(1, 2 * mm),
                        nest_tbl,
                        Spacer(1, 3 * mm),
                    ]))
                except Exception:
                    pass

    doc.build(el)
    buf.seek(0)
    filename = f"{batch_no}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ═══════════════════════════════════════════════════════════════
#  Phase 6 — PDF & Excel Export Routes
# ═══════════════════════════════════════════════════════════════

import io
import datetime as _dt
from fastapi.responses import StreamingResponse


def _build_pdf_bytes(quote_id: int, db) -> tuple[bytes, str]:
    """Generate PDF bytes for a quote. Returns (bytes, filename)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors as rc
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT

    row = db.conn.execute(
        "SELECT * FROM quotes WHERE id = ?", (quote_id,)
    ).fetchone()
    if not row:
        raise KeyError(f"Quote {quote_id} not found")

    meta = dict(row)
    import json as _json
    q_data = _json.loads(meta.get("quote_json", "{}"))
    lines = q_data.get("lines", [])
    n_data = q_data.get("n")

    quote_no = meta.get("quote_no", f"MQ-{quote_id}")
    part_name = meta.get("part_name", "")
    customer = meta.get("customer", "")
    mat = meta.get("material", "")
    t = meta.get("thickness", 0)
    pl = meta.get("length_mm", 0)
    pw = meta.get("width_mm", 0)
    qty = meta.get("qty", 1)
    total = meta.get("total", 0)
    per_pc = meta.get("per_piece", 0)
    sub = meta.get("subtotal", 0)
    overhead = meta.get("overhead", 0)
    overhead_pct = meta.get("overhead_pct", 15)
    profit = meta.get("profit", 0)
    profit_pct = meta.get("profit_pct", 20)
    weight = meta.get("weight_kg", 0)
    rate_kg = meta.get("rate_per_kg", 0)
    date_str = meta.get("created_at", "")[:10]
    try:
        date_fmt = _dt.date.fromisoformat(date_str).strftime("%d %b %Y")
    except Exception:
        date_fmt = date_str

    # --- Colour constants ---
    DARK_BG  = rc.HexColor("#1a1a2e")
    ACCENT   = rc.HexColor("#E87722")
    GREY_ROW = rc.HexColor("#f5f5f5")
    HEADER_TXT = rc.white
    BORDER   = rc.HexColor("#cccccc")
    GREEN_BG = rc.HexColor("#edf7ed")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=20*mm, bottomMargin=20*mm,
    )
    st = getSampleStyleSheet()
    normal = st["Normal"]
    bold_style = ParagraphStyle("bold_s", parent=normal, fontName="Helvetica-Bold", fontSize=9)
    small = ParagraphStyle("small_s", parent=normal, fontSize=8, textColor=rc.HexColor("#555555"))
    right_bold = ParagraphStyle("rbold", parent=normal, fontName="Helvetica-Bold",
                                fontSize=9, alignment=TA_RIGHT)
    accent_big = ParagraphStyle("accent_big", parent=normal, fontName="Helvetica-Bold",
                                fontSize=14, textColor=ACCENT, alignment=TA_RIGHT)
    el = []

    # ── Header ──────────────────────────────────────────────────
    # Use a white-text style for paragraphs inside the dark header cell
    hdr_left_style = ParagraphStyle("hdr_left", parent=normal,
                                    fontName="Helvetica-Bold", fontSize=11,
                                    textColor=rc.white)
    hdr_right_style = ParagraphStyle("hdr_right", parent=normal,
                                     fontName="Helvetica-Bold", fontSize=10,
                                     textColor=rc.white, alignment=TA_RIGHT)
    hdr_sub_style  = ParagraphStyle("hdr_sub", parent=normal, fontSize=8,
                                    textColor=rc.HexColor("#cccccc"))
    hdr_sub_right  = ParagraphStyle("hdr_sub_r", parent=normal, fontSize=8,
                                    textColor=rc.HexColor("#cccccc"), alignment=TA_RIGHT)

    from reportlab.platypus import KeepTogether
    hdr_data = [[
        [Paragraph("TATVA DYNAMICS PVT. LTD.", hdr_left_style),
         Paragraph("Sheet Metal Quotation", hdr_sub_style)],
        [Paragraph(quote_no, hdr_right_style),
         Paragraph(date_fmt, hdr_sub_right)],
    ]]
    hdr_tbl = Table(hdr_data, colWidths=[doc.width*0.6, doc.width*0.4])
    hdr_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), DARK_BG),
        ("TEXTCOLOR",  (0,0), (-1,-1), HEADER_TXT),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING", (0,0), (0,-1), 10),
        ("RIGHTPADDING", (-1,0), (-1,-1), 10),
    ]))
    el.append(hdr_tbl)
    el.append(Spacer(1, 4*mm))

    # ── Customer / Part info ────────────────────────────────────
    info_rows = [
        ["Customer", customer or "—", "Part Name", part_name],
        ["Material", f"{mat}  {t} mm", "Dimensions", f"{pl} × {pw} mm  ×  {qty} pcs"],
    ]
    info_tbl = Table(info_rows, colWidths=[25*mm, 65*mm, 25*mm, 65*mm])
    info_tbl.setStyle(TableStyle([
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTNAME", (2,0), (2,-1), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("TEXTCOLOR", (0,0), (0,-1), rc.HexColor("#555555")),
        ("TEXTCOLOR", (2,0), (2,-1), rc.HexColor("#555555")),
        ("GRID", (0,0), (-1,-1), 0.3, BORDER),
        ("BACKGROUND", (0,0), (0,-1), GREY_ROW),
        ("BACKGROUND", (2,0), (2,-1), GREY_ROW),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
    ]))
    el.append(info_tbl)
    el.append(Spacer(1, 4*mm))

    # ── STD RM Rate box ────────────────────────────────────────
    std = _c.STD_RATES_PER_KG
    std_total = sum(std.values())
    rate_row = [
        ["Punching", "Bending", "Welding & Fab.", "Powder Coat (Dual)", "TOTAL"],
        [f"Rs.{std.get('punching',6):.2f}/kg",
         f"Rs.{std.get('bending',3):.2f}/kg",
         f"Rs.{std.get('welding',4):.2f}/kg",
         f"Rs.{std.get('powder_coating_dual',24.80):.2f}/kg",
         f"Rs.{std_total:.2f}/kg"],
    ]
    rate_tbl = Table(rate_row, colWidths=[doc.width/5]*5)
    rate_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), DARK_BG),
        ("TEXTCOLOR",  (0,0), (-1,0), HEADER_TXT),
        ("BACKGROUND", (0,1), (-1,1), GREEN_BG),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTNAME", (-1,1), (-1,1), "Helvetica-Bold"),
        ("TEXTCOLOR", (-1,1), (-1,1), ACCENT),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("GRID", (0,0), (-1,-1), 0.4, BORDER),
        ("TOPPADDING", (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
    ]))
    el.append(Paragraph("<b>STD RM Rate Format</b>", bold_style))
    el.append(Spacer(1, 2*mm))
    el.append(rate_tbl)
    el.append(Spacer(1, 5*mm))

    # ── Line-item table ────────────────────────────────────────
    el.append(Paragraph("<b>Quote Breakdown</b>", bold_style))
    el.append(Spacer(1, 2*mm))

    line_data = [["Description", "Qty", "Unit", "Rate (Rs.)", "Amount (Rs.)"]]
    for l in lines:
        line_data.append([
            l.get("desc", ""),
            f"{float(l.get('qty',0)):g}",
            l.get("unit", ""),
            f"{float(l.get('rate',0)):,.2f}",
            f"{float(l.get('amt',0)):,.2f}",
        ])
    # Subtotals
    for lbl, val in [
        ("Subtotal",                    sub),
        (f"Overhead ({overhead_pct:.0f}%)",  overhead),
        (f"Profit ({profit_pct:.0f}%)",      profit),
        ("Rate / pc",                   per_pc),
        (f"TOTAL  ({qty} pcs)",          total),
    ]:
        line_data.append(["", "", "", lbl, f"{float(val):,.2f}"])

    n_lines = len(lines)
    col_w = [doc.width*0.42, doc.width*0.1, doc.width*0.1, doc.width*0.19, doc.width*0.19]
    line_tbl = Table(line_data, colWidths=col_w)
    ts = [
        ("BACKGROUND", (0,0), (-1,0), DARK_BG),
        ("TEXTCOLOR",  (0,0), (-1,0), HEADER_TXT),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("GRID",       (0,0), (-1,-1), 0.4, BORDER),
        ("ALIGN",      (1,0), (-1,-1), "RIGHT"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("LEFTPADDING", (0,0), (0,-1), 6),
    ]
    # Alternating rows
    for i in range(1, n_lines+1):
        if i % 2 == 0:
            ts.append(("BACKGROUND", (0, i), (-1, i), GREY_ROW))
    # Subtotal rows background
    sub_start = n_lines + 1
    ts.append(("BACKGROUND", (0, sub_start), (-1, -2), rc.HexColor("#fef9f0")))
    # Grand total row
    ts += [
        ("BACKGROUND", (0, -1), (-1, -1), GREEN_BG),
        ("FONTNAME",   (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR",  (3, -1), (-1, -1), ACCENT),
        ("FONTSIZE",   (3, -1), (-1, -1), 10),
    ]
    line_tbl.setStyle(TableStyle(ts))
    el.append(line_tbl)
    el.append(Spacer(1, 4*mm))

    # ── Nesting info ───────────────────────────────────────────
    if n_data:
        el.append(Paragraph("<b>Nesting</b>", bold_style))
        el.append(Spacer(1, 2*mm))
        nest_rows = [[
            "Sheet", "Parts/Sheet", "Orientation", "Utilization", "Sheets Needed"
        ], [
            n_data.get("name", "—"),
            str(n_data.get("best", "—")),
            n_data.get("orient", "—"),
            f"{n_data.get('util', 0):.1f}%",
            str(n_data.get("sheets", "—")),
        ]]
        n_tbl = Table(nest_rows, colWidths=[doc.width/5]*5)
        n_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), DARK_BG),
            ("TEXTCOLOR",  (0,0), (-1,0), HEADER_TXT),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("BACKGROUND", (0,1), (-1,1), GREY_ROW),
            ("FONTSIZE",   (0,0), (-1,-1), 8),
            ("ALIGN",      (0,0), (-1,-1), "CENTER"),
            ("GRID",       (0,0), (-1,-1), 0.4, BORDER),
            ("TOPPADDING", (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ]))
        el.append(n_tbl)
        el.append(Spacer(1, 3*mm))
        # ── Nesting diagram (matplotlib bar chart) ────────────
        try:
            from reportlab.platypus import Image as RLImage
            import io as _io2
            png_bytes = _nesting_diagram_bytes(
                n_data,
                drg_no=meta.get("part_name", ""),
                part_l=float(pl),
                part_w=float(pw),
                part_svg_str=q_data.get("geometry_svg", ""),
            )
            if not png_bytes:
                raise ValueError("No DXF geometry available for nesting diagram")
            img_buf = _io2.BytesIO(png_bytes)
            img = RLImage(img_buf, width=doc.width * 0.7, height=doc.width * 0.7 * 0.29)
            el.append(img)
        except Exception:
            pass  # Diagram is optional — never crash the PDF
        el.append(Spacer(1, 4*mm))



    # ── Grand Total Box ────────────────────────────────────────
    el.append(HRFlowable(width="100%", thickness=0.5, color=BORDER))
    el.append(Spacer(1, 3*mm))
    grand_data = [
        ["", f"Subtotal:      Rs. {float(sub):,.2f}"],
        ["", f"Overhead ({overhead_pct:.0f}%):   Rs. {float(overhead):,.2f}"],
        ["", f"Profit ({profit_pct:.0f}%):      Rs. {float(profit):,.2f}"],
        ["", f"GRAND TOTAL ({qty} pcs):  Rs. {float(total):,.2f}"],
    ]
    grand_tbl = Table(grand_data, colWidths=[doc.width*0.6, doc.width*0.4])
    grand_tbl.setStyle(TableStyle([
        ("FONTSIZE",  (0,0), (-1,-1), 9),
        ("ALIGN",     (1,0), (1,-1), "RIGHT"),
        ("FONTNAME",  (1,-1), (1,-1), "Helvetica-Bold"),
        ("TEXTCOLOR", (1,-1), (1,-1), ACCENT),
        ("FONTSIZE",  (1,-1), (1,-1), 13),
        ("TOPPADDING", (0,0), (-1,-1), 2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
    ]))
    el.append(grand_tbl)

    doc.build(el)
    buf.seek(0)
    filename = f"{quote_no}.pdf"
    return buf.read(), filename


@app.get("/api/quotes/{quote_id}/pdf", tags=["Export"])
def export_quote_pdf(quote_id: int):
    """
    Generate and stream a PDF for a saved quote.
    Returns Content-Disposition: attachment so the browser downloads it.
    """
    db = get_db()
    try:
        pdf_bytes, filename = _build_pdf_bytes(quote_id, db)
    except KeyError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="reportlab not installed — run: pip install reportlab"
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {exc}")

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/quotes/{quote_id}/xlsx", tags=["Export"])
def export_quote_xlsx(quote_id: int):
    """
    Generate and stream an Excel workbook for a saved quote.
    """
    import json as _json
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed — run: pip install openpyxl")

    db = get_db()
    row = db.conn.execute("SELECT * FROM quotes WHERE id = ?", (quote_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail=f"Quote {quote_id} not found")

    meta = dict(row)
    q_data = _json.loads(meta.get("quote_json", "{}"))
    lines = q_data.get("lines", [])
    n_data = q_data.get("n")

    quote_no   = meta.get("quote_no", f"MQ-{quote_id}")
    part_name  = meta.get("part_name", "")
    customer   = meta.get("customer", "")
    mat        = meta.get("material", "")
    t          = meta.get("thickness", 0)
    pl         = meta.get("length_mm", 0)
    pw         = meta.get("width_mm", 0)
    qty        = meta.get("qty", 1)
    total      = meta.get("total", 0)
    per_pc     = meta.get("per_piece", 0)
    sub        = meta.get("subtotal", 0)
    overhead   = meta.get("overhead", 0)
    overhead_pct = meta.get("overhead_pct", 15)
    profit     = meta.get("profit", 0)
    profit_pct = meta.get("profit_pct", 20)

    # Styles
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    title_font = Font(bold=True, size=14)
    accent_font = Font(bold=True, size=11, color="E87722")
    grey_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    green_fill = PatternFill(start_color="EDF7ED", end_color="EDF7ED", fill_type="solid")
    thin_bd   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22

    ws.merge_cells("A1:E1")
    ws["A1"] = "TATVA DYNAMICS PVT. LTD. — Sheet Metal Quotation"
    ws["A1"].font = title_font

    ws["A2"] = "Quote No:"; ws["B2"] = quote_no; ws["B2"].font = Font(bold=True)
    ws["C2"] = "Date:";     ws["D2"] = meta.get("created_at", "")[:10]
    ws["A3"] = "Part:";     ws["B3"] = part_name
    ws["A4"] = "Customer:"; ws["B4"] = customer or "—"
    ws["A5"] = "Material:"; ws["B5"] = f"{mat}  {t} mm"
    ws["A6"] = "Dimensions:"; ws["B6"] = f"{pl} × {pw} mm  ×  {qty} pcs"

    r = 8
    for h in ["Description", "Qty", "Unit", "Rate (₹)", "Amount (₹)"]:
        c = ws.cell(r, ["Description","Qty","Unit","Rate (₹)","Amount (₹)"].index(h)+1, h)
        c.font = hdr_font; c.fill = hdr_fill; c.border = thin_bd; c.alignment = Alignment(horizontal="center")
    for idx, l in enumerate(lines, start=1):
        r += 1
        row_data = [l.get("desc",""), l.get("qty",""), l.get("unit",""),
                    float(l.get("rate",0)), float(l.get("amt",0))]
        for ci, v in enumerate(row_data, 1):
            cell = ws.cell(r, ci, v); cell.border = thin_bd
            if ci >= 4: cell.number_format = '#,##0.00'
            if idx % 2 == 0: cell.fill = grey_fill

    r += 1
    for lbl, val in [
        (f"Subtotal", float(sub)),
        (f"Overhead ({overhead_pct:.0f}%)", float(overhead)),
        (f"Profit ({profit_pct:.0f}%)", float(profit)),
        ("Rate / pc", float(per_pc)),
        (f"GRAND TOTAL ({qty} pcs)", float(total)),
    ]:
        ws.cell(r, 4, lbl).font = Font(bold=True); ws.cell(r, 4, lbl).border = thin_bd
        vc = ws.cell(r, 5, val); vc.number_format = '#,##0.00'; vc.border = thin_bd
        if lbl.startswith("GRAND"):
            ws.cell(r, 4).fill = green_fill
            vc.fill = green_fill; vc.font = accent_font
        else:
            ws.cell(r, 4).font = Font(bold=True)
        r += 1

    # ── Sheet 2: Nesting ───────────────────────────────────────
    if n_data:
        wn = wb.create_sheet("Nesting")
        wn.column_dimensions["A"].width = 26
        wn.column_dimensions["B"].width = 18
        wn.column_dimensions["C"].width = 18
        wn.column_dimensions["D"].width = 16
        wn.column_dimensions["E"].width = 18
        for ci, h in enumerate(["Sheet", "Parts/Sheet", "Orientation", "Utilization %", "Sheets Needed"], 1):
            c = wn.cell(1, ci, h); c.font = hdr_font; c.fill = hdr_fill; c.border = thin_bd
        for ci, v in enumerate([
            n_data.get("name","—"), n_data.get("best","—"),
            n_data.get("orient","—"), n_data.get("util",0),
            n_data.get("sheets","—")
        ], 1):
            wn.cell(2, ci, v).border = thin_bd

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{quote_no}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ═══════════════════════════════════════════════════════════════
#  Entry point
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import uvicorn

    parser = argparse.ArgumentParser(description="Tatva Quote System API Server")
    parser.add_argument("--host", default="0.0.0.0",
                        help="Bind host (default: 0.0.0.0 — accepts LAN connections)")
    parser.add_argument("--port", type=int, default=8000,
                        help="Port to listen on (default: 8000)")
    parser.add_argument("--reload", action="store_true",
                        help="Enable auto-reload on file changes (dev mode)")
    args = parser.parse_args()

    uvicorn.run(
        "server:app",
        host=args.host,
        port=args.port,
        reload=args.reload,
    )
