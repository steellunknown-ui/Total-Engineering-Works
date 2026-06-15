"""Tab — FAB Sheet: multi-file upload + editable preview matching Kumar
Enterprises 'STD.FAB. DETAILS.' layout.

Uses `tksheet` (not ttk.Treeview) so individual cells can be painted red
when their value isn't auto-detected, and cells are editable in place with
a single click → Enter, just like Excel.
"""
import os
import re
import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from tksheet import Sheet

from ui.theme import C, F, FONT, _btn
from core.fab_grouper import (group_files, FabVariant, FabPart,
                               _recompute_missing, PROCESS_OPTIONS, _match_key)
from data.constants import STANDARD_SHEETS


# (key, display label, width, editable)
COLUMNS = [
    ("sr",      "SR NO",        50,  False),
    ("drg",     "DRG.NO.",      150, False),
    ("rev",     "REV",          50,  True),
    ("desc",    "DESCRIPTION",  200, True),
    ("mat",     "Mate",         75,  True),
    ("t",       "T",            45,  True),
    ("w",       "W",            55,  True),
    ("l",       "L",            55,  True),
    ("qty",     "QTY",          45,  True),
    ("weight",  "Weight (kg)",  80,  False),
    ("sheet",   "Sheet",        170, True),
    ("pps",     "Parts/Sheet",  80,  False),
    ("sheets",  "Sheets\n(per-drawing)", 95, False),
    ("process", "PROCESS",      150, True),
    ("files",   "Files",        80,  False),
    ("view",    "View",         70,  False),
    ("nest",    "Nesting",      80,  False),
]
COL_KEYS = [c[0] for c in COLUMNS]
COL_IDX = {c[0]: i for i, c in enumerate(COLUMNS)}

# Mapping column keys → FabVariant attribute name
VARIANT_ATTR = {
    "rev":     "rev",
    "desc":    "description",
    "mat":     "material",
    "t":       "thickness",
    "w":       "width",
    "l":       "length",
    "qty":     "qty",
    "process": "process",
    "sheet":   "sheet_name",
}
NUMERIC_ATTRS = {"thickness", "width", "length", "qty"}

RED_BG    = "#fde9e9"   # soft red — missing values (warm-tinted)
MASTER_BG = "#e5e0d6"   # warm grey — master/header rows


def _group_key_from(material: str | None, thickness: float | None) -> str:
    """Combined bucket key: <Material> · <Thickness mm>.

    e.g. CRCA 2.0 mm and CRCA 2.5 mm now sit in DIFFERENT buckets so
    the operator can see per-thickness sheet usage and cost. Parts
    with no thickness fall into a "<Material> · ?" bucket.
    """
    mat = _canonical_material(material)
    if thickness is None:
        return f"{mat} · ?"
    t = f"{thickness:.1f}".rstrip('0').rstrip('.')
    return f"{mat} · {t} mm"


def _canonical_material(mat: str | None) -> str:
    """Collapse the many ways an operator / drawing might spell the same
    material into ONE canonical bucket label.

        'CRCA', 'C.R.C.A.', 'CRCA Sheet', 'crca', 'C R C A'   → 'CRCA'
        'HR Sheet', 'HRS', 'HR', 'HOT ROLLED', 'HR SKIN PASS' → 'HR Sheet'
        'CR Sheet', 'CR', 'CRC', 'CR SKIN PASS', 'COLD ROLLED' → 'CR Sheet'
        'MS Sheet', 'MS', 'M.S.', 'MILD STEEL'                → 'MS Sheet'
        'GI Sheet', 'GI', 'GALVANIZED', 'GALV'                → 'GI Sheet'
        'SS-304', 'SS 304', 'STAINLESS 304'                   → 'SS-304'
        'AL', 'ALUMINIUM', 'ALUMINUM'                         → 'Aluminium'

    Anything we don't recognise is returned trimmed/upper-cased so at
    least 'crca sheet' and 'CRCA SHEET' still merge.
    """
    if not mat:
        return "(unspecified)"
    # Normalise: upper-case, strip dots, collapse whitespace.
    t = re.sub(r"\s+", " ", mat.upper().replace(".", "").strip())

    # Stainless first (so 'SS' isn't grabbed by anything else).
    if re.search(r"\bSS[\s\-]?304\b|\bSTAINLESS\b.*\b304\b", t):
        return "SS-304"
    if re.search(r"\bSS[\s\-]?316\b|\bSTAINLESS\b.*\b316\b", t):
        return "SS-316"
    if re.search(r"\bSTAINLESS\b|\bSS\b", t):
        return "Stainless Steel"

    # CRCA — Cold Rolled Close Annealed (very common in Indian shops).
    if re.search(r"\bC\s*R\s*C\s*A\b|\bCRCA\b", t) or \
       re.search(r"\bCOLD\s*ROLLED\s*(CLOSE[D]?\s*)?ANNEAL", t):
        return "CRCA"

    # Galvanised / GI.
    if re.search(r"\bGI\b|\bGALV", t) or "ZINC" in t:
        return "GI Sheet"

    # Hot-Rolled family — HR, HRS, HRC, HOT ROLLED, MS HR.
    if re.search(r"\bHR(?:S|C)?\b|\bHOT[\s\-]*ROLLED\b", t):
        return "HR Sheet"

    # Cold-Rolled family — CR, CRC, CR SHEET, COLD ROLLED.
    if re.search(r"\bCR(?:C)?\b|\bCOLD[\s\-]*ROLLED\b", t):
        return "CR Sheet"

    # Aluminium family.
    if re.search(r"\bAL\b|\bALUMIN[IU]?UM\b", t):
        return "Aluminium"

    # Mild-steel family — MS, MILD STEEL.
    if re.search(r"\bMS\b|\bMILD\s*STEEL\b", t):
        return "MS Sheet"

    # Unknown — return cleaned-upper so casing/whitespace doesn't fragment.
    return t


class FabSheetTab(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=C['bg'])
        self.app = app
        self.parts: list[FabPart] = []
        # Maps sheet row index → (variant, parent_part_name). Parent/header
        # rows map to (None, part.name).
        self._row_refs: list[tuple[FabVariant | None, str]] = []
        self._search_popup = None              # search-on-header popup
        self._search_popup_entry = None

        self._build_toolbar()
        self._build_sheet()

    # ── toolbar ──
    def _build_toolbar(self):
        # ── Toolbar card (white, padded, subtle border) ──
        # The card background lifts the toolbar visually off the dark
        # app shell so the buttons look contained and intentional.
        card = tk.Frame(self, bg='#ffffff',
                         highlightbackground=C['border'],
                         highlightthickness=1)
        card.pack(fill='x', padx=12, pady=(12, 6))

        bar = tk.Frame(card, bg='#ffffff')
        bar.pack(fill='x', padx=10, pady=8)

        # ── Group 1: Primary action (Upload) ──
        g1 = tk.Frame(bar, bg='#ffffff'); g1.pack(side='left')
        _btn(g1, "Upload Files", C['success'], 'white', self._upload_files,
             size=11, px=20, py=6, hover=C['success_hover']).pack(side='left')

        self._sep(bar).pack(side='left', padx=(12, 10))

        # ── Group 2: Row editing ──
        g2 = tk.Frame(bar, bg='#ffffff'); g2.pack(side='left')
        _btn(g2, "Add Row", C['primary'], 'white', self._add_row,
             size=10, px=14, py=5, hover=C['primary_hover']).pack(side='left', padx=2)
        _btn(g2, "Delete Row", '#dc2626', 'white', self._delete_row,
             size=10, px=14, py=5, hover='#b91c1c').pack(side='left', padx=2)
        _btn(g2, "Clear All", C['primary'], 'white', self._clear,
             size=10, px=14, py=5, hover=C['primary_hover']).pack(side='left', padx=2)

        # ── Group 3 (right-aligned): Export / Quote ──
        g3 = tk.Frame(bar, bg='#ffffff'); g3.pack(side='right')
        _btn(g3, "Quote", C['accent'], 'white', self._open_quote_preview,
             size=11, px=22, py=6, hover=C['accent_hover']).pack(side='right', padx=(8, 0))
        _btn(g3, "Export to Excel", C['btn_blue'], 'white',
             self._export_excel,
             size=11, px=18, py=6,
             hover=C['btn_blue_h']).pack(side='right', padx=2)

        # ── Bulk-edit row (own card) ──
        bulk_card = tk.Frame(self, bg='#ffffff',
                              highlightbackground=C['border'],
                              highlightthickness=1)
        bulk_card.pack(fill='x', padx=12, pady=(0, 6))
        bulk = tk.Frame(bulk_card, bg='#ffffff')
        bulk.pack(fill='x', padx=10, pady=8)

        tk.Label(bulk, text="Bulk set PROCESS for selected rows:",
                 font=(FONT['body'][0], 10, 'bold'),
                 bg='#ffffff', fg=C['text']).pack(side='left')
        self.bulk_process = tk.StringVar()
        cb = ttk.Combobox(bulk, textvariable=self.bulk_process,
                           values=PROCESS_OPTIONS, width=28,
                           state='readonly', font=FONT['input'])
        cb.pack(side='left', padx=10)
        _btn(bulk, "Apply to Selected", C['btn_blue'], 'white',
             self._bulk_apply_process,
             size=10, px=14, py=5, hover=C['btn_blue_h']).pack(side='left',
                                                                  padx=2)

        # The search input lives in a popup that opens when the user
        # clicks the DRG.NO. column header (see _on_header_click).
        self.search_var = tk.StringVar()
        self.search_var.trace_add(
            'write', lambda *_: self._apply_search())

        self.status_lbl = tk.Label(self, text="", bg=C['bg'])  # legacy ref

    def _sep(self, parent):
        """Vertical separator stripe for the toolbar."""
        return tk.Frame(parent, bg=C['border'], width=1)

    # ── sheet ──
    def _build_sheet(self):
        wrap = tk.Frame(self, bg=C['bg'])
        wrap.pack(fill='both', expand=True, padx=8, pady=(0, 10))

        headers = [c[1] for c in COLUMNS]
        self.sheet = Sheet(
            wrap,
            headers=headers,
            header_bg=C['table_header'],   # near-black charcoal
            header_fg='white',
            header_font=(F['body'], 10, 'bold'),
            font=(F['body'], 10, 'normal'),
            index_align='center',
            align='center',
            row_height=28,
            header_height=30,
            show_row_index=False,
            show_top_left=False,
            empty_horizontal=0,
            empty_vertical=0,
        )
        # Column widths
        for i, (_, _, w, _) in enumerate(COLUMNS):
            self.sheet.column_width(column=i, width=w)
        self.sheet.align_columns([COL_IDX['drg'], COL_IDX['desc']],
                                 align='w')

        self.sheet.enable_bindings((
            'single_select', 'row_select', 'column_select',
            'edit_cell', 'rc_select',
            'arrowkeys', 'copy', 'paste', 'delete',
        ))
        # Only allow editing on columns marked editable — block the rest.
        self.sheet.readonly_columns(
            columns=[i for i, c in enumerate(COLUMNS) if not c[3]],
            readonly=True)

        # PROCESS + Sheet columns use dropdowns.
        try:
            self.sheet.dropdown(column=COL_IDX["process"],
                                values=PROCESS_OPTIONS, set_value="")
            self.sheet.dropdown(column=COL_IDX["sheet"],
                                values=list(STANDARD_SHEETS.keys()),
                                set_value="")
        except Exception:
            pass

        # Commit cell edits back to the FabVariant model.
        self.sheet.extra_bindings('end_edit_cell', self._on_cell_edited)
        # Single-click in the View / Nest column → open the embedded
        # viewer for that variant. (Cropped to the drawing region for
        # multi-drawing DXFs; rendered as a per-sheet nest layout for
        # the Nest column.)
        self.sheet.extra_bindings('cell_select', self._on_cell_clicked)
        # Click on a column header → if it's DRG.NO., open the search
        # popup. We bind ONLY `column_select` here; do NOT register
        # `all_select_events` — that catchall would override the
        # cell_select binding above and break the View/Nest clicks.
        try:
            self.sheet.extra_bindings('column_select',
                                        self._on_header_click)
        except Exception:
            pass

        self.sheet.pack(fill='both', expand=True)

    # ── upload ──
    # Any of these extensions can be dropped in together; the grouper routes
    # each to the right parser.
    SUPPORTED_EXTS = (".dxf", ".dwg", ".pdf", ".step", ".stp", ".iges", ".igs",
                      ".xlsx", ".xls", ".jt")

    def _upload_files(self):
        """Tk 8.6 on macOS only reliably multi-selects when `filetypes` is
        unset — we open an unfiltered dialog and filter here."""
        paths = filedialog.askopenfilenames(
            title="Select files to upload (Cmd+click to pick many)")
        paths = [p for p in paths if p.lower().endswith(self.SUPPORTED_EXTS)]
        if not paths:
            messagebox.showinfo("No supported files",
                "Pick files ending in: " + ", ".join(self.SUPPORTED_EXTS))
            return
        self._ingest(paths)

    def _ingest(self, paths):
        # DWG files can't be parsed by ezdxf — warn up front so the operator
        # knows they'll be placeholders until converted to DXF.
        dwg_paths = [p for p in paths if p.lower().endswith(".dwg")]
        if dwg_paths:
            messagebox.showwarning(
                "DWG files need conversion",
                f"{len(dwg_paths)} DWG file(s) were selected. AutoCAD's DWG "
                "format is binary — this app can only read DXF.\n\n"
                "To extract L/W from these drawings:\n"
                "  • Open each in AutoCAD / DraftSight / QCAD → Save As DXF, or\n"
                "  • Install ODA File Converter (free) and batch-convert the "
                "folder to DXF, then re-upload.\n\n"
                "Proceeding will create placeholder rows for the DWGs — "
                "material/thickness will still come from any matching PDFs.")
        total = len(paths)
        n_pdfs = sum(1 for p in paths if p.lower().endswith(".pdf"))
        eta = f" (≈{n_pdfs * 20 // 60}m{n_pdfs * 20 % 60:02d}s for {n_pdfs} PDF OCR)" \
              if n_pdfs else ""
        self.app.status.set_message(f"Parsing {total} file(s)…{eta}")
        self.update_idletasks()

        def _progress(i, n, fp):
            name = os.path.basename(fp)
            self.app.status.set_message(f"Parsing {i}/{n} — {name}")
            self.update_idletasks()

        try:
            new_parts = group_files(list(paths), progress=_progress,
                                    db=self.app.db)
        except Exception as e:
            messagebox.showerror("Parse Error", str(e)); return

        # Merge with existing parts, filling in missing fields.
        for new_part in new_parts:
            existing = next((p for p in self.parts if p.name == new_part.name), None)
            if existing is None:
                self.parts.append(new_part); continue
            for nv in new_part.variants:
                match = next((v for v in existing.variants if v.name == nv.name), None)
                if match is None:
                    existing.variants.append(nv); continue
                for attr in ("material", "thickness", "length", "width"):
                    if getattr(match, attr) in (None, "") and getattr(nv, attr) not in (None, ""):
                        setattr(match, attr, getattr(nv, attr))
                if nv.dxf_path and not match.dxf_path: match.dxf_path = nv.dxf_path
                if nv.pdf_path and not match.pdf_path: match.pdf_path = nv.pdf_path

        self._refresh()
        total_rows = sum(len(p.variants) for p in self.parts)
        self.app.status.set_message(
            f"Loaded {total} file(s) — {total_rows} row(s) total")

    # ── rendering ──
    @staticmethod
    def _files_label(v: FabVariant) -> str:
        if v.has_dxf and v.has_pdf: return "DXF + PDF"
        if v.has_dxf: return "DXF only"
        if v.has_pdf: return "PDF only"
        return "—"

    @staticmethod
    def _sheet_dims(sheet_name: str) -> tuple[float, float] | None:
        """Return (length, width) in mm for the selected standard sheet."""
        return STANDARD_SHEETS.get(sheet_name)

    @staticmethod
    def _variant_row(v: FabVariant, sr_no) -> list:
        _recompute_missing(v)
        w_kg = v.weight_kg
        # Parts/Sheet + No. Sheets depend on the chosen sheet size.
        dims = FabSheetTab._sheet_dims(v.sheet_name)
        if dims:
            sl, sw = dims
            pps = v.parts_per_sheet(sl, sw)
            nsheets = v.sheets_needed(sl, sw)
        else:
            pps = 0; nsheets = 0
        return [
            str(sr_no) if sr_no != '' else '',
            v.name,
            v.rev,
            v.description,
            v.material or '',
            '' if v.thickness is None else str(v.thickness),
            '' if v.width is None else str(v.width),
            '' if v.length is None else str(v.length),
            str(v.qty),
            '' if w_kg is None else f'{w_kg:.3f}',
            v.sheet_name or '',
            '' if pps == 0 else str(pps),
            '' if nsheets == 0 else str(nsheets),
            v.process or '',
            FabSheetTab._files_label(v),
            'View' if v.has_dxf else '',
            'Nest' if v.length and v.width else '',
        ]

    @staticmethod
    def _master_row(part: FabPart, sr_no) -> list:
        # 17 columns now (added sheet, pps, sheets, view, nest).
        return [str(sr_no), part.name, part.rev, part.description,
                '', '', '', '', '', '', '', '', '', '', '', '', '']

    def _material_total_row(self, material: str, total_sheets: int,
                              row_count: int, total_qty: int = 0,
                              total_weight: float = 0.0) -> list:
        """A summary row showing per-group totals.
        The label spans DRG.NO. + REV + DESCRIPTION cells (which we
        merge in `_apply_highlights`), so the long
        'TOTAL CRCA · 2.0 mm — 30 part(s)' line never wraps.
        """
        label = f"  TOTAL  {material or 'UNSPECIFIED'}  —  {row_count} part(s)"
        return [
            '',                                          # sr
            label,                                       # drg (merged →)
            '',                                          # rev (merged)
            '',                                          # desc (merged end)
            '', '', '', '',                              # mat/t/w/l
            f'Total {total_qty}',                        # qty (total)
            f'{total_weight:.2f} kg' if total_weight else '',  # weight
            material or '',                              # sheet column label
            '',                                          # pps
            f'{total_sheets} sheets (per-drawing)',      # sheets
            '', '', '',                                  # process/files/view
            'Nest All' if total_qty > 0 else '',         # nest column
        ]

    def _refresh(self):
        # Flatten ALL variants and group strictly by variant.material so
        # mixed-material files (e.g. JV7094Q with MS + CRCA pieces)
        # have their pieces sorted into the right material buckets.
        # Use the CANONICAL material as the bucket key so 'CRCA',
        # 'C.R.C.A.', 'CRCA Sheet' all land in the same group; same for
        # 'HR Sheet' / 'HRS' / 'HR'.  The canonical key is also what the
        # TOTAL-row label displays.
        groups: dict[str, list[tuple]] = {}    # mat+thk → [(variant, part_name)]
        for part in self.parts:
            for v in part.variants:
                key = _group_key_from(v.material, v.thickness)
                groups.setdefault(key, []).append((v, part.name))

        # Order: specified groups (alphabetical) first, then anything
        # with "(unspecified)" material or unknown thickness ("· ?") last.
        def _group_sort_key(label: str):
            is_unknown = ("(unspecified)" in label) or label.endswith("· ?")
            return (1 if is_unknown else 0, -len(groups[label]), label)
        ordered_materials = sorted(groups, key=_group_sort_key)

        data = []
        self._row_refs = []
        sr = 0

        for material in ordered_materials:
            entries = groups[material]
            mat_total_sheets = 0
            mat_total_qty = 0
            mat_total_weight = 0.0
            for v, part_name in entries:
                sr += 1
                data.append(self._variant_row(v, sr))
                self._row_refs.append((v, part_name))
                dims = self._sheet_dims(v.sheet_name)
                if dims:
                    mat_total_sheets += v.sheets_needed(*dims)
                mat_total_qty += int(v.qty or 0)
                if v.weight_kg is not None:
                    mat_total_weight += v.weight_kg * (v.qty or 1)

            data.append(self._material_total_row(
                material, mat_total_sheets, len(entries),
                mat_total_qty, mat_total_weight))
            self._row_refs.append(("TOTAL", material))

        self.sheet.set_sheet_data(data, reset_col_positions=False,
                                  reset_row_positions=True,
                                  redraw=False)
        self._apply_highlights()
        self.sheet.refresh()

    def _refresh_material_total(self, material: str) -> None:
        """Recompute and update the TOTAL row for a single material in
        place — used after a cell edit that affected sheets-needed."""
        if not material: return
        # Find the TOTAL row for this material.
        target_row = None
        for i, ref in enumerate(self._row_refs):
            if ref[0] == "TOTAL" and ref[1] == material:
                target_row = i; break
        if target_row is None: return
        # Recompute total from all variants in this group (matched by
        # the combined material+thickness key).
        total = 0; count = 0; tot_qty = 0; tot_w = 0.0
        for ref in self._row_refs:
            v = ref[0]
            if v in (None, "TOTAL"): continue
            if _group_key_from(v.material, v.thickness) != material: continue
            count += 1
            dims = self._sheet_dims(v.sheet_name)
            if dims:
                total += v.sheets_needed(*dims)
            tot_qty += int(v.qty or 0)
            if v.weight_kg is not None:
                tot_w += v.weight_kg * (v.qty or 1)
        new_row = self._material_total_row(material, total, count,
                                            tot_qty, tot_w)
        # Update each cell in the total row.
        for col_idx, val in enumerate(new_row):
            self.sheet.set_cell_data(target_row, col_idx, val, redraw=False)
        # set_cell_data can drop the yellow row-highlight on some tksheet
        # versions — re-paint to be safe.
        try:
            self.sheet.highlight_rows(rows=[target_row],
                                       bg='#f3eee5', redraw=False)
        except Exception:
            pass
        self.sheet.refresh()

    def _set_search(self, value: str) -> None:
        """Programmatic clear-button helper."""
        self.search_var.set(value)

    def _on_header_click(self, event=None) -> None:
        """When the user clicks a column header, open a tiny search
        popup if the DRG.NO. column was the one clicked."""
        col = None
        s = getattr(event, 'selected', None)
        if s is not None:
            col = getattr(s, 'column', None)
            if col is None and isinstance(s, (tuple, list)) and len(s) >= 2:
                col = s[1]
        if col is None and hasattr(event, 'get'):
            col = event.get('column')
        if col != COL_IDX['drg']:
            return
        self._open_search_popup()

    def _open_search_popup(self) -> None:
        """Open a small floating Entry over the DRG.NO. column header.
        Closes on Escape or clicking outside. Live-filters as you type."""
        # Avoid re-opening if a popup is already up.
        if getattr(self, '_search_popup', None) is not None:
            try:
                self._search_popup.lift()
                self._search_popup_entry.focus_set()
                return
            except Exception:
                self._search_popup = None
        try:
            # Locate the DRG.NO. header on screen
            x = self.sheet.winfo_rootx()
            y = self.sheet.winfo_rooty()
            # Sum widths up to (but not including) the DRG column to get
            # the X offset of the DRG.NO. header within the sheet.
            offset_x = 0
            for i, (_, _, w, _) in enumerate(COLUMNS):
                if i >= COL_IDX['drg']:
                    break
                offset_x += w
        except Exception:
            x, y, offset_x = 100, 100, 0

        win = tk.Toplevel(self)
        win.overrideredirect(True)        # borderless popup
        win.transient(self.winfo_toplevel())
        win.configure(bg=C['accent'])
        win.geometry(f"+{x + offset_x}+{y - 2}")
        # 2-px accent border via outer frame
        inner = tk.Frame(win, bg='white')
        inner.pack(padx=2, pady=2)

        tk.Label(inner, text="Search:", bg='white', fg=C['accent'],
                 font=(FONT['body'][0], 10, 'bold')).pack(side='left',
                                                              padx=(8, 4))
        entry = tk.Entry(inner, textvariable=self.search_var,
                          width=22, font=FONT['input'],
                          relief='flat', bd=0,
                          bg='white', fg=C['text'],
                          highlightthickness=0)
        entry.pack(side='left', padx=(0, 6), pady=4, ipady=2)

        def _close(*_):
            try:
                self.search_var.set("")
            except Exception:
                pass
            try:
                win.destroy()
            except Exception:
                pass
            self._search_popup = None
            self._search_popup_entry = None

        clear_btn = tk.Button(inner, text="Close", bd=0, cursor='hand2',
                               bg='white', fg=C['text2'],
                               font=(FONT['body'][0], 9, 'bold'),
                               command=_close)
        clear_btn.pack(side='left', padx=(0, 6))

        win.bind("<Escape>", lambda e: _close())
        # Close when focus moves away (clicking the sheet, etc.)
        entry.bind("<FocusOut>", lambda e: self.after(150, lambda:
            self._maybe_close_search_popup()))

        entry.focus_set()
        self._search_popup = win
        self._search_popup_entry = entry

    def _maybe_close_search_popup(self) -> None:
        """If neither the popup window nor its Entry currently has
        focus, dismiss it."""
        win = getattr(self, '_search_popup', None)
        if win is None:
            return
        try:
            focused = win.focus_displayof()
            if focused and (focused is win or
                             str(focused).startswith(str(win))):
                return
            win.destroy()
            self._search_popup = None
            self._search_popup_entry = None
        except Exception:
            pass

    def _apply_search(self) -> None:
        """Hide rows whose DRG.NO. doesn't contain the search query.
        Group TOTAL and master-header rows stay visible so the layout
        doesn't collapse mid-table."""
        q = (self.search_var.get() or "").strip().lower()
        # tksheet exposes display_rows / hide_rows for filtering.
        if not q:
            try:
                self.sheet.display_rows("all")
            except Exception:
                pass
            return
        keep: list[int] = []
        for i, ref in enumerate(self._row_refs):
            v, part_name = ref
            if v == "TOTAL" or v is None:
                # Header/total rows always visible
                keep.append(i); continue
            if q in (v.name or "").lower() \
               or q in (v.description or "").lower() \
               or q in (part_name or "").lower():
                keep.append(i)
        try:
            self.sheet.display_rows(keep, all_displayed=False)
            self.sheet.refresh()
        except Exception:
            pass

    def _apply_highlights(self):
        """Paint per-cell and per-row highlights:
           • Master header rows → grey background
           • Material TOTAL rows → light yellow + bold (visual divider),
             with DRG.NO.+REV+DESCRIPTION cells merged so the long
             'TOTAL CRCA · 2.0 mm — 30 part(s)' label fits on ONE line.
           • Missing Material/Thickness/L/W cells → red background"""
        self.sheet.dehighlight_all()
        # Clear any previous merges so this refresh starts clean.
        # tksheet's API for "unmerge everything" varies between
        # versions — try several common forms.
        for fn_name in ("del_all_merge_cells", "unmerge_all_cells"):
            fn = getattr(self.sheet, fn_name, None)
            if fn:
                try:
                    fn(); break
                except Exception:
                    pass
        else:
            try:
                self.sheet.unmerge_cells(all=True)
            except Exception:
                pass

        TOTAL_BG = '#fef3e2'   # very light burnt-orange tint
        drg_idx  = COL_IDX['drg']
        desc_idx = COL_IDX['desc']
        for row_idx, ref in enumerate(self._row_refs):
            variant, _part_name = ref
            if variant == "TOTAL":
                # Yellow band across the row.
                self.sheet.highlight_rows(
                    rows=[row_idx], bg=TOTAL_BG, redraw=False)
                # ── Make the long TOTAL label fit on ONE line ──
                # tksheet's merge API differs across versions, so we
                # try the new Span syntax first, fall back to the old
                # merge_cells signature, and as a last resort just
                # shove the label into the wider DESCRIPTION column.
                merged = False
                try:
                    span = self.sheet.span(
                        (row_idx, drg_idx),
                        (row_idx, desc_idx))
                    span.merge()
                    merged = True
                except Exception:
                    try:
                        self.sheet.merge_cells(
                            r=row_idx, c=drg_idx,
                            rows=1, columns=3)
                        merged = True
                    except Exception:
                        merged = False
                if not merged:
                    # Fallback — move the label into DESCRIPTION (200 px,
                    # wide enough to hold "TOTAL CRCA · 1 mm — 1 part(s)")
                    # and clear DRG/REV so the row doesn't double up.
                    label = self.sheet.get_cell_data(row_idx, drg_idx)
                    self.sheet.set_cell_data(row_idx, drg_idx, '',
                                              redraw=False)
                    self.sheet.set_cell_data(row_idx, desc_idx, label,
                                              redraw=False)
                continue
            if variant is None:
                # Master header row — light grey across all columns
                self.sheet.highlight_rows(
                    rows=[row_idx], bg=MASTER_BG, redraw=False)
                continue
            # Variant row — per-cell red on each unidentified field
            for attr_key, col_key in (("material", "mat"),
                                       ("thickness", "t"),
                                       ("width", "w"),
                                       ("length", "l")):
                if getattr(variant, attr_key) in (None, ""):
                    self.sheet.highlight_cells(
                        row=row_idx, column=COL_IDX[col_key],
                        bg=RED_BG, redraw=False)

    # ── quote preview ──
    def _open_quote_preview(self):
        from tkinter import messagebox
        if not self.parts:
            messagebox.showinfo("Quote",
                "Upload files first to build a quote."); return
        try:
            from ui.quote_preview import open_quote_preview
        except Exception as e:
            messagebox.showerror("Quote", f"Couldn't load preview: {e}"); return
        open_quote_preview(self, self.parts, self.app)

    # ── group-level nesting (TOTAL row → "Nest All") ──
    def _open_group_nesting(self, group_key: str) -> None:
        """Collect every variant in this material+thickness bucket and
        open the multi-part nesting viewer."""
        variants = [v for v, _ in (
            (ref[0], ref[1]) for ref in self._row_refs
        ) if v not in (None, "TOTAL")
            and _group_key_from(v.material, v.thickness) == group_key]
        if not variants:
            messagebox.showinfo("Nesting",
                f"No parts found in '{group_key}'."); return
        try:
            from ui.nesting_viewer import open_group_nesting_viewer
        except Exception as e:
            messagebox.showerror("Nesting",
                f"Couldn't load group nesting viewer: {e}"); return
        # Use the most common sheet across variants in this group.
        sheets = [v.sheet_name for v in variants if v.sheet_name]
        sheet = max(set(sheets), key=sheets.count) if sheets else \
                "1220 × 2440  (4'×8')"
        open_group_nesting_viewer(self, group_key, variants, sheet_name=sheet)

    # ── view ──
    def _on_cell_clicked(self, event):
        """Open the DXF viewer when the user clicks a "View" cell."""
        row = col = None
        # tksheet 7.x: cell_select fires with EventDataDict containing
        # `selected` (a Span-like object with .row / .column).
        s = getattr(event, 'selected', None)
        if s is not None:
            row = getattr(s, 'row', None)
            col = getattr(s, 'column', None)
            if (row is None or col is None) and isinstance(s, (tuple, list)):
                if len(s) >= 1: row = s[0]
                if len(s) >= 2: col = s[1]
        if row is None and hasattr(event, 'get'):
            row = event.get('row')
        if col is None and hasattr(event, 'get'):
            col = event.get('column')
        if row is None or col is None:
            return
        if col not in (COL_IDX['view'], COL_IDX['nest']):
            return
        if row >= len(self._row_refs):
            return
        variant, _ = self._row_refs[row]

        # TOTAL row + Nest column → group-level nesting viewer.
        if variant == "TOTAL" and col == COL_IDX['nest']:
            group_key = self._row_refs[row][1]
            self._open_group_nesting(group_key)
            return
        if variant in (None, "TOTAL"):
            return

        # ── DXF viewer ──
        if col == COL_IDX['view']:
            if not getattr(variant, 'dxf_path', None):
                messagebox.showinfo("No DXF",
                    f"No DXF file is associated with {variant.name}.")
                return
            try:
                from ui.dxf_viewer import open_dxf_viewer
            except Exception as e:
                messagebox.showerror("Viewer error",
                    f"Couldn't load DXF viewer: {e}")
                return
            title = f"{variant.name}"
            if variant.description:
                title += f" — {variant.description}"
            open_dxf_viewer(self, variant.dxf_path,
                             bbox=getattr(variant, 'cluster_bbox', None),
                             title=title)
            return

        # ── Nesting viewer ──
        if col == COL_IDX['nest']:
            if not (variant.length and variant.width):
                messagebox.showinfo("Nesting",
                    f"{variant.name} has no L × W to nest.")
                return
            try:
                from ui.nesting_viewer import open_nesting_viewer
            except Exception as e:
                messagebox.showerror("Nesting viewer",
                    f"Couldn't load nesting viewer: {e}")
                return
            title = f"Nesting — {variant.name}"
            if variant.description:
                title += f" — {variant.description}"
            open_nesting_viewer(self, variant, title=title)
            return

    # ── editing ──
    def _on_cell_edited(self, event):
        """tksheet passes an event object with .row/.column/.value (attr or
        dict access). tksheet 7.x uses EventDataDict which supports both."""
        # Debug: log every event to /tmp/fab_edit.log so we can verify the
        # handler is being called and what keys it's receiving.
        try:
            with open('/tmp/fab_edit.log', 'a') as _f:
                _f.write(f'event_type={type(event).__name__}  repr={event!r}\n')
        except Exception:
            pass

        row = col = value = None
        # tksheet 7.x EventDataDict — supports attribute AND dict access.
        for attr_name in ('row', 'selected'):
            v = getattr(event, attr_name, None)
            if v is not None:
                if hasattr(v, 'row'):
                    row = v.row
                elif isinstance(v, (tuple, list)) and len(v) >= 1:
                    row = v[0]
                else:
                    row = v
                if row is not None: break
        col = getattr(event, 'column', None)
        if col is None and hasattr(event, 'selected'):
            s = event.selected
            if hasattr(s, 'column'):
                col = s.column
            elif isinstance(s, (tuple, list)) and len(s) >= 2:
                col = s[1]
        value = getattr(event, 'value', None)
        # Fallback to dict access
        if row is None and hasattr(event, 'get'):
            row = event.get('row')
        if col is None and hasattr(event, 'get'):
            col = event.get('column')
        if value is None and hasattr(event, 'get'):
            value = event.get('value')

        try:
            with open('/tmp/fab_edit.log', 'a') as _f:
                _f.write(f'  parsed row={row} col={col} value={value!r}\n')
        except Exception:
            pass

        if row is None or col is None or row >= len(self._row_refs):
            return
        variant, part_name = self._row_refs[row]
        # Material TOTAL rows are read-only summaries — ignore edits.
        if variant == "TOTAL":
            return
        key = COL_KEYS[col]

        # Header row — only rev/desc apply, and to the FabPart.
        if variant is None:
            part = next((p for p in self.parts if p.name == part_name), None)
            if part is None: return
            if key == "rev": part.rev = value
            elif key == "desc": part.description = value
            return

        attr = VARIANT_ATTR.get(key)
        if attr is None: return

        # Capture old group key BEFORE the edit — if the user changes
        # material OR thickness, the variant moves to a different
        # bucket and both old + new TOTAL rows have to refresh.
        old_material = _group_key_from(variant.material, variant.thickness)

        if attr in NUMERIC_ATTRS:
            try:
                if value in ("", None):
                    setattr(variant, attr, None if attr != "qty" else 1)
                else:
                    if attr == "qty":
                        variant.qty = int(float(value))
                    else:
                        setattr(variant, attr, float(value))
            except ValueError:
                messagebox.showerror("Invalid", f"{key.upper()} must be a number")
                # Put the original value back
                current = getattr(variant, attr)
                self.sheet.set_cell_data(row, col,
                    '' if current is None else str(current), redraw=True)
                return
        else:
            setattr(variant, attr, value)

        # Recompute weight + re-highlight this row.
        _recompute_missing(variant)
        w_kg = variant.weight_kg
        self.sheet.set_cell_data(
            row, COL_IDX["weight"],
            '' if w_kg is None else f'{w_kg:.3f}', redraw=False)
        # Recompute Parts/Sheet + No. Sheets.
        dims = self._sheet_dims(variant.sheet_name)
        if dims:
            sl, sw = dims
            pps = variant.parts_per_sheet(sl, sw)
            nsheets = variant.sheets_needed(sl, sw)
        else:
            pps, nsheets = 0, 0
        try:
            with open('/tmp/fab_edit.log', 'a') as _f:
                _f.write(f'  RECOMPUTE: variant.qty={variant.qty} L={variant.length} W={variant.width} '
                         f'sheet={variant.sheet_name!r} dims={dims} pps={pps} nsheets={nsheets}\n')
        except Exception: pass
        self.sheet.set_cell_data(row, COL_IDX["pps"],
            '' if pps == 0 else str(pps), redraw=False)
        self.sheet.set_cell_data(row, COL_IDX["sheets"],
            '' if nsheets == 0 else str(nsheets), redraw=False)
        # Clear any per-cell red on this row, then reapply based on current state.
        for col_key in ("mat", "t", "w", "l"):
            self.sheet.dehighlight_cells(row=row, column=COL_IDX[col_key],
                                         redraw=False)
        for attr_key, col_key in (("material", "mat"),
                                   ("thickness", "t"),
                                   ("width", "w"),
                                   ("length", "l")):
            if getattr(variant, attr_key) in (None, ""):
                self.sheet.highlight_cells(
                    row=row, column=COL_IDX[col_key], bg=RED_BG, redraw=False)

        # ── Live update of the group TOTAL row(s) ──
        # If material OR thickness changed, the variant may now belong to
        # a different bucket — easiest is a full refresh so the row ends
        # up under the right group header. Otherwise patch the totals in
        # place for stability.
        new_material = _group_key_from(variant.material, variant.thickness)
        if attr in ("material", "thickness") and new_material != old_material:
            self._refresh()
            return
        self._refresh_material_total(new_material)
        self.sheet.refresh()

    # ── add / delete / clear ──
    def _add_row(self):
        part = FabPart(name=f"NEW-{len(self.parts)+1}")
        part.variants.append(FabVariant(name=part.name))
        self.parts.append(part)
        self._refresh()

    def _selected_rows(self) -> set[int]:
        """Return the set of row indices selected in the sheet, accepting
        any of tksheet's selection modes: row-select (clicking the row
        header), range selection (drag), or a single-cell click (use that
        cell's row).
        """
        rows: set[int] = set()
        try:
            rows |= set(self.sheet.get_selected_rows())
        except Exception:
            pass
        try:
            for r, _c in self.sheet.get_selected_cells():
                rows.add(r)
        except Exception:
            pass
        try:
            curr = self.sheet.get_currently_selected()
            if curr is not None:
                r = getattr(curr, 'row', None)
                if r is None and isinstance(curr, (list, tuple)) and curr:
                    r = curr[0]
                if isinstance(r, int):
                    rows.add(r)
        except Exception:
            pass
        return rows

    def _delete_row(self):
        rows = self._selected_rows()
        if not rows:
            messagebox.showinfo("Select",
                "Click on a cell in the row you want to delete first."); return
        for row_idx in sorted(rows, reverse=True):
            if row_idx >= len(self._row_refs): continue
            variant, part_name = self._row_refs[row_idx]
            if variant == "TOTAL":
                continue   # can't delete a summary row
            if variant is None:
                self.parts = [p for p in self.parts if p.name != part_name]
            else:
                for part in self.parts:
                    if part.name == part_name and variant in part.variants:
                        part.variants.remove(variant)
                        if not part.variants:
                            self.parts.remove(part)
                        break
        self._refresh()
        self.app.status.set_message(f"Deleted {len(rows)} row(s)")

    def _clear(self):
        if not self.parts: return
        if not messagebox.askyesno("Clear", "Remove all rows?"):
            return
        self.parts.clear()
        self._refresh()
        self.app.status.set_message("FAB sheet cleared")

    # ── bulk edit ──
    def _bulk_apply_process(self):
        """Write the dropdown's PROCESS value into the PROCESS cell of
        every selected row in one click — useful for quickly overriding
        dozens of rows that all share the same process."""
        process = self.bulk_process.get().strip()
        if not process:
            messagebox.showinfo("Pick a process",
                "Choose a PROCESS from the dropdown first."); return
        rows = self._selected_rows()
        if not rows:
            messagebox.showinfo("No selection",
                "Click a cell in the target row(s) first (or drag to select)."); return
        changed = 0
        for row_idx in rows:
            if row_idx >= len(self._row_refs): continue
            variant, _ = self._row_refs[row_idx]
            if variant is None: continue  # skip master headers
            if variant == "TOTAL": continue  # skip material totals
            variant.process = process
            self.sheet.set_cell_data(row_idx, COL_IDX["process"],
                                      process, redraw=False)
            changed += 1
        self.sheet.refresh()
        self.app.status.set_message(
            f"Set PROCESS = '{process}' on {changed} row(s)")

    # ── export ──
    def _export_excel(self):
        if not self.parts:
            messagebox.showinfo("Empty", "Nothing to export — upload files first.")
            return
        default = f"FAB_Sheet_{datetime.date.today().isoformat()}.xlsx"
        fp = filedialog.asksaveasfilename(
            title="Save FAB Sheet as Excel",
            defaultextension=".xlsx",
            initialfile=default,
            filetypes=[("Excel Workbook", "*.xlsx")])
        if not fp: return
        try:
            self._write_excel(fp)
        except Exception as e:
            messagebox.showerror("Export failed", str(e)); return
        self.app.status.set_message(f"Exported: {os.path.basename(fp)}")
        messagebox.showinfo("Exported", f"Saved to:\n{fp}")

    def _write_excel(self, fp: str):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "STD.FAB. DETAILS"

        header_fill = PatternFill("solid", fgColor="1E293B")
        header_font = Font(color="FFFFFF", bold=True)
        master_fill = PatternFill("solid", fgColor="E2E8F0")
        miss_fill = PatternFill("solid", fgColor="FFE4E4")
        thin = Side(border_style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")

        # Header row
        headers = ["SR NO", "DRG.NO.", "REV", "DESCRIPTION",
                   "Mate", "T", "W", "L", "QTY", "Weight (kg)",
                   "Sheet", "Parts/Sheet", "No. Sheets",
                   "PROCESS", "Files"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = header_fill; cell.font = header_font
            cell.alignment = center; cell.border = border
        ws.row_dimensions[1].height = 22

        # Data
        r = 2; sr = 0
        for part in self.parts:
            variants = part.variants
            if len(variants) == 1:
                sr += 1
                self._write_variant_row(ws, r, sr, variants[0],
                                         miss_fill, border, center, left)
                r += 1
            else:
                sr += 1
                for c, val in enumerate([sr, part.name, part.rev,
                                          part.description, "", "", "", "",
                                          "", "", "", "", "", "", ""], 1):
                    cell = ws.cell(r, c, val)
                    cell.fill = master_fill; cell.font = Font(bold=True)
                    cell.alignment = left if c in (2, 4) else center
                    cell.border = border
                r += 1
                for v in variants:
                    self._write_variant_row(ws, r, "", v,
                                             miss_fill, border, center, left)
                    r += 1

        # Column widths
        widths = [8, 24, 8, 34, 12, 7, 10, 10, 7, 13,
                  22, 13, 13, 24, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        ws.freeze_panes = "A2"
        wb.save(fp)

    @staticmethod
    def _write_variant_row(ws, r, sr, v: FabVariant,
                           miss_fill, border, center, left):
        _recompute_missing(v)
        w_kg = v.weight_kg
        dims = FabSheetTab._sheet_dims(v.sheet_name)
        if dims:
            pps = v.parts_per_sheet(*dims)
            nsheets = v.sheets_needed(*dims)
        else:
            pps, nsheets = 0, 0
        vals = [
            sr, v.name, v.rev, v.description,
            v.material or "",
            "" if v.thickness is None else v.thickness,
            "" if v.width is None else v.width,
            "" if v.length is None else v.length,
            v.qty,
            "" if w_kg is None else w_kg,
            v.sheet_name or "",
            "" if pps == 0 else pps,
            "" if nsheets == 0 else nsheets,
            v.process or "",
            FabSheetTab._files_label(v),
        ]
        # Map attr → column to know which cells to paint red on missing.
        miss_cols = set()
        if v.material in (None, ""): miss_cols.add(5)
        if v.thickness is None: miss_cols.add(6)
        if v.width is None: miss_cols.add(7)
        if v.length is None: miss_cols.add(8)
        for c, val in enumerate(vals, 1):
            cell = ws.cell(r, c, val)
            cell.alignment = left if c in (2, 4) else center
            cell.border = border
            if c in miss_cols:
                cell.fill = miss_fill
